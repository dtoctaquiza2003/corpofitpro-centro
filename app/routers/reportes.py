from datetime import date, datetime, time, timedelta, timezone
from io import BytesIO
from typing import Dict, List, Optional, Set, Tuple

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import Date, and_, cast, exists, func, or_
from sqlalchemy.orm import Session, aliased, joinedload
from collections import defaultdict
from ..models.alerta import Alerta
from ..models.notificacion import Notificacion
from ..models.transferencia import Transferencia
from ..auth.dependencies import get_current_secretary, get_current_user
from ..dependencies.db import get_db
from ..models.consultorio import Consultorio
from ..models.gimnasio import MembresiaGimnasio
from ..models.paciente import Paciente
from ..models.pago import Pago
from ..models.sesion_terapia import SesionTerapia
from ..models.tratamiento_paciente import TratamientoPaciente
from ..models.usuario import Usuario
from ..schemas.reporte import (
    CajaSemanalDetalleOut,
    CajaSemanalPagoOut,
    ClinicaSemanalOut,
    DashboardAccionesOut,
    DeudaAcumuladaOut,
    DeudaAcumuladaPacienteOut,
    DashboardLiteOut,
    DashboardResumenOut,
    FisioDetalleOut,
    FisioDetalleDiaSueldoOut,
    FisioDetallePacienteOut,
    FisioSemanalOut,
    MetodoPagoTotalOut,
    ReporteDiaOut,
    ReporteFiltroConsultorioOut,
    ReporteFiltroTerapeutaOut,
    ReporteFiltrosOut,
    PendienteSemanaDetalleOut,
    PendienteSemanaPacienteOut,
    ReporteSemanalResponse,
    ResumenEstadoPagosOut,
    SaldoFavorDetalleOut,
    SaldoFavorPacienteOut,
    SesionPorDia,
    TerapiasReporteOut,
    TratamientoRealizadoOut,
)

router = APIRouter(prefix="/api/reportes", tags=["reportes"])


# -----------------------------------------------------------------------------
# Utilidades generales
# -----------------------------------------------------------------------------

DIAS_SEMANA = [
    "Lunes",
    "Martes",
    "Miércoles",
    "Jueves",
    "Viernes",
    "Sábado",
    "Domingo",
]

# Terapias CORPOFIT - sueldo fijo por atención:
# - Lunes a viernes: $3.50 por atención.
# - Sábado y domingo: $4.00 por atención.
# - Si de lunes a viernes el fisioterapeuta supera 15 atenciones en un día,
#   todas las atenciones de ese día se pagan a $4.00.
SUELDO_FISIO_LUNES_VIERNES = 3.50
SUELDO_FISIO_FIN_SEMANA = 4.00
SUELDO_FISIO_BONO_DIARIO = 4.00
SUELDO_FISIO_MULTIPLE_EXTREMIDAD = 5.00
# Terapias que, por carga de trabajo, pagan $5.00 al fisio por cada sesión.
TERAPIAS_SUELDO_ESPECIAL_5 = (
    "descarga completa precio especial",
    "descarga muscular completa",
)
UMBRAL_ATENCIONES_BONO_DIARIO = 15

# Porcentajes antiguos: se mantienen solo para compatibilidad con reportes
# de clínica/caja que aún usen distribución porcentual.
PORCENTAJE_FISIO_TERAPIA = 0.35
PORCENTAJE_CLINICA_TERAPIA = 0.65
PORCENTAJE_FISIO_TERAPIA_FIN_SEMANA = 0.40
PORCENTAJE_CLINICA_TERAPIA_FIN_SEMANA = 0.60
PORCENTAJE_FISIO_GIMNASIO = 0.50
PORCENTAJE_CLINICA_GIMNASIO = 0.50

# Compatibilidad con cálculos antiguos de terapias.
PORCENTAJE_FISIO = PORCENTAJE_FISIO_TERAPIA
PORCENTAJE_CLINICA = PORCENTAJE_CLINICA_TERAPIA

ECUADOR_TZ = timezone(timedelta(hours=-5))


def now_ecuador() -> datetime:
    return datetime.now(ECUADOR_TZ)


def fecha_ecuador() -> date:
    return now_ecuador().date()


def _es_fin_semana(fecha: Optional[date]) -> bool:
    """True si la fecha cae sábado o domingo."""
    return bool(fecha and fecha.weekday() in {5, 6})


def _texto_normalizado(valor: Optional[str]) -> str:
    return (valor or "").strip().lower()


def _tratamiento_tarifa_especial_5(tratamiento) -> bool:
    """True si la sesión debe pagar $5 al fisio.

    Aplica cuando el tratamiento está marcado como multi-extremidad o cuando
    el tipo de tratamiento corresponde a descargas completas trabajosas.
    """
    return _motivo_tarifa_especial_5(tratamiento) is not None


def _motivo_tarifa_especial_5(tratamiento) -> Optional[str]:
    """Motivo legible de la tarifa especial de $5.

    Se usa en el detalle del sueldo para que contabilidad pueda auditar
    por qué una sesión no se pagó con la tarifa base de $3.50/$4.00.
    """
    if not tratamiento:
        return None

    if bool(getattr(tratamiento, "multiple_extremidad", False)):
        return "Más de una extremidad"

    nombre = _texto_normalizado(getattr(tratamiento, "tipotratamiento", ""))

    if "descarga completa precio especial" in nombre:
        return "Descarga completa precio especial"

    if "descarga muscular completa" in nombre:
        return "Descarga muscular completa"

    return None


def porcentaje_fisio_terapia_por_fecha(fecha: Optional[date]) -> float:
    return (
        PORCENTAJE_FISIO_TERAPIA_FIN_SEMANA
        if _es_fin_semana(fecha)
        else PORCENTAJE_FISIO_TERAPIA
    )


def porcentaje_clinica_terapia_por_fecha(fecha: Optional[date]) -> float:
    return (
        PORCENTAJE_CLINICA_TERAPIA_FIN_SEMANA
        if _es_fin_semana(fecha)
        else PORCENTAJE_CLINICA_TERAPIA
    )


def ganancia_fisio_terapia(monto: float, fecha: Optional[date]) -> float:
    return float(monto or 0) * porcentaje_fisio_terapia_por_fecha(fecha)


def ganancia_clinica_terapia(monto: float, fecha: Optional[date]) -> float:
    return float(monto or 0) * porcentaje_clinica_terapia_por_fecha(fecha)

def sueldo_fisio_por_atencion(
    fecha: Optional[date],
    atenciones_dia: int = 0,
    multiple_extremidad: bool = False,
) -> float:
    """Valor fijo que gana el fisioterapeuta por una atención de terapia.

    Reglas:
    - Más de una extremidad o terapia especial de descarga completa: $5.00 por esa sesión.
    - Sábado y domingo: $4.00 por atención.
    - Lunes a viernes: $3.50 por atención.
    - Si lunes-viernes supera 15 atenciones en ese día, todas pagan $4.00.
    """
    if multiple_extremidad:
        return SUELDO_FISIO_MULTIPLE_EXTREMIDAD

    if _es_fin_semana(fecha):
        return SUELDO_FISIO_FIN_SEMANA

    if atenciones_dia > UMBRAL_ATENCIONES_BONO_DIARIO:
        return SUELDO_FISIO_BONO_DIARIO

    return SUELDO_FISIO_LUNES_VIERNES


def sueldo_fisio_terapia(
    fecha: Optional[date],
    atenciones_dia: int = 0,
    multiple_extremidad: bool = False,
) -> float:
    return sueldo_fisio_por_atencion(fecha, atenciones_dia, multiple_extremidad)

def rango_fechas_ecuador(desde: date, hasta: date) -> tuple[datetime, datetime]:
    """Devuelve [inicio, fin) del rango usando día calendario de Ecuador.

    Los pagos se guardan en UTC, pero los cortes de caja/reportes deben
    hacerse por fecha de Ecuador, sin depender de la zona horaria de Render
    ni de Supabase.
    """
    inicio = datetime.combine(desde, time.min).replace(tzinfo=ECUADOR_TZ)
    fin = datetime.combine(hasta + timedelta(days=1), time.min).replace(
        tzinfo=ECUADOR_TZ
    )
    return inicio, fin


def fecha_pago_ecuador_expr():
    """Fecha de caja del pago.

    Prioridad:
    1. fechapagoreal, cuando la secretaria registró una fecha contable/caja.
    2. fecha local Ecuador derivada de fechapago UTC.

    Esto evita que pagos registrados cerca de medianoche se vayan al día
    incorrecto del cuadre.
    """
    return func.coalesce(
        Pago.fechapagoreal,
        cast(func.timezone("America/Guayaquil", Pago.fechapago), Date),
    )


def filtro_fechapago_ecuador(desde: date, hasta: date):
    fecha_expr = fecha_pago_ecuador_expr()
    return and_(
        fecha_expr >= desde,
        fecha_expr <= hasta,
    )


def fin_dia_ecuador(fecha: date) -> datetime:
    return datetime.combine(fecha + timedelta(days=1), time.min).replace(
        tzinfo=ECUADOR_TZ
    )


def _validar_dia_semana(dia_semana: Optional[int]) -> Optional[int]:
    if dia_semana is None:
        return None
    if dia_semana < 0 or dia_semana > 6:
        raise HTTPException(
            status_code=400,
            detail="El día de la semana debe estar entre 0=Lunes y 6=Domingo.",
        )
    return dia_semana


def _aplicar_filtro_dia_sesion(query, dia_semana: Optional[int]):
    dia_semana = _validar_dia_semana(dia_semana)
    if dia_semana is None:
        return query
    return query.filter(func.extract("isodow", SesionTerapia.fecha) == dia_semana + 1)


def filtro_dia_pago_ecuador(dia_semana: Optional[int]):
    dia_semana = _validar_dia_semana(dia_semana)
    if dia_semana is None:
        return True
    return func.extract("isodow", fecha_pago_ecuador_expr()) == dia_semana + 1


def _generar_dias_reporte_filtrados(desde: date, hasta: date, dia_semana: Optional[int]) -> List[ReporteDiaOut]:
    dias = _generar_dias_reporte(desde, hasta)
    dia_semana = _validar_dia_semana(dia_semana)
    if dia_semana is None:
        return dias
    return [item for item in dias if item.fecha.weekday() == dia_semana]


def _nombre_usuario(usuario: Optional[Usuario]) -> str:
    if not usuario:
        return "Sin terapeuta"
    return f"{usuario.nombres} {usuario.apellidos}".strip()


def _nombre_paciente(paciente: Optional[Paciente]) -> str:
    if not paciente:
        return "Sin paciente"
    return f"{paciente.nombres} {paciente.apellidos}".strip()


def _nombre_consultorio(consultorio: Optional[Consultorio]) -> str:
    if not consultorio:
        return "Sin consultorio"
    return consultorio.nombre or "Sin consultorio"


def _precio_aplicado(tratamiento: Optional[TratamientoPaciente]) -> float:
    if not tratamiento or tratamiento.precio_sesion_aplicado is None:
        return 0.0
    return float(tratamiento.precio_sesion_aplicado)


def _es_paciente_ecuasanitas(paciente: Optional[Paciente]) -> bool:
    """Devuelve True si el paciente está cubierto por Ecuasanitas.

    Regla de negocio: Ecuasanitas SOLO aplica a terapias.
    No usar esta condición para perdonar, anular o excluir cobros de
    gimnasio mensual ni gimnasio diario; gimnasio se cobra normal.

    Usa el campo nuevo `esecuasanitas` y mantiene compatibilidad con
    pacientes antiguos que solo tenían `tiposeguro = Ecuasanitas`.
    """
    if not paciente:
        return False

    if bool(getattr(paciente, "esecuasanitas", False)):
        return True

    tipo_seguro = (getattr(paciente, "tiposeguro", None) or "").strip().lower()
    return "ecuasanitas" in tipo_seguro


def _condicion_paciente_ecuasanitas():
    return or_(
        Paciente.esecuasanitas == True,
        Paciente.tiposeguro.ilike("%ecuasanitas%"),
    )


def _columna_paciente_alerta():
    """Devuelve la columna FK del paciente en Alerta.

    En algunas versiones del modelo el atributo se llama `pacienteid`,
    y en otras `paciente_id` mapeado a la columna real `pacienteid`.
    Esto evita que el reporte falle por diferencia de nombres en el ORM.
    """
    columna = getattr(Alerta, "pacienteid", None)
    if columna is not None:
        return columna

    columna = getattr(Alerta, "paciente_id", None)
    if columna is not None:
        return columna

    raise RuntimeError("El modelo Alerta no tiene pacienteid ni paciente_id.")


def _pago_no_anulado_filter():
    return or_(Pago.anulado == False, Pago.anulado.is_(None))


def _normalizar_texto_pago(value: Optional[str]) -> str:
    texto = (value or "").strip().lower()
    traducciones = str.maketrans("áéíóúüñ", "aeiouun")
    return texto.translate(traducciones)


def _es_metodo_sin_caja(metodo: Optional[str]) -> bool:
    texto = _normalizar_texto_pago(metodo)
    return any(
        token in texto
        for token in ("cortesia", "exoner", "canje", "especie", "convenio")
    )


def _metodo_sin_caja_filter():
    metodo = func.lower(func.coalesce(Pago.metodopago, ""))
    return or_(
        metodo.like("%cortesia%"),
        metodo.like("%cortesía%"),
        metodo.like("%exoner%"),
        metodo.like("%canje%"),
        metodo.like("%especie%"),
        metodo.like("%convenio%"),
    )


def _pago_de_caja_filter():
    """Pagos que sí representan dinero cobrado dentro del sistema.

    Excluye pagos previos y también cortesías/canjes/exoneraciones, porque
    estos cubren la deuda del paciente pero no aumentan la caja real.
    Incluye recuperación de cartera porque ese dinero se cobra hoy, aunque
    no esté asociado a una sesión/tratamiento registrado.
    """
    return and_(
        or_(Pago.espagoprevio == False, Pago.espagoprevio.is_(None)),
        ~_metodo_sin_caja_filter(),
    )


def _normalizar_metodo_pago(metodo: Optional[str]) -> str:
    value = _normalizar_texto_pago(metodo)

    if "transfer" in value:
        return "transferencia"
    if "efectivo" in value:
        return "efectivo"
    if "tarjeta" in value or "card" in value:
        return "tarjeta"
    return "otros"


def _sumar_metodo_en_resumen(resumen: Dict[str, float], metodo: Optional[str], monto: float) -> None:
    categoria = _normalizar_metodo_pago(metodo)
    monto = float(monto or 0)
    if categoria == "efectivo":
        resumen["total_efectivo"] = resumen.get("total_efectivo", 0.0) + monto
    elif categoria == "transferencia":
        resumen["total_transferencia"] = resumen.get("total_transferencia", 0.0) + monto
    elif categoria == "tarjeta":
        resumen["total_tarjeta"] = resumen.get("total_tarjeta", 0.0) + monto
    else:
        resumen["total_otros_metodos"] = resumen.get("total_otros_metodos", 0.0) + monto


def _metodo_nombre_canonico(metodo: Optional[str]) -> str:
    categoria = _normalizar_metodo_pago(metodo)
    if categoria == "efectivo":
        return "Efectivo"
    if categoria == "transferencia":
        return "Transferencia"
    if categoria == "tarjeta":
        return "Tarjeta"
    return (metodo or "Otros").strip() or "Otros"


def _totales_pendientes_transferencia(
    db: Session,
    current_user: Usuario,
    desde: date,
    hasta: date,
    terapeutaid: Optional[int] = None,
    consultorioid: Optional[int] = None,
    dia_semana: Optional[int] = None,
) -> Tuple[float, int]:
    query = db.query(Pago).filter(
        filtro_fechapago_ecuador(desde, hasta),
        filtro_dia_pago_ecuador(dia_semana),
        Pago.estadopago == 1,
        _pago_no_anulado_filter(),
        _pago_de_caja_filter(),
        Pago.tratamientopacienteid != None,
        Pago.metodopago.ilike("%transfer%"),
    )

    query = _aplicar_filtros_pagos(
        query,
        current_user,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
    )

    total, cantidad = query.with_entities(
        func.coalesce(func.sum(Pago.monto), 0),
        func.count(Pago.id),
    ).one()
    return round(float(total or 0), 2), int(cantidad or 0)


def _crear_por_metodo_desde_totales(totales: Dict[str, float], extra: Optional[List[MetodoPagoTotalOut]] = None) -> List[MetodoPagoTotalOut]:
    items: List[MetodoPagoTotalOut] = []
    pares = [
        ("Efectivo", float(totales.get("total_efectivo", 0.0) or 0.0)),
        ("Transferencia", float(totales.get("total_transferencia", 0.0) or 0.0)),
        ("Tarjeta", float(totales.get("total_tarjeta", 0.0) or 0.0)),
        ("Otros", float(totales.get("total_otros_metodos", 0.0) or 0.0)),
    ]
    for nombre, total in pares:
        if round(total, 2) > 0:
            items.append(MetodoPagoTotalOut(metodo=nombre, total=round(total, 2)))
    if extra:
        items.extend(extra)
    return items


def _query_recuperacion_cartera(
    db: Session,
    current_user: Usuario,
    desde: date,
    hasta: date,
    terapeutaid: Optional[int] = None,
    consultorioid: Optional[int] = None,
    dia_semana: Optional[int] = None,
):
    """Pagos de recuperación de cartera para caja/reportes.

    No se asocian a tratamiento, sesión ni terapeuta, por eso:
    - se muestran solo como caja recuperada;
    - no se suman a productividad ni a comisión de fisioterapeuta;
    - si se filtra por terapeuta, no deben aparecer.
    """
    query = db.query(Pago).filter(
        Pago.esrecuperacioncartera == True,
        Pago.estadopago == 2,
        _pago_no_anulado_filter(),
        filtro_fechapago_ecuador(desde, hasta),
        filtro_dia_pago_ecuador(dia_semana),
    )

    if current_user.rol == 2 or terapeutaid is not None:
        return query.filter(Pago.id == -1)

    consultorio_resuelto = _resolver_consultorioid_para_rol(
        current_user,
        consultorioid,
    )

    if consultorio_resuelto is not None:
        cobrador = aliased(Usuario)
        query = query.join(
            cobrador,
            cobrador.id == Pago.creado_por_id,
        ).filter(
            cobrador.consultorioid == consultorio_resuelto,
        )

    return query



def _default_range(desde: Optional[date], hasta: Optional[date]) -> tuple[date, date]:
    today = fecha_ecuador()

    if desde is None:
        desde = today.replace(day=1)

    if hasta is None:
        hasta = today

    if hasta < desde:
        raise HTTPException(
            status_code=400,
            detail="La fecha hasta no puede ser menor que desde.",
        )

    return desde, hasta


def _validar_acceso_reportes(current_user: Usuario) -> None:
    if current_user.rol not in (1, 2, 3):
        raise HTTPException(
            status_code=403,
            detail="No autorizado",
        )


def _validar_filtros_para_rol(
    current_user: Usuario,
    terapeutaid: Optional[int] = None,
) -> None:
    _validar_acceso_reportes(current_user)

    if current_user.rol == 2 and terapeutaid is not None and terapeutaid != current_user.id:
        raise HTTPException(
            status_code=403,
            detail="Un fisioterapeuta solo puede consultar su propio reporte.",
        )


def _resolver_consultorioid_para_rol(
    current_user: Usuario,
    consultorioid: Optional[int] = None,
) -> Optional[int]:
    """
    Resuelve el consultorio permitido según el rol.

    Secretario:
        Siempre queda limitado a current_user.consultorioid.
        Si intenta enviar otro consultorioid, se bloquea.

    Terapeuta:
        Puede quedar filtrado por su propio terapeuta.
        Si tiene consultorio y manda otro consultorio, se bloquea.

    Jefe:
        Puede consultar todo o filtrar por consultorio.
    """
    _validar_acceso_reportes(current_user)

    if current_user.rol == 1:
        if current_user.consultorioid is None:
            raise HTTPException(
                status_code=403,
                detail="El secretario no tiene consultorio asignado.",
            )

        if consultorioid is not None and consultorioid != current_user.consultorioid:
            raise HTTPException(
                status_code=403,
                detail="Un secretario solo puede consultar reportes de su consultorio.",
            )

        return current_user.consultorioid

    if current_user.rol == 2:
        if (
            consultorioid is not None
            and current_user.consultorioid is not None
            and consultorioid != current_user.consultorioid
        ):
            raise HTTPException(
                status_code=403,
                detail="Un fisioterapeuta solo puede consultar su propio consultorio.",
            )

        return consultorioid

    return consultorioid


def _validar_terapeuta_para_secretario(
    current_user: Usuario,
    terapeuta: Usuario,
) -> None:
    if current_user.rol != 1:
        return

    if current_user.consultorioid is None:
        raise HTTPException(
            status_code=403,
            detail="El secretario no tiene consultorio asignado.",
        )

    if terapeuta.consultorioid != current_user.consultorioid:
        raise HTTPException(
            status_code=403,
            detail="No autorizado para consultar reportes de otro consultorio.",
        )


def _sesion_finalizada_tratamiento_en_consultorio_exists(consultorioid: int):
    """
    Pacientes compartidos:
    Un tratamiento también pertenece operativamente a un consultorio si
    ya tiene sesiones finalizadas atendidas por terapeutas de ese consultorio.
    Así Atahualpa ve la sesión de un paciente del Centro si fue atendido
    por una terapeuta de Atahualpa.
    """
    terapeuta_sesion = aliased(Usuario)

    return exists().where(
        and_(
            SesionTerapia.tratamientopacienteid == TratamientoPaciente.id,
            SesionTerapia.terapeutaid == terapeuta_sesion.id,
            terapeuta_sesion.consultorioid == consultorioid,
            SesionTerapia.horasalida != None,
        )
    )


def _tratamiento_visible_para_consultorio_filter(consultorioid: int):
    """
    Visibilidad por consultorio para tratamientos/pagos de terapia.

    Incluye:
    1. Pacientes registrados en el consultorio.
    2. Pacientes de otro consultorio que ya fueron atendidos por un
       terapeuta del consultorio actual.
    """
    return or_(
        Paciente.consultorioid == consultorioid,
        _sesion_finalizada_tratamiento_en_consultorio_exists(consultorioid),
    )


def _tratamiento_visible_para_terapeuta_filter(terapeutaid: int):
    """
    Visibilidad para terapeutas:
    - Pacientes asignados directamente.
    - Pacientes compartidos/cedidos que ya tienen una sesión atendida por él.
    """
    return or_(
        Paciente.terapeutaasignadoid == terapeutaid,
        exists().where(
            and_(
                SesionTerapia.tratamientopacienteid == TratamientoPaciente.id,
                SesionTerapia.terapeutaid == terapeutaid,
                SesionTerapia.horasalida != None,
            )
        ),
    )


def _aplicar_filtros_sesiones(
    query,
    current_user: Usuario,
    terapeutaid: Optional[int] = None,
    consultorioid: Optional[int] = None,
):
    """Aplica rol + filtros opcionales a consultas de SesionTerapia."""
    _validar_filtros_para_rol(current_user, terapeutaid)

    consultorioid = _resolver_consultorioid_para_rol(
        current_user,
        consultorioid,
    )

    if current_user.rol == 2:
        query = query.filter(SesionTerapia.terapeutaid == current_user.id)

    elif terapeutaid is not None:
        query = query.filter(SesionTerapia.terapeutaid == terapeutaid)

    if consultorioid is not None:
        # Consultorio operativo: se filtra por el consultorio del terapeuta
        # que atendió, no por el consultorio de origen del paciente.
        terapeuta_sesion = aliased(Usuario)
        query = query.join(
            terapeuta_sesion,
            terapeuta_sesion.id == SesionTerapia.terapeutaid,
        ).filter(
            terapeuta_sesion.consultorioid == consultorioid,
        )

    return query


def _aplicar_filtros_tratamientos(
    query,
    current_user: Usuario,
    terapeutaid: Optional[int] = None,
    consultorioid: Optional[int] = None,
):
    """Aplica rol + filtros opcionales a consultas de TratamientoPaciente."""
    _validar_filtros_para_rol(current_user, terapeutaid)

    consultorioid = _resolver_consultorioid_para_rol(
        current_user,
        consultorioid,
    )

    query = query.join(Paciente, Paciente.id == TratamientoPaciente.pacienteid)

    if current_user.rol == 2:
        query = query.filter(
            _tratamiento_visible_para_terapeuta_filter(current_user.id)
        )

    elif terapeutaid is not None:
        query = query.filter(
            _tratamiento_visible_para_terapeuta_filter(terapeutaid)
        )

    if consultorioid is not None:
        query = query.filter(
            _tratamiento_visible_para_consultorio_filter(consultorioid)
        )

    return query


def _aplicar_filtros_pagos(
    query,
    current_user: Usuario,
    terapeutaid: Optional[int] = None,
    consultorioid: Optional[int] = None,
):
    """Aplica rol + filtros opcionales a consultas de Pago."""
    _validar_filtros_para_rol(current_user, terapeutaid)

    consultorioid = _resolver_consultorioid_para_rol(
        current_user,
        consultorioid,
    )

    query = (
        query.join(
            TratamientoPaciente,
            TratamientoPaciente.id == Pago.tratamientopacienteid,
        )
        .join(Paciente, Paciente.id == TratamientoPaciente.pacienteid)
    )

    if current_user.rol == 2:
        query = query.filter(
            _tratamiento_visible_para_terapeuta_filter(current_user.id)
        )

    elif terapeutaid is not None:
        query = query.filter(
            _tratamiento_visible_para_terapeuta_filter(terapeutaid)
        )

    if consultorioid is not None:
        query = query.filter(
            _tratamiento_visible_para_consultorio_filter(consultorioid)
        )

    return query


def _resolver_consultorioid_gimnasio_para_rol(
    current_user: Usuario,
    consultorioid: Optional[int] = None,
) -> Optional[int]:
    """Aplica la misma seguridad por rol a los pagos de gimnasio."""
    return _resolver_consultorioid_para_rol(current_user, consultorioid)



def _gimnasio_responsable_expr():
    return func.coalesce(
        MembresiaGimnasio.responsablegimnasioid,
        Paciente.terapeutaasignadoid,
    )


def _gimnasio_consultorio_expr():
    return func.coalesce(
        MembresiaGimnasio.consultorioid,
        Paciente.consultorioid,
    )

def _aplicar_filtros_pagos_gimnasio(
    query,
    current_user: Usuario,
    terapeutaid: Optional[int] = None,
    consultorioid: Optional[int] = None,
):
    """Aplica rol + filtros a pagos de gimnasio.

    Gimnasio no tiene sesión de terapia asociada. Se asigna al responsable
    guardado en la membresía y al consultorio operativo donde se creó el
    gimnasio. Para datos antiguos sin esos campos, se mantiene el respaldo al
    terapeuta principal y al consultorio de origen del paciente.
    """
    _validar_filtros_para_rol(current_user, terapeutaid)
    consultorioid = _resolver_consultorioid_gimnasio_para_rol(
        current_user,
        consultorioid,
    )

    responsable_expr = _gimnasio_responsable_expr()
    consultorio_expr = _gimnasio_consultorio_expr()

    if current_user.rol == 2:
        query = query.filter(responsable_expr == current_user.id)
    elif terapeutaid is not None:
        query = query.filter(responsable_expr == terapeutaid)

    if consultorioid is not None:
        query = query.filter(consultorio_expr == consultorioid)

    return query


def _pagos_gimnasio_detalle_base_query(
    db: Session,
    current_user: Usuario,
    desde: date,
    hasta: date,
    terapeutaid: Optional[int] = None,
    consultorioid: Optional[int] = None,
    dia_semana: Optional[int] = None,
    solo_caja: bool = True,
):
    """Pagos de gimnasio mensual / pase diario con filtros de reporte."""
    dia_semana = _validar_dia_semana(dia_semana)

    query = (
        db.query(Pago, MembresiaGimnasio, Paciente, Usuario)
        .select_from(Pago)
        .join(MembresiaGimnasio, MembresiaGimnasio.id == Pago.membresiagimnasioid)
        .join(Paciente, Paciente.id == MembresiaGimnasio.pacienteid)
        .outerjoin(Usuario, Usuario.id == _gimnasio_responsable_expr())
        .filter(
            Pago.membresiagimnasioid != None,
            filtro_fechapago_ecuador(desde, hasta),
            filtro_dia_pago_ecuador(dia_semana),
            Pago.estadopago == 2,
            _pago_no_anulado_filter(),
        )
    )

    if solo_caja:
        query = query.filter(_pago_de_caja_filter())

    return _aplicar_filtros_pagos_gimnasio(
        query,
        current_user,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
    )


def _totales_pagos_gimnasio_caja(
    db: Session,
    current_user: Usuario,
    desde: date,
    hasta: date,
    terapeutaid: Optional[int] = None,
    consultorioid: Optional[int] = None,
    dia_semana: Optional[int] = None,
) -> Tuple[float, Dict[str, float], List[MetodoPagoTotalOut], Dict[date, Dict[str, float]]]:
    """Totales de gimnasio que sí entran a caja."""
    rows = (
        _pagos_gimnasio_detalle_base_query(
            db=db,
            current_user=current_user,
            desde=desde,
            hasta=hasta,
            terapeutaid=terapeutaid,
            consultorioid=consultorioid,
            dia_semana=dia_semana,
            solo_caja=True,
        )
        .with_entities(
            fecha_pago_ecuador_expr().label("fecha_pago"),
            Pago.metodopago,
            func.coalesce(func.sum(Pago.monto), 0),
        )
        .group_by(fecha_pago_ecuador_expr(), Pago.metodopago)
        .all()
    )

    totales_metodo = {
        "total_efectivo": 0.0,
        "total_transferencia": 0.0,
        "total_tarjeta": 0.0,
        "total_otros_metodos": 0.0,
    }
    por_metodo_map: Dict[str, float] = defaultdict(float)
    por_dia: Dict[date, Dict[str, float]] = defaultdict(lambda: {
        "total": 0.0,
        "efectivo": 0.0,
        "transferencia": 0.0,
        "tarjeta": 0.0,
        "otros": 0.0,
    })

    total_gimnasio = 0.0
    for fecha_pago, metodo, total in rows:
        total_float = float(total or 0)
        if total_float <= 0:
            continue
        total_gimnasio += total_float
        _sumar_metodo_en_resumen(totales_metodo, metodo, total_float)
        por_metodo_map[_metodo_nombre_canonico(metodo)] += total_float

        categoria = _normalizar_metodo_pago(metodo)
        por_dia[fecha_pago]["total"] += total_float
        por_dia[fecha_pago][categoria] += total_float

    por_metodo = [
        MetodoPagoTotalOut(metodo=metodo, total=round(float(total or 0), 2))
        for metodo, total in sorted(por_metodo_map.items())
        if round(float(total or 0), 2) > 0
    ]

    return round(total_gimnasio, 2), totales_metodo, por_metodo, por_dia


def _totales_pendientes_transferencia_gimnasio(
    db: Session,
    current_user: Usuario,
    desde: date,
    hasta: date,
    terapeutaid: Optional[int] = None,
    consultorioid: Optional[int] = None,
    dia_semana: Optional[int] = None,
) -> Tuple[float, int]:
    """Transferencias de gimnasio pendientes de verificación."""
    query = (
        db.query(Pago)
        .select_from(Pago)
        .join(MembresiaGimnasio, MembresiaGimnasio.id == Pago.membresiagimnasioid)
        .join(Paciente, Paciente.id == MembresiaGimnasio.pacienteid)
        .filter(
            Pago.membresiagimnasioid != None,
            filtro_fechapago_ecuador(desde, hasta),
            filtro_dia_pago_ecuador(dia_semana),
            Pago.estadopago == 1,
            _pago_no_anulado_filter(),
            _pago_de_caja_filter(),
            Pago.metodopago.ilike("%transfer%"),
        )
    )

    query = _aplicar_filtros_pagos_gimnasio(
        query,
        current_user,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
    )

    total, cantidad = query.with_entities(
        func.coalesce(func.sum(Pago.monto), 0),
        func.count(Pago.id),
    ).one()
    return round(float(total or 0), 2), int(cantidad or 0)


def _pagos_gimnasio_por_terapeuta(
    db: Session,
    current_user: Usuario,
    desde: date,
    hasta: date,
    terapeutaid: Optional[int] = None,
    consultorioid: Optional[int] = None,
    dia_semana: Optional[int] = None,
):
    """
    Pagos verificados de gimnasio mensual y pase diario agrupados por terapeuta.

    Importante: Ecuasanitas NO cubre gimnasio. Aunque el paciente sea
    Ecuasanitas, gimnasio mensual y gimnasio diario se cobran normal.

    Regla de negocio:
    - Terapia: 35% para fisioterapeuta de lunes a viernes; 40% sábado y domingo.
    - Gimnasio mensual y diario: 50% para fisioterapeuta.

    La asignación se hace por el responsable de gimnasio guardado en la
    membresía. Para datos antiguos sin responsable, se usa como respaldo el
    terapeuta principal del paciente.
    """
    _validar_filtros_para_rol(current_user, terapeutaid)
    consultorioid = _resolver_consultorioid_gimnasio_para_rol(
        current_user,
        consultorioid,
    )

    responsable_expr = _gimnasio_responsable_expr()
    consultorio_expr = _gimnasio_consultorio_expr()

    query = (
        db.query(
            responsable_expr.label("terapeutaid"),
            Usuario.nombres.label("nombres"),
            Usuario.apellidos.label("apellidos"),
            Usuario.consultorioid.label("consultorioid"),
            func.coalesce(func.sum(Pago.monto), 0).label("total_pagado"),
        )
        .select_from(Pago)
        .join(
            MembresiaGimnasio,
            MembresiaGimnasio.id == Pago.membresiagimnasioid,
        )
        .join(Paciente, Paciente.id == MembresiaGimnasio.pacienteid)
        .join(Usuario, Usuario.id == responsable_expr)
        .filter(
            Pago.membresiagimnasioid != None,
            Pago.estadopago == 2,
            _pago_no_anulado_filter(),
            _pago_de_caja_filter(),
            filtro_fechapago_ecuador(desde, hasta),
            filtro_dia_pago_ecuador(dia_semana),
            responsable_expr != None,
            Usuario.rol == 2,
            Usuario.activo == True,
        )
    )

    if current_user.rol == 2:
        query = query.filter(responsable_expr == current_user.id)
    elif terapeutaid is not None:
        query = query.filter(responsable_expr == terapeutaid)

    if consultorioid is not None:
        query = query.filter(consultorio_expr == consultorioid)

    return (
        query.group_by(
            responsable_expr,
            Usuario.nombres,
            Usuario.apellidos,
            Usuario.consultorioid,
        )
        .all()
    )


def _pagos_gimnasio_por_consultorio(
    db: Session,
    current_user: Usuario,
    desde: date,
    hasta: date,
    terapeutaid: Optional[int] = None,
    consultorioid: Optional[int] = None,
    dia_semana: Optional[int] = None,
):
    """Pagos verificados de gimnasio mensual/diario agrupados por consultorio.

    Ecuasanitas no afecta esta consulta: gimnasio mensual y diario se cobran.
    """
    _validar_filtros_para_rol(current_user, terapeutaid)
    consultorioid = _resolver_consultorioid_gimnasio_para_rol(
        current_user,
        consultorioid,
    )

    responsable_expr = _gimnasio_responsable_expr()
    consultorio_expr = _gimnasio_consultorio_expr()

    query = (
        db.query(
            consultorio_expr.label("consultorioid"),
            func.coalesce(func.sum(Pago.monto), 0).label("total_pagado"),
        )
        .select_from(Pago)
        .join(
            MembresiaGimnasio,
            MembresiaGimnasio.id == Pago.membresiagimnasioid,
        )
        .join(Paciente, Paciente.id == MembresiaGimnasio.pacienteid)
        .filter(
            Pago.membresiagimnasioid != None,
            Pago.estadopago == 2,
            _pago_no_anulado_filter(),
            _pago_de_caja_filter(),
            filtro_fechapago_ecuador(desde, hasta),
            filtro_dia_pago_ecuador(dia_semana),
        )
    )

    if current_user.rol == 2:
        query = query.filter(responsable_expr == current_user.id)
    elif terapeutaid is not None:
        query = query.filter(responsable_expr == terapeutaid)

    if consultorioid is not None:
        query = query.filter(consultorio_expr == consultorioid)

    return query.group_by(consultorio_expr).all()


def _obtener_consultorios_map(db: Session) -> Dict[Optional[int], str]:
    rows = db.query(Consultorio).all()
    return {c.id: _nombre_consultorio(c) for c in rows}


def _generar_dias_reporte(desde: date, hasta: date) -> List[ReporteDiaOut]:
    dias: List[ReporteDiaOut] = []
    actual = desde

    while actual <= hasta:
        dias.append(
            ReporteDiaOut(
                fecha=actual,
                dia=DIAS_SEMANA[actual.weekday()],
                sesiones=0,
                total_generado=0,
                pagos_verificados=0,
            )
        )
        actual = actual + timedelta(days=1)

    return dias


# -----------------------------------------------------------------------------
# Cálculos de cuentas y pagos aplicados
# -----------------------------------------------------------------------------

def _calcular_cuentas_tratamientos(
    db: Session,
    current_user: Usuario,
    tratamiento_ids: Optional[Set[int]] = None,
    terapeutaid: Optional[int] = None,
    consultorioid: Optional[int] = None,
) -> Dict[int, Dict[str, float]]:
    """
    Calcula cuentas de tratamientos usando toda la vida del tratamiento.
    """
    tratamientos_query = db.query(TratamientoPaciente).options(
        joinedload(TratamientoPaciente.paciente)
    )

    tratamientos_query = _aplicar_filtros_tratamientos(
        tratamientos_query,
        current_user,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
    )

    if tratamiento_ids is not None:
        if not tratamiento_ids:
            return {}

        tratamientos_query = tratamientos_query.filter(
            TratamientoPaciente.id.in_(tratamiento_ids)
        )

    tratamientos = tratamientos_query.all()
    ids = {t.id for t in tratamientos}

    if not ids:
        return {}

    sesiones_por_tratamiento = dict(
        db.query(
            SesionTerapia.tratamientopacienteid,
            func.count(SesionTerapia.id),
        )
        .filter(
            SesionTerapia.tratamientopacienteid.in_(ids),
            SesionTerapia.horasalida != None,
        )
        .group_by(SesionTerapia.tratamientopacienteid)
        .all()
    )

    pagos_rows = (
        db.query(
            Pago.tratamientopacienteid,
            Pago.estadopago,
            Pago.espagoprevio,
            Pago.metodopago,
            func.coalesce(func.sum(Pago.monto), 0),
        )
        .filter(
            Pago.tratamientopacienteid.in_(ids),
            _pago_no_anulado_filter(),
        )
        .group_by(
            Pago.tratamientopacienteid,
            Pago.estadopago,
            Pago.espagoprevio,
            Pago.metodopago,
        )
        .all()
    )

    pagos_map: Dict[int, Dict[str, float]] = {}

    for tratamiento_id, estado, es_previo, metodo, total in pagos_rows:
        if tratamiento_id is None:
            continue

        item = pagos_map.setdefault(
            tratamiento_id,
            {
                "pagado_caja": 0.0,
                "pago_previo": 0.0,
                "pagado_sin_caja": 0.0,
                "pendiente_verificacion": 0.0,
            },
        )

        total_float = float(total or 0)

        if estado == 2:
            if bool(es_previo):
                item["pago_previo"] += total_float
            elif _es_metodo_sin_caja(metodo):
                item["pagado_sin_caja"] += total_float
            else:
                item["pagado_caja"] += total_float
        elif estado == 1:
            item["pendiente_verificacion"] += total_float

    result: Dict[int, Dict[str, float]] = {}

    for tratamiento in tratamientos:
        sesiones = int(sesiones_por_tratamiento.get(tratamiento.id, 0) or 0)
        precio = _precio_aplicado(tratamiento)
        total_generado = sesiones * precio
        pagos_item = pagos_map.get(
            tratamiento.id,
            {
                "pagado_caja": 0.0,
                "pago_previo": 0.0,
                "pagado_sin_caja": 0.0,
                "pendiente_verificacion": 0.0,
            },
        )
        pagado_caja = float(pagos_item.get("pagado_caja", 0.0) or 0.0)
        pago_previo = float(pagos_item.get("pago_previo", 0.0) or 0.0)
        pagado_sin_caja = float(pagos_item.get("pagado_sin_caja", 0.0) or 0.0)
        pagado_verificado = pagado_caja + pago_previo + pagado_sin_caja
        pendiente_verificacion = float(
            pagos_item.get("pendiente_verificacion", 0.0) or 0.0
        )
        es_ecuasanitas = _es_paciente_ecuasanitas(tratamiento.paciente)
        cubierto_ecuasanitas = total_generado if es_ecuasanitas else 0.0

        # Ecuasanitas SOLO aplica a tratamientos/sesiones de terapia.
        # La deuda de terapia no se marca como pendiente del paciente.
        # La sesión sí genera comisión para el terapeuta:
        # 35% de lunes a viernes y 40% sábado/domingo. Gimnasio mensual/diario se cobra normal y
        # no entra en esta cuenta.
        if es_ecuasanitas:
            saldo = 0.0
            saldo_favor = max(pagado_verificado - total_generado, 0.0)
        else:
            saldo = max(total_generado - pagado_verificado, 0.0)
            saldo_favor = max(pagado_verificado - total_generado, 0.0)

        result[tratamiento.id] = {
            "precio": precio,
            "sesiones": float(sesiones),
            "total_generado": total_generado,
            "pagado_verificado": pagado_verificado,
            "pagado_caja_verificado": pagado_caja,
            "pago_previo_verificado": pago_previo,
            "pendiente_verificacion": pendiente_verificacion,
            "saldo": saldo,
            "saldo_favor": saldo_favor,
            "es_ecuasanitas": 1.0 if es_ecuasanitas else 0.0,
            "cubierto_ecuasanitas": cubierto_ecuasanitas,
        }

    return result


def _pagos_aplicados_a_rango_por_tratamiento(
    db: Session,
    tratamiento_ids: Set[int],
    desde: date,
    hasta: date,
) -> Dict[int, float]:
    """
    Calcula cuánto pago verificado está disponible para cubrir sesiones del rango.
    """
    if not tratamiento_ids:
        return {}

    tratamientos = (
        db.query(TratamientoPaciente)
        .filter(TratamientoPaciente.id.in_(tratamiento_ids))
        .all()
    )

    precios = {t.id: _precio_aplicado(t) for t in tratamientos}

    sesiones_antes = dict(
        db.query(
            SesionTerapia.tratamientopacienteid,
            func.count(SesionTerapia.id),
        )
        .filter(
            SesionTerapia.tratamientopacienteid.in_(tratamiento_ids),
            SesionTerapia.horasalida != None,
            SesionTerapia.fecha < desde,
        )
        .group_by(SesionTerapia.tratamientopacienteid)
        .all()
    )

    pagos_hasta = dict(
        db.query(
            Pago.tratamientopacienteid,
            func.coalesce(func.sum(Pago.monto), 0),
        )
        .filter(
            Pago.tratamientopacienteid.in_(tratamiento_ids),
            Pago.estadopago == 2,
            _pago_no_anulado_filter(),
            or_(
                Pago.espagoprevio == True,
                fecha_pago_ecuador_expr() <= hasta,
            ),
        )
        .group_by(Pago.tratamientopacienteid)
        .all()
    )

    disponible_para_rango: Dict[int, float] = {}

    for tratamiento_id in tratamiento_ids:
        generado_antes = float(sesiones_antes.get(tratamiento_id, 0) or 0) * precios.get(
            tratamiento_id,
            0.0,
        )
        pagado = float(pagos_hasta.get(tratamiento_id, 0) or 0)
        disponible_para_rango[tratamiento_id] = max(pagado - generado_antes, 0.0)

    return disponible_para_rango



def _cobertura_sesiones_filtradas(
    db: Session,
    sesiones_objetivo: List[SesionTerapia],
    hasta: date,
) -> Dict[int, Tuple[float, float]]:
    """Devuelve cuánto está pagado y pendiente para cada sesión visible.

    Importante: consume pagos con FIFO real desde las sesiones más antiguas del
    mismo paciente + tratamiento hasta la fecha de corte. Esto evita que, al
    filtrar por un día de la semana, una sesión aparezca pagada solo porque no se
    descontaron las sesiones anteriores de esa misma semana.

    Retorna:
        {sesion_id: (monto_aplicado, monto_pendiente)}
    """
    sesiones_validas = [
        s
        for s in sesiones_objetivo
        if s.id is not None
        and s.pacienteid is not None
        and s.tratamientopacienteid is not None
    ]

    if not sesiones_validas:
        return {}

    sesiones_objetivo_ids = {int(s.id) for s in sesiones_validas}
    paciente_ids = {int(s.pacienteid) for s in sesiones_validas}
    tratamiento_ids = {int(s.tratamientopacienteid) for s in sesiones_validas}
    claves_objetivo = {
        (int(s.pacienteid), int(s.tratamientopacienteid))
        for s in sesiones_validas
    }

    pagos_rows = (
        db.query(
            Pago.pacienteid,
            Pago.tratamientopacienteid,
            func.coalesce(func.sum(Pago.monto), 0),
        )
        .filter(
            Pago.pacienteid.in_(paciente_ids),
            Pago.tratamientopacienteid.in_(tratamiento_ids),
            Pago.estadopago == 2,
            _pago_no_anulado_filter(),
            or_(Pago.esrecuperacioncartera == False, Pago.esrecuperacioncartera.is_(None)),
            or_(
                Pago.espagoprevio == True,
                fecha_pago_ecuador_expr() <= hasta,
            ),
        )
        .group_by(Pago.pacienteid, Pago.tratamientopacienteid)
        .all()
    )

    disponible_por_clave: Dict[Tuple[int, int], float] = {
        (int(paciente_id), int(tratamiento_id)): float(total or 0.0)
        for paciente_id, tratamiento_id, total in pagos_rows
        if paciente_id is not None and tratamiento_id is not None
    }

    sesiones_historicas = (
        db.query(SesionTerapia)
        .options(
            joinedload(SesionTerapia.tratamiento_paciente),
            joinedload(SesionTerapia.terapeuta),
        )
        .filter(
            SesionTerapia.pacienteid.in_(paciente_ids),
            SesionTerapia.tratamientopacienteid.in_(tratamiento_ids),
            SesionTerapia.horasalida != None,
            SesionTerapia.fecha <= hasta,
        )
        .order_by(
            SesionTerapia.pacienteid,
            SesionTerapia.tratamientopacienteid,
            SesionTerapia.fecha,
            SesionTerapia.horaingreso,
            SesionTerapia.id,
        )
        .all()
    )

    cobertura: Dict[int, Tuple[float, float]] = {}

    for sesion in sesiones_historicas:
        if sesion.id is None or sesion.pacienteid is None or sesion.tratamientopacienteid is None:
            continue

        clave = (int(sesion.pacienteid), int(sesion.tratamientopacienteid))
        if clave not in claves_objetivo:
            continue

        # FIFO contable real:
        # Los pagos del tratamiento se consumen desde la sesión más antigua,
        # aunque esa sesión la haya hecho otro terapeuta o se haya atendido en
        # otra sede. Luego solo se devuelve cobertura para las sesiones visibles
        # del filtro actual. Esto evita marcar como pagada una sesión reciente
        # cuando los pagos ya fueron consumidos por sesiones anteriores.
        precio = float(_precio_aplicado(sesion.tratamiento_paciente) or 0.0)
        disponible = float(disponible_por_clave.get(clave, 0.0) or 0.0)
        aplicado = min(precio, disponible)
        pendiente = max(precio - aplicado, 0.0)
        disponible_por_clave[clave] = max(disponible - aplicado, 0.0)

        if int(sesion.id) in sesiones_objetivo_ids:
            cobertura[int(sesion.id)] = (round(aplicado, 2), round(pendiente, 2))

    for sesion in sesiones_validas:
        sid = int(sesion.id)
        if sid not in cobertura:
            precio = float(_precio_aplicado(sesion.tratamiento_paciente) or 0.0)
            cobertura[sid] = (0.0, round(precio, 2))

    return cobertura

def _deuda_acumulada_reporte(
    db: Session,
    current_user: Usuario,
    desde: date,
    hasta: date,
    terapeutaid: Optional[int] = None,
    consultorioid: Optional[int] = None,
    dia_semana: Optional[int] = None,
) -> DeudaAcumuladaOut:
    """Calcula cartera acumulada sin mezclarla con caja semanal.

    Regla importante:
    - La deuda acumulada NO se basa solo en pacientes que vinieron esta semana.
    - Se revisan todas las sesiones finalizadas hasta `hasta`, respetando filtros
      de fisioterapeuta, clínica operativa y día cuando correspondan.
    - Los pagos verificados del MISMO paciente + tratamiento cubren las sesiones
      desde la más antigua hasta la más nueva.
    - Las sesiones que siguen sin cubrir son deuda acumulada.

    Esto evita que el pago de otro paciente o de otro tratamiento cubra una deuda,
    y también evita ocultar deudas antiguas de pacientes que no vinieron en la semana.
    """
    dia_semana = _validar_dia_semana(dia_semana)
    consultorio_resuelto = _resolver_consultorioid_para_rol(
        current_user,
        consultorioid,
    )

    sesiones_corte_query = (
        db.query(SesionTerapia)
        .filter(
            SesionTerapia.fecha <= hasta,
            SesionTerapia.horasalida != None,
            SesionTerapia.tratamientopacienteid != None,
        )
    )

    sesiones_corte_query = _aplicar_filtro_dia_sesion(
        sesiones_corte_query,
        dia_semana,
    )

    sesiones_corte_query = _aplicar_filtros_sesiones(
        sesiones_corte_query,
        current_user,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
    )

    tratamiento_ids = {
        row[0]
        for row in sesiones_corte_query.with_entities(
            SesionTerapia.tratamientopacienteid
        ).distinct().all()
        if row[0] is not None
    }

    if not tratamiento_ids:
        return DeudaAcumuladaOut(
            desde=desde,
            hasta=hasta,
            total_deuda=0,
            total_sesiones_pendientes=0,
            pacientes=[],
        )

    pagos_rows = (
        db.query(
            Pago.pacienteid,
            Pago.tratamientopacienteid,
            func.coalesce(func.sum(Pago.monto), 0),
        )
        .filter(
            Pago.tratamientopacienteid.in_(tratamiento_ids),
            Pago.estadopago == 2,
            _pago_no_anulado_filter(),
            or_(Pago.esrecuperacioncartera == False, Pago.esrecuperacioncartera.is_(None)),
            or_(
                Pago.espagoprevio == True,
                fecha_pago_ecuador_expr() <= hasta,
            ),
        )
        .group_by(Pago.pacienteid, Pago.tratamientopacienteid)
        .all()
    )

    disponible_por_tratamiento: Dict[Tuple[int, int], float] = {
        (int(paciente_id), int(tratamiento_id)): float(total or 0)
        for paciente_id, tratamiento_id, total in pagos_rows
        if paciente_id is not None and tratamiento_id is not None
    }

    sesiones_historicas = (
        db.query(SesionTerapia)
        .options(
            joinedload(SesionTerapia.paciente),
            joinedload(SesionTerapia.terapeuta),
            joinedload(SesionTerapia.tratamiento_paciente),
        )
        .filter(
            SesionTerapia.tratamientopacienteid.in_(tratamiento_ids),
            SesionTerapia.horasalida != None,
            SesionTerapia.fecha <= hasta,
        )
        .order_by(
            SesionTerapia.tratamientopacienteid,
            SesionTerapia.pacienteid,
            SesionTerapia.fecha,
            SesionTerapia.horaingreso,
            SesionTerapia.id,
        )
        .all()
    )

    consultorios_map = _obtener_consultorios_map(db)
    agrupado: Dict[Tuple[int, Optional[int], int, float], Dict] = {}

    def incluir_en_detalle(sesion: SesionTerapia) -> bool:
        if current_user.rol == 2 and sesion.terapeutaid != current_user.id:
            return False

        if terapeutaid is not None and sesion.terapeutaid != terapeutaid:
            return False

        if consultorio_resuelto is not None:
            terapeuta = getattr(sesion, "terapeuta", None)
            if terapeuta is None or terapeuta.consultorioid != consultorio_resuelto:
                return False

        if dia_semana is not None and sesion.fecha.weekday() != dia_semana:
            return False

        return True

    # FIFO contable real:
    # No se limita por terapeuta/consultorio. Las sesiones anteriores del mismo
    # paciente + tratamiento deben consumir primero los pagos disponibles.
    # Después solo se reportan las sesiones que pertenecen al filtro actual.

    for sesion in sesiones_historicas:
        tratamiento = sesion.tratamiento_paciente

        if tratamiento is None or sesion.paciente is None:
            continue

        # Paciente Ecuasanitas también se considera cuenta por cobrar del paciente
        # si el copago/valor de la sesión no está cubierto.

        precio = _precio_aplicado(tratamiento)
        clave_pago = (int(sesion.pacienteid), int(sesion.tratamientopacienteid))
        disponible = disponible_por_tratamiento.get(clave_pago, 0.0)
        aplicado = min(precio, disponible)
        pendiente = round(max(precio - aplicado, 0.0), 2)
        disponible_por_tratamiento[clave_pago] = max(disponible - aplicado, 0.0)

        if pendiente <= 0:
            continue

        if not incluir_en_detalle(sesion):
            continue

        terapeuta = getattr(sesion, "terapeuta", None)
        terapeuta_id = sesion.terapeutaid
        consultorio_operativo = terapeuta.consultorioid if terapeuta is not None else None

        key = (
            int(sesion.pacienteid),
            terapeuta_id,
            int(sesion.tratamientopacienteid),
            round(precio, 2),
        )

        item = agrupado.setdefault(
            key,
            {
                "pacienteid": int(sesion.pacienteid),
                "paciente": _nombre_paciente(sesion.paciente),
                "terapeutaid": terapeuta_id,
                "terapeuta": _nombre_usuario(terapeuta),
                "consultorioid": consultorio_operativo,
                "consultorio": consultorios_map.get(
                    consultorio_operativo,
                    "Sin consultorio",
                ),
                "tratamientopacienteid": int(sesion.tratamientopacienteid),
                "tratamiento": tratamiento.tipotratamiento or "Tratamiento",
                "sesiones_debe": 0,
                "valor_sesion": round(precio, 2),
                "total_deuda": 0.0,
                "fechas_pendientes": [],
            },
        )

        item["sesiones_debe"] = int(item["sesiones_debe"]) + 1
        item["total_deuda"] = float(item["total_deuda"]) + pendiente
        item["fechas_pendientes"].append(sesion.fecha)

    pacientes = [
        DeudaAcumuladaPacienteOut(
            pacienteid=int(item["pacienteid"]),
            paciente=str(item["paciente"]),
            terapeutaid=item["terapeutaid"],
            terapeuta=str(item["terapeuta"]),
            consultorioid=item["consultorioid"],
            consultorio=str(item["consultorio"]),
            tratamientopacienteid=int(item["tratamientopacienteid"]),
            tratamiento=str(item["tratamiento"]),
            sesiones_debe=int(item["sesiones_debe"]),
            valor_sesion=round(float(item["valor_sesion"]), 2),
            total_deuda=round(float(item["total_deuda"]), 2),
            fechas_pendientes=sorted(item["fechas_pendientes"]),
        )
        for item in agrupado.values()
    ]

    pacientes.sort(
        key=lambda item: (
            item.paciente.lower(),
            item.terapeuta.lower(),
            item.tratamientopacienteid,
        )
    )

    return DeudaAcumuladaOut(
        desde=desde,
        hasta=hasta,
        total_deuda=round(sum(item.total_deuda for item in pacientes), 2),
        total_sesiones_pendientes=sum(item.sesiones_debe for item in pacientes),
        pacientes=pacientes,
    )


# -----------------------------------------------------------------------------
# Filtros para el frontend
# -----------------------------------------------------------------------------

@router.get("/filtros", response_model=ReporteFiltrosOut)
def obtener_filtros_reportes(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    _validar_acceso_reportes(current_user)

    consultorios_query = db.query(Consultorio).filter(Consultorio.activo == True)
    terapeutas_query = db.query(Usuario).filter(
        Usuario.activo == True,
        Usuario.rol == 2,
    )

    if current_user.rol == 1:
        consultorioid = _resolver_consultorioid_para_rol(current_user)

        consultorios_query = consultorios_query.filter(
            Consultorio.id == consultorioid
        )

        terapeutas_query = terapeutas_query.filter(
            Usuario.consultorioid == consultorioid
        )

    elif current_user.rol == 2:
        terapeutas_query = terapeutas_query.filter(
            Usuario.id == current_user.id
        )

        if current_user.consultorioid is not None:
            consultorios_query = consultorios_query.filter(
                Consultorio.id == current_user.consultorioid
            )

    consultorios = [
        ReporteFiltroConsultorioOut(
            id=c.id,
            nombre=_nombre_consultorio(c),
        )
        for c in consultorios_query.order_by(Consultorio.nombre).all()
    ]

    terapeutas = [
        ReporteFiltroTerapeutaOut(
            id=t.id,
            nombre=_nombre_usuario(t),
            consultorioid=t.consultorioid,
        )
        for t in terapeutas_query.order_by(
            Usuario.apellidos,
            Usuario.nombres,
        ).all()
    ]

    return ReporteFiltrosOut(
        terapeutas=terapeutas,
        consultorios=consultorios,
    )


# -----------------------------------------------------------------------------
# Dashboard
# -----------------------------------------------------------------------------

@router.get("/dashboard-acciones", response_model=DashboardAccionesOut)
def dashboard_acciones(
    consultorioid: Optional[int] = Query(None),
    terapeutaid: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    """
    Endpoint liviano para el panel principal.
    Versión optimizada: consolida las 11 queries originales en 5
    usando COUNT con CASE condicional por tabla.
    """
    from sqlalchemy import case as sa_case

    _validar_filtros_para_rol(current_user, terapeutaid)

    hoy = fecha_ecuador()
    inicio_semana_dt = datetime.combine(
        hoy - timedelta(days=hoy.weekday()), time.min
    )
    hace_7_dias = hoy - timedelta(days=7)

    consultorio_resuelto = _resolver_consultorioid_para_rol(
        current_user, consultorioid
    )

    # ------------------------------------------------------------------
    # QUERY 1 — Sesiones: hoy / en curso / finalizadas hoy
    # 3 COUNTs → 1 sola pasada con CASE
    # ------------------------------------------------------------------
    sesiones_q = db.query(
        func.count(
            sa_case(
                (SesionTerapia.fecha == hoy, 1),
            )
        ).label("hoy"),
        func.count(
            sa_case(
                (SesionTerapia.horasalida == None, 1),
            )
        ).label("en_curso"),
        func.count(
            sa_case(
                (
                    (SesionTerapia.fecha == hoy)
                    & (SesionTerapia.horasalida != None),
                    1,
                ),
            )
        ).label("finalizadas_hoy"),
    ).select_from(SesionTerapia)

    if current_user.rol == 2:
        sesiones_q = sesiones_q.filter(
            SesionTerapia.terapeutaid == current_user.id
        )
    elif terapeutaid is not None:
        sesiones_q = sesiones_q.filter(
            SesionTerapia.terapeutaid == terapeutaid
        )

    if consultorio_resuelto is not None:
        # Consultorio operativo: cuenta la sesión para el consultorio del
        # terapeuta que atendió, aunque el paciente sea de otra sede.
        terapeuta_sesion = aliased(Usuario)
        sesiones_q = sesiones_q.join(
            terapeuta_sesion,
            terapeuta_sesion.id == SesionTerapia.terapeutaid,
        ).filter(terapeuta_sesion.consultorioid == consultorio_resuelto)

    row_sesiones = sesiones_q.one()
    sesiones_hoy = row_sesiones.hoy or 0
    sesiones_en_curso = row_sesiones.en_curso or 0
    sesiones_finalizadas_hoy = row_sesiones.finalizadas_hoy or 0

    # ------------------------------------------------------------------
    # QUERY 2 — Pacientes: activos y nuevos esta semana
    # 2 COUNTs → 1 sola pasada con CASE
    # ------------------------------------------------------------------
    pacientes_q = db.query(
        func.count(
            sa_case(
                (Paciente.estadopaciente == 1, 1),
            )
        ).label("activos"),
        func.count(
            sa_case(
                (
                    (Paciente.estadopaciente == 1)
                    & (Paciente.fechainicio >= inicio_semana_dt),
                    1,
                ),
            )
        ).label("nuevos_semana"),
    ).select_from(Paciente)

    if current_user.rol == 2:
        pacientes_q = pacientes_q.filter(
            Paciente.terapeutaasignadoid == current_user.id
        )
    elif current_user.rol == 1:
        pacientes_q = pacientes_q.filter(
            Paciente.consultorioid == consultorio_resuelto
        )
    elif current_user.rol == 3:
        if consultorioid is not None:
            pacientes_q = pacientes_q.filter(
                Paciente.consultorioid == consultorioid
            )
        if terapeutaid is not None:
            pacientes_q = pacientes_q.filter(
                Paciente.terapeutaasignadoid == terapeutaid
            )

    row_pac = pacientes_q.one()
    pacientes_activos = row_pac.activos or 0
    pacientes_nuevos_semana = row_pac.nuevos_semana or 0

    # ------------------------------------------------------------------
    # QUERY 3 — Tratamientos: activos y sin sesión en 7 días
    # 2 COUNTs → 1 sola pasada con EXISTS como subquery
    # ------------------------------------------------------------------
    sesion_reciente_exists = exists().where(
        SesionTerapia.tratamientopacienteid == TratamientoPaciente.id,
        SesionTerapia.horasalida != None,
        SesionTerapia.fecha >= hace_7_dias,
    )

    tratamientos_q = (
        db.query(
            func.count(TratamientoPaciente.id).label("activos"),
            func.count(
                sa_case(
                    (~sesion_reciente_exists, 1),
                )
            ).label("sin_sesion_7_dias"),
        )
        .select_from(TratamientoPaciente)
        .join(Paciente, Paciente.id == TratamientoPaciente.pacienteid)
        .filter(TratamientoPaciente.activo == True)
    )

    if current_user.rol == 2:
        tratamientos_q = tratamientos_q.filter(
            _tratamiento_visible_para_terapeuta_filter(current_user.id)
        )
    elif terapeutaid is not None:
        tratamientos_q = tratamientos_q.filter(
            _tratamiento_visible_para_terapeuta_filter(terapeutaid)
        )

    if consultorio_resuelto is not None:
        tratamientos_q = tratamientos_q.filter(
            _tratamiento_visible_para_consultorio_filter(consultorio_resuelto)
        )

    row_trat = tratamientos_q.one()
    tratamientos_activos = row_trat.activos or 0
    tratamientos_sin_sesion_7_dias = row_trat.sin_sesion_7_dias or 0

    # ------------------------------------------------------------------
    # QUERY 4 — Pagos pendientes + alertas no leídas
    # Siguen siendo queries separadas porque tocan tablas distintas
    # con JOINs diferentes, consolidarlas no ahorraría nada.
    # ------------------------------------------------------------------
    pagos_q = (
        db.query(func.count(Pago.id))
        .join(
            TratamientoPaciente,
            TratamientoPaciente.id == Pago.tratamientopacienteid,
        )
        .join(Paciente, Paciente.id == TratamientoPaciente.pacienteid)
        .filter(
            Pago.estadopago == 1,
            Pago.anulado == False,
            Pago.tratamientopacienteid != None,
        )
    )

    if current_user.rol == 2:
        pagos_q = pagos_q.filter(
            _tratamiento_visible_para_terapeuta_filter(current_user.id)
        )
    elif terapeutaid is not None:
        pagos_q = pagos_q.filter(
            _tratamiento_visible_para_terapeuta_filter(terapeutaid)
        )
    if consultorio_resuelto is not None:
        pagos_q = pagos_q.filter(
            _tratamiento_visible_para_consultorio_filter(consultorio_resuelto)
        )

    transferencias_pendientes = pagos_q.scalar() or 0

    alertas_q = (
        db.query(func.count(Alerta.id))
        .join(Paciente, Paciente.id == _columna_paciente_alerta())
        .filter(Alerta.leida == False)
    )

    if current_user.rol == 2:
        alertas_q = alertas_q.filter(
            Paciente.terapeutaasignadoid == current_user.id
        )
    elif current_user.rol == 1:
        alertas_q = alertas_q.filter(
            Paciente.consultorioid == consultorio_resuelto
        )
    elif current_user.rol == 3 and consultorioid is not None:
        alertas_q = alertas_q.filter(
            Paciente.consultorioid == consultorioid
        )

    alertas_no_leidas = alertas_q.scalar() or 0

    # ------------------------------------------------------------------
    # QUERY 5 — Notificaciones no leídas + cesiones activas
    # 2 COUNTs simples sobre tablas sin JOIN pesado → 1 query con UNION
    # es más complejo que útil; mejor 2 scalars directos.
    # ------------------------------------------------------------------
    notificaciones_no_leidas = (
        db.query(func.count(Notificacion.id))
        .filter(
            Notificacion.usuarioid == current_user.id,
            Notificacion.leida == False,
        )
        .scalar()
        or 0
    )

    cesiones_q = db.query(func.count(Transferencia.id)).filter(
        Transferencia.activo == True
    )

    if current_user.rol == 1:
        terapeutas_ids = [
            row.id
            for row in db.query(Usuario.id)
            .filter(
                Usuario.rol == 2,
                Usuario.activo == True,
                Usuario.consultorioid == consultorio_resuelto,
            )
            .all()
        ]
        if terapeutas_ids:
            cesiones_q = cesiones_q.filter(
                Transferencia.terapeuta_origen_id.in_(terapeutas_ids),
                Transferencia.terapeuta_destino_id.in_(terapeutas_ids),
            )
        else:
            cesiones_q = cesiones_q.filter(Transferencia.id == -1)
    elif current_user.rol == 2:
        cesiones_q = cesiones_q.filter(
            Transferencia.terapeuta_destino_id == current_user.id
        )

    cesiones_activas = cesiones_q.scalar() or 0

    return DashboardAccionesOut(
        sesiones_hoy=sesiones_hoy,
        sesiones_en_curso=sesiones_en_curso,
        sesiones_finalizadas_hoy=sesiones_finalizadas_hoy,
        pacientes_activos=pacientes_activos,
        pacientes_nuevos_semana=pacientes_nuevos_semana,
        tratamientos_activos=tratamientos_activos,
        tratamientos_sin_sesion_7_dias=tratamientos_sin_sesion_7_dias,
        transferencias_pendientes=transferencias_pendientes,
        alertas_no_leidas=alertas_no_leidas,
        notificaciones_no_leidas=notificaciones_no_leidas,
        cesiones_activas=cesiones_activas,
    )


@router.get("/dashboard-resumen", response_model=DashboardResumenOut)
def dashboard_resumen(
    consultorioid: Optional[int] = Query(None),
    terapeutaid: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    hoy = fecha_ecuador()

    sesiones_hoy_query = db.query(SesionTerapia).filter(
        SesionTerapia.fecha == hoy
    )

    sesiones_hoy_query = _aplicar_filtros_sesiones(
        sesiones_hoy_query,
        current_user,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
    )

    sesiones_hoy = sesiones_hoy_query.count()

    pacientes_atendidos_hoy = (
        sesiones_hoy_query
        .with_entities(SesionTerapia.pacienteid)
        .distinct()
        .count()
    )

    tratamientos_activos_query = db.query(TratamientoPaciente).filter(
        TratamientoPaciente.activo == True
    )

    tratamientos_activos_query = _aplicar_filtros_tratamientos(
        tratamientos_activos_query,
        current_user,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
    )

    tratamientos_activos = tratamientos_activos_query.count()

    pagos_hoy_query = db.query(Pago).filter(
        filtro_fechapago_ecuador(hoy, hoy),
        Pago.estadopago == 2,
        _pago_no_anulado_filter(),
        _pago_de_caja_filter(),
        Pago.tratamientopacienteid != None,
    )

    pagos_hoy_query = _aplicar_filtros_pagos(
        pagos_hoy_query,
        current_user,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
    )

    ingresos_hoy = float(
        pagos_hoy_query
        .with_entities(func.coalesce(func.sum(Pago.monto), 0))
        .scalar()
        or 0
    )

    cuentas = _calcular_cuentas_tratamientos(
        db,
        current_user,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
    )

    cuentas_pendientes = sum(
        1 for item in cuentas.values() if item["saldo"] > 0
    )

    saldo_pendiente_total = sum(
        item["saldo"] for item in cuentas.values()
    )

    saldo_a_favor_total = sum(
        item["saldo_favor"] for item in cuentas.values()
    )

    transferencias_pendientes_query = db.query(Pago).filter(
        Pago.estadopago == 1,
        Pago.anulado == False,
        Pago.tratamientopacienteid != None,
    )

    transferencias_pendientes_query = _aplicar_filtros_pagos(
        transferencias_pendientes_query,
        current_user,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
    )

    transferencias_pendientes = transferencias_pendientes_query.count()

    return DashboardResumenOut(
        sesiones_hoy=sesiones_hoy,
        pacientes_atendidos_hoy=pacientes_atendidos_hoy,
        tratamientos_activos=tratamientos_activos,
        ingresos_hoy=round(ingresos_hoy, 2),
        cuentas_pendientes=cuentas_pendientes,
        saldo_pendiente_total=round(saldo_pendiente_total, 2),
        transferencias_pendientes=transferencias_pendientes,
        saldo_a_favor_total=round(saldo_a_favor_total, 2),
    )

@router.get("/dashboard-lite", response_model=DashboardLiteOut)
def dashboard_lite(
    consultorioid: Optional[int] = Query(None),
    terapeutaid: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    resumen = dashboard_resumen(
        consultorioid=consultorioid,
        terapeutaid=terapeutaid,
        db=db,
        current_user=current_user,
    )

    # Alertas no leídas
    alertas_query = db.query(Alerta).join(
        Paciente,
        Paciente.id == _columna_paciente_alerta(),
    ).filter(
        Alerta.leida == False,
    )

    if current_user.rol == 2:
        alertas_query = alertas_query.filter(
            Paciente.terapeutaasignadoid == current_user.id,
        )

    elif current_user.rol == 1:
        if current_user.consultorioid is None:
            raise HTTPException(
                status_code=403,
                detail="El secretario no tiene consultorio asignado.",
            )

        alertas_query = alertas_query.filter(
            Paciente.consultorioid == current_user.consultorioid,
        )

    elif current_user.rol == 3:
        if consultorioid is not None:
            alertas_query = alertas_query.filter(
                Paciente.consultorioid == consultorioid,
            )

    else:
        raise HTTPException(
            status_code=403,
            detail="No autorizado",
        )

    alertas_no_leidas = alertas_query.count()

    # Notificaciones no leídas del usuario actual
    notificaciones_no_leidas = (
        db.query(Notificacion)
        .filter(
            Notificacion.usuarioid == current_user.id,
            Notificacion.leida == False,
        )
        .count()
    )

    # Cesiones activas
    cesiones_query = db.query(Transferencia).filter(
        Transferencia.activo == True,
    )

    if current_user.rol == 1:
        if current_user.consultorioid is None:
            raise HTTPException(
                status_code=403,
                detail="El secretario no tiene consultorio asignado.",
            )

        terapeutas_ids = [
            row.id
            for row in db.query(Usuario.id)
            .filter(
                Usuario.rol == 2,
                Usuario.activo == True,
                Usuario.consultorioid == current_user.consultorioid,
            )
            .all()
        ]

        cesiones_query = cesiones_query.filter(
            Transferencia.terapeuta_origen_id.in_(terapeutas_ids),
            Transferencia.terapeuta_destino_id.in_(terapeutas_ids),
        )

    elif current_user.rol == 2:
        cesiones_query = cesiones_query.filter(
            Transferencia.terapeuta_destino_id == current_user.id,
        )

    elif current_user.rol == 3:
        pass

    else:
        raise HTTPException(
            status_code=403,
            detail="No autorizado",
        )

    cesiones_activas = cesiones_query.count()

    return DashboardLiteOut(
        sesiones_hoy=resumen.sesiones_hoy,
        pacientes_atendidos_hoy=resumen.pacientes_atendidos_hoy,
        tratamientos_activos=resumen.tratamientos_activos,
        ingresos_hoy=resumen.ingresos_hoy,
        cuentas_pendientes=resumen.cuentas_pendientes,
        saldo_pendiente_total=resumen.saldo_pendiente_total,
        transferencias_pendientes=resumen.transferencias_pendientes,
        saldo_a_favor_total=resumen.saldo_a_favor_total,
        alertas_no_leidas=alertas_no_leidas,
        notificaciones_no_leidas=notificaciones_no_leidas,
        cesiones_activas=cesiones_activas,
    )


# -----------------------------------------------------------------------------
# Reporte semanal antiguo: se mantiene por compatibilidad
# -----------------------------------------------------------------------------

@router.get("/sesiones/semana", response_model=ReporteSemanalResponse)
def reporte_semanal(
    fecha_inicio: date,
    terapeuta_id: Optional[int] = Query(None),
    consultorio_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_secretary),
):
    fecha_fin = fecha_inicio + timedelta(days=6)

    query = db.query(SesionTerapia).filter(
        SesionTerapia.fecha.between(fecha_inicio, fecha_fin)
    )

    query = _aplicar_filtros_sesiones(
        query,
        current_user,
        terapeutaid=terapeuta_id,
        consultorioid=consultorio_id,
    )

    detalle = query.options(
        joinedload(SesionTerapia.paciente),
        joinedload(SesionTerapia.terapeuta),
    ).all()

    conteo_por_dia_query = db.query(
        SesionTerapia.fecha,
        func.count(SesionTerapia.id).label("cantidad"),
    ).filter(
        SesionTerapia.fecha.between(fecha_inicio, fecha_fin)
    )

    conteo_por_dia_query = _aplicar_filtros_sesiones(
        conteo_por_dia_query,
        current_user,
        terapeutaid=terapeuta_id,
        consultorioid=consultorio_id,
    )

    conteo_por_dia = (
        conteo_por_dia_query
        .group_by(SesionTerapia.fecha)
        .all()
    )

    mapa_conteo = {item.fecha: item.cantidad for item in conteo_por_dia}

    sesiones_por_dia: List[SesionPorDia] = []

    for i, dia in enumerate(DIAS_SEMANA):
        dia_fecha = fecha_inicio + timedelta(days=i)

        sesiones_por_dia.append(
            SesionPorDia(
                dia=dia,
                fecha=dia_fecha,
                cantidad=mapa_conteo.get(dia_fecha, 0),
            )
        )

    detalle_final = [
        {
            "id": s.id,
            "fecha": s.fecha,
            "paciente": _nombre_paciente(s.paciente),
            "terapeuta": _nombre_usuario(s.terapeuta),
        }
        for s in detalle
    ]

    return {
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "sesiones_por_dia": sesiones_por_dia,
        "total_sesiones": len(detalle),
        "detalle": detalle_final,
    }


# -----------------------------------------------------------------------------
# Reporte general de terapias
# -----------------------------------------------------------------------------

@router.get("/terapias", response_model=TerapiasReporteOut)
def reporte_terapias(
    desde: Optional[date] = Query(None),
    hasta: Optional[date] = Query(None),
    terapeutaid: Optional[int] = Query(None),
    consultorioid: Optional[int] = Query(None),
    dia_semana: Optional[int] = Query(None, ge=0, le=6),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    desde, hasta = _default_range(desde, hasta)
    dia_semana = _validar_dia_semana(dia_semana)

    sesiones_query = (
        db.query(SesionTerapia)
        .options(
            joinedload(SesionTerapia.tratamiento_paciente),
            joinedload(SesionTerapia.paciente),
        )
        .filter(
            SesionTerapia.fecha.between(desde, hasta),
            SesionTerapia.horasalida != None,
            SesionTerapia.tratamientopacienteid != None,
        )
    )

    sesiones_query = _aplicar_filtro_dia_sesion(sesiones_query, dia_semana)

    sesiones_query = _aplicar_filtros_sesiones(
        sesiones_query,
        current_user,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
    )

    sesiones = sesiones_query.all()

    tratamiento_ids = {
        s.tratamientopacienteid
        for s in sesiones
        if s.tratamientopacienteid
    }

    cuentas = _calcular_cuentas_tratamientos(
        db,
        current_user,
        tratamiento_ids=tratamiento_ids,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
    )

    total_generado = sum(
        _precio_aplicado(s.tratamiento_paciente)
        for s in sesiones
    )

    total_ecuasanitas = sum(
        _precio_aplicado(s.tratamiento_paciente)
        for s in sesiones
        if _es_paciente_ecuasanitas(s.paciente)
    )

    sesiones_ecuasanitas = sum(
        1 for s in sesiones if _es_paciente_ecuasanitas(s.paciente)
    )

    # IMPORTANTE:
    # El resumen semanal NO debe usar la deuda acumulada del tratamiento completo.
    # Si se mezcla producción semanal con saldos históricos, aparecen pendientes
    # exagerados. Aquí dejamos:
    # - Pagado: dinero cobrado dentro del rango filtrado.
    # - Pendiente: producción del rango que no se cobró en el rango.
    # - Deuda acumulada: se calcula aparte y se consulta en su propio detalle.
    saldo_a_favor = sum(
        item["saldo_favor"]
        for item in cuentas.values()
    )

    total_pagado_verificado = 0.0
    total_pago_previo_verificado = 0.0

    pendiente_semana_detalle = _pendiente_semana_detalle(
        db=db,
        current_user=current_user,
        desde=desde,
        hasta=hasta,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
        dia_semana=dia_semana,
    )
    total_pendiente = float(pendiente_semana_detalle.total_pendiente or 0.0)

    pendiente_verificacion_total = sum(
        item["pendiente_verificacion"]
        for item in cuentas.values()
    )

    transferencias_pendientes = sum(
        1
        for item in cuentas.values()
        if item["pendiente_verificacion"] > 0
    )

    pagos_query = db.query(Pago).filter(
        filtro_fechapago_ecuador(desde, hasta),
        filtro_dia_pago_ecuador(dia_semana),
        Pago.estadopago == 2,
        _pago_no_anulado_filter(),
        _pago_de_caja_filter(),
        Pago.tratamientopacienteid != None,
    )

    pagos_query = _aplicar_filtros_pagos(
        pagos_query,
        current_user,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
    )

    total_pagado_verificado = float(
        pagos_query
        .with_entities(func.coalesce(func.sum(Pago.monto), 0))
        .scalar()
        or 0
    )

    # Si se filtra por fisioterapeuta, la caja no puede ser el pago completo
    # del tratamiento. Se reparte FIFO y solo se cuenta la parte que cubre
    # sesiones realizadas por ese fisio.
    caja_asignada_fisio = _caja_terapia_asignada_a_fisio(
        db=db,
        current_user=current_user,
        desde=desde,
        hasta=hasta,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
        dia_semana=dia_semana,
    )
    if caja_asignada_fisio is not None:
        total_pagado_verificado = float(caja_asignada_fisio["total"] or 0.0)

    transferencias_pendientes_total, transferencias_pendientes_cantidad = _totales_pendientes_transferencia(
        db=db,
        current_user=current_user,
        desde=desde,
        hasta=hasta,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
        dia_semana=dia_semana,
    )
    transferencias_gimnasio_total, transferencias_gimnasio_cantidad = _totales_pendientes_transferencia_gimnasio(
        db=db,
        current_user=current_user,
        desde=desde,
        hasta=hasta,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
        dia_semana=dia_semana,
    )
    transferencias_pendientes_total = round(
        float(transferencias_pendientes_total or 0) + float(transferencias_gimnasio_total or 0),
        2,
    )
    transferencias_pendientes_cantidad = int(transferencias_pendientes_cantidad or 0) + int(
        transferencias_gimnasio_cantidad or 0
    )

    pago_previo_query = db.query(Pago).filter(
        filtro_fechapago_ecuador(desde, hasta),
        filtro_dia_pago_ecuador(dia_semana),
        Pago.estadopago == 2,
        _pago_no_anulado_filter(),
        Pago.espagoprevio == True,
        Pago.tratamientopacienteid != None,
    )

    pago_previo_query = _aplicar_filtros_pagos(
        pago_previo_query,
        current_user,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
    )

    total_pago_previo_verificado = float(
        pago_previo_query
        .with_entities(func.coalesce(func.sum(Pago.monto), 0))
        .scalar()
        or 0
    )

    # No calcular Pend. semana como: generado - caja.
    # Eso permite que un pago adelantado de un paciente reduzca pendientes de otro.
    # El pendiente correcto ya fue calculado por paciente + tratamiento, consumiendo
    # pagos acumulados en orden FIFO hasta el fin del rango.

    deuda_acumulada = _deuda_acumulada_reporte(
        db=db,
        current_user=current_user,
        desde=desde,
        hasta=hasta,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
        dia_semana=dia_semana,
    )

    if caja_asignada_fisio is not None:
        totales_metodo = caja_asignada_fisio["totales"]
        por_metodo = list(caja_asignada_fisio["por_metodo"])
    else:
        pagos_metodo_rows = (
            pagos_query
            .with_entities(
                Pago.metodopago,
                func.coalesce(func.sum(Pago.monto), 0),
            )
            .group_by(Pago.metodopago)
            .all()
        )

        totales_metodo = {
            "total_efectivo": 0.0,
            "total_transferencia": 0.0,
            "total_tarjeta": 0.0,
            "total_otros_metodos": 0.0,
            "transferencias_pendientes_total": transferencias_pendientes_total,
            "transferencias_pendientes_cantidad": transferencias_pendientes_cantidad,
        }
        por_metodo_map: Dict[str, float] = defaultdict(float)
        for metodo, total in pagos_metodo_rows:
            total_float = float(total or 0)
            _sumar_metodo_en_resumen(totales_metodo, metodo, total_float)
            por_metodo_map[_metodo_nombre_canonico(metodo)] += total_float

        por_metodo = [
            MetodoPagoTotalOut(
                metodo=metodo,
                total=round(float(total or 0), 2),
            )
            for metodo, total in sorted(por_metodo_map.items())
            if round(float(total or 0), 2) > 0
        ]

    # Mantener los pendientes de transferencia totales: terapia + gimnasio.
    totales_metodo["transferencias_pendientes_total"] = transferencias_pendientes_total
    totales_metodo["transferencias_pendientes_cantidad"] = transferencias_pendientes_cantidad

    total_gimnasio_pagado, totales_gimnasio, por_metodo_gimnasio, gimnasio_por_dia = _totales_pagos_gimnasio_caja(
        db=db,
        current_user=current_user,
        desde=desde,
        hasta=hasta,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
        dia_semana=dia_semana,
    )

    if total_gimnasio_pagado > 0:
        total_pagado_verificado += total_gimnasio_pagado
        for key in ("total_efectivo", "total_transferencia", "total_tarjeta", "total_otros_metodos"):
            totales_metodo[key] = float(totales_metodo.get(key, 0.0) or 0.0) + float(
                totales_gimnasio.get(key, 0.0) or 0.0
            )

        por_metodo_final: Dict[str, float] = defaultdict(float)
        for item in por_metodo:
            por_metodo_final[item.metodo] += float(item.total or 0)
        for item in por_metodo_gimnasio:
            por_metodo_final[item.metodo] += float(item.total or 0)
        por_metodo = [
            MetodoPagoTotalOut(metodo=metodo, total=round(total, 2))
            for metodo, total in sorted(por_metodo_final.items())
            if round(total, 2) > 0
        ]

    tratamiento_map: Dict[str, Dict[str, float]] = {}

    dias_map: Dict[date, ReporteDiaOut] = {
        item.fecha: item
        for item in _generar_dias_reporte_filtrados(desde, hasta, dia_semana)
    }

    for sesion in sesiones:
        tratamiento = sesion.tratamiento_paciente
        nombre = tratamiento.tipotratamiento if tratamiento else "Sin tratamiento"
        precio = _precio_aplicado(tratamiento)

        item = tratamiento_map.setdefault(
            nombre,
            {
                "sesiones": 0,
                "total": 0.0,
            },
        )

        item["sesiones"] += 1
        item["total"] += precio

        if sesion.fecha in dias_map:
            dias_map[sesion.fecha].sesiones += 1
            dias_map[sesion.fecha].total_generado = round(
                dias_map[sesion.fecha].total_generado + precio,
                2,
            )
            if _es_paciente_ecuasanitas(sesion.paciente):
                dias_map[sesion.fecha].cubierto_ecuasanitas = round(
                    dias_map[sesion.fecha].cubierto_ecuasanitas + precio,
                    2,
                )

    fecha_pago_expr = fecha_pago_ecuador_expr()

    if caja_asignada_fisio is not None:
        for fecha_pago, data in caja_asignada_fisio["por_dia"].items():
            if fecha_pago in dias_map:
                dias_map[fecha_pago].pagos_verificados = round(
                    dias_map[fecha_pago].pagos_verificados + float(data.get("total", 0.0) or 0.0),
                    2,
                )
                dias_map[fecha_pago].pagos_efectivo = round(
                    dias_map[fecha_pago].pagos_efectivo + float(data.get("efectivo", 0.0) or 0.0),
                    2,
                )
                dias_map[fecha_pago].pagos_transferencia = round(
                    dias_map[fecha_pago].pagos_transferencia + float(data.get("transferencia", 0.0) or 0.0),
                    2,
                )
                dias_map[fecha_pago].pagos_tarjeta = round(
                    dias_map[fecha_pago].pagos_tarjeta + float(data.get("tarjeta", 0.0) or 0.0),
                    2,
                )
    else:
        pagos_por_dia = (
            pagos_query
            .with_entities(
                fecha_pago_expr.label("fecha_pago"),
                Pago.metodopago,
                func.coalesce(func.sum(Pago.monto), 0),
            )
            .group_by(fecha_pago_expr, Pago.metodopago)
            .all()
        )

        for fecha_pago, metodo, total in pagos_por_dia:
            if fecha_pago in dias_map:
                total_float = float(total or 0)
                categoria = _normalizar_metodo_pago(metodo)
                dias_map[fecha_pago].pagos_verificados = round(
                    dias_map[fecha_pago].pagos_verificados + total_float,
                    2,
                )
                if categoria == "efectivo":
                    dias_map[fecha_pago].pagos_efectivo = round(dias_map[fecha_pago].pagos_efectivo + total_float, 2)
                elif categoria == "transferencia":
                    dias_map[fecha_pago].pagos_transferencia = round(dias_map[fecha_pago].pagos_transferencia + total_float, 2)
                elif categoria == "tarjeta":
                    dias_map[fecha_pago].pagos_tarjeta = round(dias_map[fecha_pago].pagos_tarjeta + total_float, 2)

    for fecha_pago, data in gimnasio_por_dia.items():
        if fecha_pago in dias_map:
            total_float = float(data.get("total", 0.0) or 0.0)
            dias_map[fecha_pago].pagos_verificados = round(
                dias_map[fecha_pago].pagos_verificados + total_float,
                2,
            )
            dias_map[fecha_pago].pagos_gimnasio = round(
                dias_map[fecha_pago].pagos_gimnasio + total_float,
                2,
            )
            dias_map[fecha_pago].pagos_efectivo = round(
                dias_map[fecha_pago].pagos_efectivo + float(data.get("efectivo", 0.0) or 0.0),
                2,
            )
            dias_map[fecha_pago].pagos_transferencia = round(
                dias_map[fecha_pago].pagos_transferencia + float(data.get("transferencia", 0.0) or 0.0),
                2,
            )
            dias_map[fecha_pago].pagos_tarjeta = round(
                dias_map[fecha_pago].pagos_tarjeta + float(data.get("tarjeta", 0.0) or 0.0),
                2,
            )

    recuperacion_cartera_query = _query_recuperacion_cartera(
        db=db,
        current_user=current_user,
        desde=desde,
        hasta=hasta,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
        dia_semana=dia_semana,
    )

    total_recuperacion_cartera = float(
        recuperacion_cartera_query
        .with_entities(func.coalesce(func.sum(Pago.monto), 0))
        .scalar()
        or 0
    )

    if total_recuperacion_cartera > 0:
        total_pagado_verificado += total_recuperacion_cartera
        _sumar_metodo_en_resumen(totales_metodo, "Recuperación de cartera", total_recuperacion_cartera)
        por_metodo.append(
            MetodoPagoTotalOut(
                metodo="Recuperación de cartera",
                total=round(total_recuperacion_cartera, 2),
            )
        )

    recuperacion_por_dia = (
        recuperacion_cartera_query
        .with_entities(
            fecha_pago_expr.label("fecha_pago"),
            func.coalesce(func.sum(Pago.monto), 0),
        )
        .group_by(fecha_pago_expr)
        .all()
    )

    for fecha_pago, total in recuperacion_por_dia:
        if fecha_pago in dias_map:
            dias_map[fecha_pago].pagos_verificados = round(
                dias_map[fecha_pago].pagos_verificados + float(total or 0),
                2,
            )

    tratamientos_mas = sorted(
        [
            TratamientoRealizadoOut(
                tratamiento=nombre,
                sesiones=int(data["sesiones"]),
                total_generado=round(float(data["total"]), 2),
            )
            for nombre, data in tratamiento_map.items()
        ],
        key=lambda item: item.sesiones,
        reverse=True,
    )[:10]

    return TerapiasReporteOut(
        desde=desde,
        hasta=hasta,
        total_sesiones=len(sesiones),
        total_generado=round(total_generado, 2),
        total_pagado_verificado=round(total_pagado_verificado, 2),
        total_efectivo=round(float(totales_metodo.get("total_efectivo", 0.0) or 0.0), 2),
        total_transferencia=round(float(totales_metodo.get("total_transferencia", 0.0) or 0.0), 2),
        total_tarjeta=round(float(totales_metodo.get("total_tarjeta", 0.0) or 0.0), 2),
        total_otros_metodos=round(float(totales_metodo.get("total_otros_metodos", 0.0) or 0.0), 2),
        transferencias_pendientes_total=round(float(totales_metodo.get("transferencias_pendientes_total", transferencias_pendientes_total) or 0.0), 2),
        transferencias_pendientes_cantidad=int(totales_metodo.get("transferencias_pendientes_cantidad", transferencias_pendientes_cantidad) or 0),
        total_pago_previo_verificado=round(total_pago_previo_verificado, 2),
        total_gimnasio_pagado=round(total_gimnasio_pagado, 2),
        total_ecuasanitas=round(total_ecuasanitas, 2),
        sesiones_ecuasanitas=sesiones_ecuasanitas,
        total_pendiente=round(total_pendiente, 2),
        deuda_acumulada_total=round(deuda_acumulada.total_deuda, 2),
        deuda_acumulada_sesiones=deuda_acumulada.total_sesiones_pendientes,
        saldo_a_favor=round(saldo_a_favor, 2),
        transferencias_pendientes=transferencias_pendientes_cantidad,
        pendiente_verificacion_total=round(pendiente_verificacion_total, 2),
        por_metodo_pago=por_metodo,
        tratamientos_mas_realizados=tratamientos_mas,
        sesiones_por_dia=list(dias_map.values()),
        estado_pagos=ResumenEstadoPagosOut(
            pagado_verificado=round(max(total_pagado_verificado - total_gimnasio_pagado, 0.0), 2),
            gimnasio_pagado=round(total_gimnasio_pagado, 2),
            pago_previo=round(total_pago_previo_verificado, 2),
            pendiente_cobro=round(total_pendiente, 2),
            saldo_a_favor=round(saldo_a_favor, 2),
            pendiente_verificacion=round(pendiente_verificacion_total, 2),
            cubierto_ecuasanitas=round(total_ecuasanitas, 2),
        ),
    )


def _fecha_pago_local_ecuador(pago: Pago) -> date:
    fecha_real = getattr(pago, "fechapagoreal", None)
    if fecha_real is not None:
        return fecha_real

    value = pago.fechapago
    if value is None:
        return fecha_ecuador()
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)
    return value.astimezone(ECUADOR_TZ).date()


def _sesiones_reporte_filtradas(
    db: Session,
    current_user: Usuario,
    desde: date,
    hasta: date,
    terapeutaid: Optional[int] = None,
    consultorioid: Optional[int] = None,
    dia_semana: Optional[int] = None,
) -> List[SesionTerapia]:
    dia_semana = _validar_dia_semana(dia_semana)

    query = (
        db.query(SesionTerapia)
        .options(
            joinedload(SesionTerapia.paciente),
            joinedload(SesionTerapia.terapeuta),
            joinedload(SesionTerapia.tratamiento_paciente),
        )
        .filter(
            SesionTerapia.fecha.between(desde, hasta),
            SesionTerapia.horasalida != None,
            SesionTerapia.tratamientopacienteid != None,
        )
    )

    query = _aplicar_filtro_dia_sesion(query, dia_semana)
    query = _aplicar_filtros_sesiones(
        query,
        current_user,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
    )

    return (
        query.order_by(
            SesionTerapia.fecha,
            SesionTerapia.horaingreso,
            SesionTerapia.id,
        )
        .all()
    )


def _pagos_detalle_base_query(
    db: Session,
    current_user: Usuario,
    desde: date,
    hasta: date,
    terapeutaid: Optional[int] = None,
    consultorioid: Optional[int] = None,
    dia_semana: Optional[int] = None,
    incluir_pago_previo: bool = False,
    solo_caja: bool = True,
):
    """Query base de pagos con los mismos filtros visuales de reportes.

    Se hace manualmente en vez de usar _aplicar_filtros_pagos para poder traer
    Pago + Tratamiento + Paciente en el mismo resultado sin duplicar joins.
    """
    _validar_filtros_para_rol(current_user, terapeutaid)
    dia_semana = _validar_dia_semana(dia_semana)
    consultorio_resuelto = _resolver_consultorioid_para_rol(
        current_user,
        consultorioid,
    )

    query = (
        db.query(Pago, TratamientoPaciente, Paciente)
        .join(TratamientoPaciente, TratamientoPaciente.id == Pago.tratamientopacienteid)
        .join(Paciente, Paciente.id == TratamientoPaciente.pacienteid)
        .filter(
            filtro_fechapago_ecuador(desde, hasta),
            filtro_dia_pago_ecuador(dia_semana),
            Pago.estadopago == 2,
            _pago_no_anulado_filter(),
            Pago.tratamientopacienteid != None,
        )
    )

    if solo_caja:
        query = query.filter(_pago_de_caja_filter())
    elif not incluir_pago_previo:
        query = query.filter(or_(Pago.espagoprevio == False, Pago.espagoprevio.is_(None)))

    query = query.filter(
        or_(Pago.esrecuperacioncartera == False, Pago.esrecuperacioncartera.is_(None))
    )

    if current_user.rol == 2:
        query = query.filter(_tratamiento_visible_para_terapeuta_filter(current_user.id))
    elif terapeutaid is not None:
        query = query.filter(_tratamiento_visible_para_terapeuta_filter(terapeutaid))

    if consultorio_resuelto is not None:
        query = query.filter(_tratamiento_visible_para_consultorio_filter(consultorio_resuelto))

    return query


def _terapeutas_por_tratamiento_en_sesiones(sesiones: List[SesionTerapia]) -> Dict[Tuple[int, int], str]:
    nombres: Dict[Tuple[int, int], Set[str]] = {}
    for sesion in sesiones:
        if sesion.pacienteid is None or sesion.tratamientopacienteid is None:
            continue
        key = (int(sesion.pacienteid), int(sesion.tratamientopacienteid))
        nombres.setdefault(key, set()).add(_nombre_usuario(getattr(sesion, "terapeuta", None)))
    return {key: ", ".join(sorted(value)) for key, value in nombres.items() if value}


def _pagos_gimnasio_detalle_out(
    db: Session,
    current_user: Usuario,
    desde: date,
    hasta: date,
    terapeutaid: Optional[int] = None,
    consultorioid: Optional[int] = None,
    dia_semana: Optional[int] = None,
) -> List[CajaSemanalPagoOut]:
    """Detalle de pagos de gimnasio para el bottom sheet de caja."""
    rows = (
        _pagos_gimnasio_detalle_base_query(
            db=db,
            current_user=current_user,
            desde=desde,
            hasta=hasta,
            terapeutaid=terapeutaid,
            consultorioid=consultorioid,
            dia_semana=dia_semana,
            solo_caja=True,
        )
        .order_by(Pago.fechapago, Pago.id)
        .all()
    )

    pagos: List[CajaSemanalPagoOut] = []
    for pago, membresia, paciente, terapeuta in rows:
        modalidad = _texto_normalizado(getattr(membresia, "modalidad", ""))
        if "diario" in modalidad or "dia" in modalidad:
            concepto = "Gimnasio diario"
        else:
            concepto = "Membresía gimnasio"

        pagos.append(
            CajaSemanalPagoOut(
                pagoid=int(pago.id),
                es_gimnasio=True,
                membresiagimnasioid=int(pago.membresiagimnasioid or 0),
                fecha=_fecha_pago_local_ecuador(pago),
                pacienteid=int(paciente.id),
                paciente=_nombre_paciente(paciente),
                terapeuta=_nombre_usuario(terapeuta),
                tratamientopacienteid=0,
                tratamiento=concepto,
                metodo=pago.metodopago or "Sin método",
                monto=round(float(pago.monto or 0), 2),
                valor_sesion=0.0,
                sesiones_pagadas=0.0,
                sesiones_realizadas_semana=0,
            )
        )

    return pagos



def _caja_terapia_asignada_a_fisio(
    db: Session,
    current_user: Usuario,
    desde: date,
    hasta: date,
    terapeutaid: Optional[int] = None,
    consultorioid: Optional[int] = None,
    dia_semana: Optional[int] = None,
):
    """Caja de terapias repartida por sesiones reales del fisioterapeuta.

    Cuando se filtra por fisio, NO se debe sumar el pago completo del
    tratamiento, porque ese tratamiento puede tener sesiones hechas por varios
    terapeutas. Aquí se consume el pago FIFO por paciente + tratamiento y solo
    se cuenta la parte que cubre sesiones atendidas por el fisio filtrado.
    """
    terapeuta_objetivo: Optional[int] = current_user.id if current_user.rol == 2 else terapeutaid
    if terapeuta_objetivo is None:
        return None

    pagos_rango = (
        _pagos_detalle_base_query(
            db=db,
            current_user=current_user,
            desde=desde,
            hasta=hasta,
            terapeutaid=terapeuta_objetivo,
            consultorioid=consultorioid,
            dia_semana=dia_semana,
            solo_caja=True,
        )
        .order_by(fecha_pago_ecuador_expr(), Pago.fechapago, Pago.id)
        .all()
    )

    if not pagos_rango:
        pendiente_total, pendiente_cantidad = _totales_pendientes_transferencia(
            db=db,
            current_user=current_user,
            desde=desde,
            hasta=hasta,
            terapeutaid=terapeuta_objetivo,
            consultorioid=consultorioid,
            dia_semana=dia_semana,
        )
        return {
            "total": 0.0,
            "totales": {
                "total_efectivo": 0.0,
                "total_transferencia": 0.0,
                "total_tarjeta": 0.0,
                "total_otros_metodos": 0.0,
                "transferencias_pendientes_total": pendiente_total,
                "transferencias_pendientes_cantidad": pendiente_cantidad,
            },
            "por_metodo": [],
            "por_dia": {},
            "pagos_detalle": [],
        }

    pagos_rango_ids = {int(pago.id) for pago, _, _ in pagos_rango}
    claves = {
        (int(pago.pacienteid), int(pago.tratamientopacienteid))
        for pago, _, _ in pagos_rango
        if pago.pacienteid is not None and pago.tratamientopacienteid is not None
    }
    tratamiento_ids = {tratamiento_id for _, tratamiento_id in claves}
    paciente_ids = {paciente_id for paciente_id, _ in claves}

    sesiones = (
        db.query(SesionTerapia)
        .options(
            joinedload(SesionTerapia.paciente),
            joinedload(SesionTerapia.terapeuta),
            joinedload(SesionTerapia.tratamiento_paciente),
        )
        .filter(
            SesionTerapia.pacienteid.in_(paciente_ids),
            SesionTerapia.tratamientopacienteid.in_(tratamiento_ids),
            SesionTerapia.horasalida != None,
            SesionTerapia.fecha <= hasta,
        )
        .order_by(
            SesionTerapia.pacienteid,
            SesionTerapia.tratamientopacienteid,
            SesionTerapia.fecha,
            SesionTerapia.horaingreso,
            SesionTerapia.id,
        )
        .all()
    )

    sesiones_por_clave: Dict[Tuple[int, int], List[SesionTerapia]] = defaultdict(list)
    for sesion in sesiones:
        if sesion.pacienteid is None or sesion.tratamientopacienteid is None:
            continue
        key = (int(sesion.pacienteid), int(sesion.tratamientopacienteid))
        if key in claves:
            sesiones_por_clave[key].append(sesion)

    pagos_hasta_corte = (
        db.query(Pago)
        .filter(
            Pago.pacienteid.in_(paciente_ids),
            Pago.tratamientopacienteid.in_(tratamiento_ids),
            Pago.estadopago == 2,
            _pago_no_anulado_filter(),
            or_(Pago.esrecuperacioncartera == False, Pago.esrecuperacioncartera.is_(None)),
            fecha_pago_ecuador_expr() <= hasta,
        )
        .order_by(
            Pago.pacienteid,
            Pago.tratamientopacienteid,
            fecha_pago_ecuador_expr(),
            Pago.fechapago,
            Pago.id,
        )
        .all()
    )

    pagos_por_clave: Dict[Tuple[int, int], List[Pago]] = defaultdict(list)
    for pago in pagos_hasta_corte:
        if pago.pacienteid is None or pago.tratamientopacienteid is None:
            continue
        key = (int(pago.pacienteid), int(pago.tratamientopacienteid))
        if key in claves:
            pagos_por_clave[key].append(pago)

    totales = {
        "total_efectivo": 0.0,
        "total_transferencia": 0.0,
        "total_tarjeta": 0.0,
        "total_otros_metodos": 0.0,
        "transferencias_pendientes_total": 0.0,
        "transferencias_pendientes_cantidad": 0,
    }
    por_metodo: Dict[str, float] = defaultdict(float)
    por_dia: Dict[date, Dict[str, float]] = defaultdict(
        lambda: {"total": 0.0, "efectivo": 0.0, "transferencia": 0.0, "tarjeta": 0.0}
    )
    pagos_detalle: Dict[int, Dict] = {}

    for key, pagos_key in pagos_por_clave.items():
        sesiones_key = sesiones_por_clave.get(key, [])
        sesion_index = 0
        restante_sesion = 0.0
        sesion_actual: Optional[SesionTerapia] = None

        def avanzar_sesion():
            nonlocal sesion_index, restante_sesion, sesion_actual
            while sesion_index < len(sesiones_key):
                sesion_actual = sesiones_key[sesion_index]
                sesion_index += 1
                precio = _precio_aplicado(getattr(sesion_actual, "tratamiento_paciente", None))
                restante_sesion = round(float(precio or 0.0), 2)
                if restante_sesion > 0:
                    return True
            sesion_actual = None
            restante_sesion = 0.0
            return False

        avanzar_sesion()

        for pago in pagos_key:
            restante_pago = float(pago.monto or 0.0)
            es_pago_rango = int(pago.id) in pagos_rango_ids and not bool(getattr(pago, "espagoprevio", False))

            while restante_pago > 0 and sesion_actual is not None:
                aplicado = min(restante_pago, restante_sesion)
                restante_pago = round(restante_pago - aplicado, 2)
                restante_sesion = round(restante_sesion - aplicado, 2)

                if es_pago_rango and sesion_actual.terapeutaid == terapeuta_objetivo:
                    fecha_caja = _fecha_pago_local_ecuador(pago)
                    metodo = pago.metodopago or "Sin método"
                    categoria = _normalizar_metodo_pago(metodo)
                    monto = round(float(aplicado or 0.0), 2)
                    if monto > 0:
                        _sumar_metodo_en_resumen(totales, metodo, monto)
                        por_metodo[_metodo_nombre_canonico(metodo)] += monto
                        por_dia[fecha_caja]["total"] += monto
                        if categoria in {"efectivo", "transferencia", "tarjeta"}:
                            por_dia[fecha_caja][categoria] += monto

                        detalle = pagos_detalle.setdefault(
                            int(pago.id),
                            {
                                "pago": pago,
                                "sesion": sesion_actual,
                                "monto": 0.0,
                                "sesiones_pagadas": 0.0,
                            },
                        )
                        detalle["monto"] += monto
                        precio_sesion = _precio_aplicado(getattr(sesion_actual, "tratamiento_paciente", None))
                        if precio_sesion > 0:
                            detalle["sesiones_pagadas"] += monto / precio_sesion

                if restante_sesion <= 0:
                    avanzar_sesion()

    pendiente_total, pendiente_cantidad = _totales_pendientes_transferencia(
        db=db,
        current_user=current_user,
        desde=desde,
        hasta=hasta,
        terapeutaid=terapeuta_objetivo,
        consultorioid=consultorioid,
        dia_semana=dia_semana,
    )
    totales["transferencias_pendientes_total"] = pendiente_total
    totales["transferencias_pendientes_cantidad"] = pendiente_cantidad

    pagos_detalle_lista: List[CajaSemanalPagoOut] = []
    for data in pagos_detalle.values():
        pago = data["pago"]
        sesion = data["sesion"]
        tratamiento = getattr(sesion, "tratamiento_paciente", None)
        paciente = getattr(sesion, "paciente", None)
        precio = _precio_aplicado(tratamiento)
        pagos_detalle_lista.append(
            CajaSemanalPagoOut(
                pagoid=int(pago.id),
                fecha=_fecha_pago_local_ecuador(pago),
                pacienteid=int(sesion.pacienteid),
                paciente=_nombre_paciente(paciente),
                terapeuta=_nombre_usuario(getattr(sesion, "terapeuta", None)),
                tratamientopacienteid=int(sesion.tratamientopacienteid),
                tratamiento=(tratamiento.tipotratamiento if tratamiento else None) or "Tratamiento",
                metodo=pago.metodopago or "Sin método",
                monto=round(float(data["monto"] or 0.0), 2),
                valor_sesion=round(precio, 2),
                sesiones_pagadas=round(float(data["sesiones_pagadas"] or 0.0), 2),
                sesiones_realizadas_semana=0,
            )
        )

    pagos_detalle_lista.sort(key=lambda item: (item.fecha, item.pagoid))

    return {
        "total": round(sum(float(v or 0.0) for k, v in totales.items() if k.startswith("total_")), 2),
        "totales": {k: round(float(v or 0.0), 2) if isinstance(v, float) else v for k, v in totales.items()},
        "por_metodo": [
            MetodoPagoTotalOut(metodo=metodo, total=round(total, 2))
            for metodo, total in sorted(por_metodo.items())
            if round(total, 2) > 0
        ],
        "por_dia": por_dia,
        "pagos_detalle": pagos_detalle_lista,
    }


def _pagos_por_clave_para_pendiente(
    db: Session,
    current_user: Usuario,
    desde: date,
    hasta: date,
    tratamiento_ids: Set[int],
    terapeutaid: Optional[int] = None,
    consultorioid: Optional[int] = None,
    dia_semana: Optional[int] = None,
) -> Dict[Tuple[int, int], float]:
    if not tratamiento_ids:
        return {}

    rows = (
        _pagos_detalle_base_query(
            db=db,
            current_user=current_user,
            desde=desde,
            hasta=hasta,
            terapeutaid=terapeutaid,
            consultorioid=consultorioid,
            dia_semana=dia_semana,
            incluir_pago_previo=True,
            solo_caja=False,
        )
        .filter(Pago.tratamientopacienteid.in_(tratamiento_ids))
        .with_entities(
            Pago.pacienteid,
            Pago.tratamientopacienteid,
            func.coalesce(func.sum(Pago.monto), 0),
        )
        .group_by(Pago.pacienteid, Pago.tratamientopacienteid)
        .all()
    )

    return {
        (int(paciente_id), int(tratamiento_id)): float(total or 0)
        for paciente_id, tratamiento_id, total in rows
        if paciente_id is not None and tratamiento_id is not None
    }



def _pendiente_semana_detalle(
    db: Session,
    current_user: Usuario,
    desde: date,
    hasta: date,
    terapeutaid: Optional[int] = None,
    consultorioid: Optional[int] = None,
    dia_semana: Optional[int] = None,
) -> PendienteSemanaDetalleOut:
    """Pendiente real de las sesiones filtradas.

    Regla clave para que el reporte cuadre:
    - Caja semanal es dinero cobrado en la semana/día.
    - Pendiente semana NO es generado - caja global.
    - Cada pago solo puede cubrir sesiones del mismo paciente y mismo tratamiento.
    - Se consumen pagos acumulados en orden FIFO, desde la sesión más antigua.

    Ejemplo: si un paciente pagó $50 y esta semana solo usó una sesión de $10,
    los $40 sobrantes quedan como saldo a favor de ese mismo paciente; no reducen
    la deuda de otros pacientes.
    """
    sesiones_filtradas = _sesiones_reporte_filtradas(
        db=db,
        current_user=current_user,
        desde=desde,
        hasta=hasta,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
        dia_semana=dia_semana,
    )

    sesiones_filtradas = [
        s for s in sesiones_filtradas
        if s.pacienteid is not None and s.tratamientopacienteid is not None
    ]

    if not sesiones_filtradas:
        return PendienteSemanaDetalleOut(desde=desde, hasta=hasta)

    sesiones_filtradas_ids = {int(s.id) for s in sesiones_filtradas if s.id is not None}
    paciente_ids = {int(s.pacienteid) for s in sesiones_filtradas}
    tratamiento_ids = {int(s.tratamientopacienteid) for s in sesiones_filtradas}
    claves_visibles = {
        (int(s.pacienteid), int(s.tratamientopacienteid))
        for s in sesiones_filtradas
    }

    pagos_rows = (
        db.query(
            Pago.pacienteid,
            Pago.tratamientopacienteid,
            func.coalesce(func.sum(Pago.monto), 0),
        )
        .filter(
            Pago.pacienteid.in_(paciente_ids),
            Pago.tratamientopacienteid.in_(tratamiento_ids),
            Pago.estadopago == 2,
            _pago_no_anulado_filter(),
            or_(Pago.esrecuperacioncartera == False, Pago.esrecuperacioncartera.is_(None)),
            or_(
                Pago.espagoprevio == True,
                fecha_pago_ecuador_expr() <= hasta,
            ),
        )
        .group_by(Pago.pacienteid, Pago.tratamientopacienteid)
        .all()
    )

    disponible_por_clave: Dict[Tuple[int, int], float] = {}
    for paciente_id, tratamiento_id, total in pagos_rows:
        if paciente_id is None or tratamiento_id is None:
            continue
        key = (int(paciente_id), int(tratamiento_id))
        if key not in claves_visibles:
            continue
        disponible_por_clave[key] = float(total or 0)

    sesiones_historicas = (
        db.query(SesionTerapia)
        .options(
            joinedload(SesionTerapia.paciente),
            joinedload(SesionTerapia.terapeuta),
            joinedload(SesionTerapia.tratamiento_paciente),
        )
        .filter(
            SesionTerapia.pacienteid.in_(paciente_ids),
            SesionTerapia.tratamientopacienteid.in_(tratamiento_ids),
            SesionTerapia.horasalida != None,
            SesionTerapia.fecha <= hasta,
        )
        .order_by(
            SesionTerapia.pacienteid,
            SesionTerapia.tratamientopacienteid,
            SesionTerapia.fecha,
            SesionTerapia.horaingreso,
            SesionTerapia.id,
        )
        .all()
    )

    consultorios_map = _obtener_consultorios_map(db)
    agrupado: Dict[Tuple[int, Optional[int], int, float], Dict] = {}

    for sesion in sesiones_historicas:
        if sesion.pacienteid is None or sesion.tratamientopacienteid is None:
            continue

        key_pago = (int(sesion.pacienteid), int(sesion.tratamientopacienteid))
        if key_pago not in claves_visibles:
            continue

        tratamiento = sesion.tratamiento_paciente
        paciente = sesion.paciente
        if tratamiento is None or paciente is None:
            continue

        precio = _precio_aplicado(tratamiento)

        # Pacientes Ecuasanitas: el copago/valor de sesión pendiente se cobra
        # directamente al paciente, por eso SÍ cuenta como pendiente normal.

        # Cuando hay filtro de día de la semana, las sesiones de la misma semana
        # pero de un día diferente al filtrado NO deben consumir el saldo disponible
        # en el FIFO. Solo deben consumirlo las sesiones de semanas anteriores
        # (para el cálculo histórico correcto) y las del propio día filtrado.
        # Esto evita que, por ejemplo, una sesión del lunes consuma el saldo
        # antes de que el bucle llegue a la sesión del martes filtrado, haciendo
        # que la deuda del martes desaparezca incorrectamente.
        es_sesion_dentro_del_rango = sesion.fecha >= desde and sesion.fecha <= hasta
        es_sesion_de_otro_dia_filtrado = (
            dia_semana is not None
            and es_sesion_dentro_del_rango
            and sesion.fecha.weekday() != dia_semana
        )
        if es_sesion_de_otro_dia_filtrado:
            # No consumir saldo, no reportar: saltar completamente.
            continue

        disponible = disponible_por_clave.get(key_pago, 0.0)
        aplicado = min(precio, disponible)
        pendiente = round(max(precio - aplicado, 0.0), 2)
        disponible_por_clave[key_pago] = max(disponible - aplicado, 0.0)

        # Solo se reporta el pendiente de las sesiones que pertenecen al filtro
        # actual. Las sesiones anteriores solo sirven para consumir saldo en FIFO.
        if int(sesion.id) not in sesiones_filtradas_ids:
            continue

        if pendiente <= 0:
            continue

        terapeuta = getattr(sesion, "terapeuta", None)
        consultorio_operativo = terapeuta.consultorioid if terapeuta is not None else None
        key = (
            int(sesion.pacienteid),
            sesion.terapeutaid,
            int(sesion.tratamientopacienteid),
            round(precio, 2),
        )
        item = agrupado.setdefault(
            key,
            {
                "pacienteid": int(sesion.pacienteid),
                "paciente": _nombre_paciente(paciente),
                "terapeutaid": sesion.terapeutaid,
                "terapeuta": _nombre_usuario(terapeuta),
                "consultorioid": consultorio_operativo,
                "consultorio": consultorios_map.get(consultorio_operativo, "Sin consultorio"),
                "tratamientopacienteid": int(sesion.tratamientopacienteid),
                "tratamiento": tratamiento.tipotratamiento or "Tratamiento",
                "sesiones_pendientes": 0,
                "valor_sesion": round(precio, 2),
                "total_pendiente": 0.0,
                "fechas_pendientes": [],
            },
        )
        item["sesiones_pendientes"] = int(item["sesiones_pendientes"]) + 1
        item["total_pendiente"] = float(item["total_pendiente"]) + pendiente
        item["fechas_pendientes"].append(sesion.fecha)

    pacientes = [
        PendienteSemanaPacienteOut(
            pacienteid=int(item["pacienteid"]),
            paciente=str(item["paciente"]),
            terapeutaid=item["terapeutaid"],
            terapeuta=str(item["terapeuta"]),
            consultorioid=item["consultorioid"],
            consultorio=str(item["consultorio"]),
            tratamientopacienteid=int(item["tratamientopacienteid"]),
            tratamiento=str(item["tratamiento"]),
            sesiones_pendientes=int(item["sesiones_pendientes"]),
            valor_sesion=round(float(item["valor_sesion"]), 2),
            total_pendiente=round(float(item["total_pendiente"]), 2),
            fechas_pendientes=sorted(item["fechas_pendientes"]),
        )
        for item in agrupado.values()
    ]
    pacientes.sort(key=lambda item: (item.paciente.lower(), item.terapeuta.lower()))

    return PendienteSemanaDetalleOut(
        desde=desde,
        hasta=hasta,
        total_pendiente=round(sum(item.total_pendiente for item in pacientes), 2),
        total_sesiones_pendientes=sum(item.sesiones_pendientes for item in pacientes),
        pacientes=pacientes,
    )


@router.get("/caja-semanal-detalle", response_model=CajaSemanalDetalleOut)
def reporte_caja_semanal_detalle(
    desde: date,
    hasta: date,
    terapeutaid: Optional[int] = Query(None),
    consultorioid: Optional[int] = Query(None),
    dia_semana: Optional[int] = Query(None, ge=0, le=6),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    desde, hasta = _default_range(desde, hasta)
    dia_semana = _validar_dia_semana(dia_semana)

    caja_asignada_fisio = _caja_terapia_asignada_a_fisio(
        db=db,
        current_user=current_user,
        desde=desde,
        hasta=hasta,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
        dia_semana=dia_semana,
    )
    if caja_asignada_fisio is not None:
        totales = caja_asignada_fisio["totales"]
        pagos_asignados = list(caja_asignada_fisio["pagos_detalle"])
        pagos_gimnasio = _pagos_gimnasio_detalle_out(
            db=db,
            current_user=current_user,
            desde=desde,
            hasta=hasta,
            terapeutaid=terapeutaid,
            consultorioid=consultorioid,
            dia_semana=dia_semana,
        )
        total_gimnasio = round(sum(item.monto for item in pagos_gimnasio), 2)
        for item in pagos_gimnasio:
            _sumar_metodo_en_resumen(totales, item.metodo, item.monto)
        pagos_asignados.extend(pagos_gimnasio)

        pendiente_gym_total, pendiente_gym_cantidad = _totales_pendientes_transferencia_gimnasio(
            db=db,
            current_user=current_user,
            desde=desde,
            hasta=hasta,
            terapeutaid=terapeutaid,
            consultorioid=consultorioid,
            dia_semana=dia_semana,
        )
        transferencias_total = round(float(totales.get("transferencias_pendientes_total", 0.0) or 0.0) + pendiente_gym_total, 2)
        transferencias_cantidad = int(totales.get("transferencias_pendientes_cantidad", 0) or 0) + pendiente_gym_cantidad

        return CajaSemanalDetalleOut(
            desde=desde,
            hasta=hasta,
            total_caja=round(float(caja_asignada_fisio["total"] or 0.0) + total_gimnasio, 2),
            total_pagos=len(pagos_asignados),
            total_sesiones_pagadas=round(sum(item.sesiones_pagadas for item in pagos_asignados), 2),
            total_gimnasio=total_gimnasio,
            total_efectivo=round(float(totales.get("total_efectivo", 0.0) or 0.0), 2),
            total_transferencia=round(float(totales.get("total_transferencia", 0.0) or 0.0), 2),
            total_tarjeta=round(float(totales.get("total_tarjeta", 0.0) or 0.0), 2),
            total_otros_metodos=round(float(totales.get("total_otros_metodos", 0.0) or 0.0), 2),
            transferencias_pendientes_total=transferencias_total,
            transferencias_pendientes_cantidad=transferencias_cantidad,
            pagos=pagos_asignados,
        )

    sesiones = _sesiones_reporte_filtradas(
        db=db,
        current_user=current_user,
        desde=desde,
        hasta=hasta,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
        dia_semana=dia_semana,
    )
    sesiones_por_clave: Dict[Tuple[int, int], int] = {}
    for sesion in sesiones:
        if sesion.pacienteid is None or sesion.tratamientopacienteid is None:
            continue
        key = (int(sesion.pacienteid), int(sesion.tratamientopacienteid))
        sesiones_por_clave[key] = sesiones_por_clave.get(key, 0) + 1
    terapeutas_por_clave = _terapeutas_por_tratamiento_en_sesiones(sesiones)

    rows = (
        _pagos_detalle_base_query(
            db=db,
            current_user=current_user,
            desde=desde,
            hasta=hasta,
            terapeutaid=terapeutaid,
            consultorioid=consultorioid,
            dia_semana=dia_semana,
            solo_caja=True,
        )
        .order_by(Pago.fechapago, Pago.id)
        .all()
    )

    pagos: List[CajaSemanalPagoOut] = []
    for pago, tratamiento, paciente in rows:
        precio = _precio_aplicado(tratamiento)
        key = (int(pago.pacienteid), int(pago.tratamientopacienteid))
        sesiones_pagadas = round(float(pago.monto or 0) / precio, 2) if precio > 0 else 0.0
        pagos.append(
            CajaSemanalPagoOut(
                pagoid=int(pago.id),
                fecha=_fecha_pago_local_ecuador(pago),
                pacienteid=int(paciente.id),
                paciente=_nombre_paciente(paciente),
                terapeuta=terapeutas_por_clave.get(key, "Fisio no asociado a sesión de la semana"),
                tratamientopacienteid=int(tratamiento.id),
                tratamiento=tratamiento.tipotratamiento or "Tratamiento",
                metodo=pago.metodopago or "Sin método",
                monto=round(float(pago.monto or 0), 2),
                valor_sesion=round(precio, 2),
                sesiones_pagadas=sesiones_pagadas,
                sesiones_realizadas_semana=int(sesiones_por_clave.get(key, 0)),
            )
        )

    pagos_gimnasio = _pagos_gimnasio_detalle_out(
        db=db,
        current_user=current_user,
        desde=desde,
        hasta=hasta,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
        dia_semana=dia_semana,
    )
    total_gimnasio = round(sum(item.monto for item in pagos_gimnasio), 2)
    pagos.extend(pagos_gimnasio)

    totales_metodo = {
        "total_efectivo": 0.0,
        "total_transferencia": 0.0,
        "total_tarjeta": 0.0,
        "total_otros_metodos": 0.0,
    }
    for item in pagos:
        _sumar_metodo_en_resumen(totales_metodo, item.metodo, item.monto)

    transferencias_pendientes_total, transferencias_pendientes_cantidad = _totales_pendientes_transferencia(
        db=db,
        current_user=current_user,
        desde=desde,
        hasta=hasta,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
        dia_semana=dia_semana,
    )

    transferencias_gym_total, transferencias_gym_cantidad = _totales_pendientes_transferencia_gimnasio(
        db=db,
        current_user=current_user,
        desde=desde,
        hasta=hasta,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
        dia_semana=dia_semana,
    )
    transferencias_pendientes_total = round(
        float(transferencias_pendientes_total or 0) + float(transferencias_gym_total or 0),
        2,
    )
    transferencias_pendientes_cantidad = int(transferencias_pendientes_cantidad or 0) + int(
        transferencias_gym_cantidad or 0
    )

    return CajaSemanalDetalleOut(
        desde=desde,
        hasta=hasta,
        total_caja=round(sum(item.monto for item in pagos), 2),
        total_pagos=len(pagos),
        total_sesiones_pagadas=round(sum(item.sesiones_pagadas for item in pagos), 2),
        total_gimnasio=total_gimnasio,
        total_efectivo=round(float(totales_metodo.get("total_efectivo", 0.0) or 0.0), 2),
        total_transferencia=round(float(totales_metodo.get("total_transferencia", 0.0) or 0.0), 2),
        total_tarjeta=round(float(totales_metodo.get("total_tarjeta", 0.0) or 0.0), 2),
        total_otros_metodos=round(float(totales_metodo.get("total_otros_metodos", 0.0) or 0.0), 2),
        transferencias_pendientes_total=transferencias_pendientes_total,
        transferencias_pendientes_cantidad=transferencias_pendientes_cantidad,
        pagos=pagos,
    )


@router.get("/pendiente-semana-detalle", response_model=PendienteSemanaDetalleOut)
def reporte_pendiente_semana_detalle(
    desde: date,
    hasta: date,
    terapeutaid: Optional[int] = Query(None),
    consultorioid: Optional[int] = Query(None),
    dia_semana: Optional[int] = Query(None, ge=0, le=6),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    desde, hasta = _default_range(desde, hasta)
    dia_semana = _validar_dia_semana(dia_semana)
    return _pendiente_semana_detalle(
        db=db,
        current_user=current_user,
        desde=desde,
        hasta=hasta,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
        dia_semana=dia_semana,
    )


@router.get("/saldo-favor-detalle", response_model=SaldoFavorDetalleOut)
def reporte_saldo_favor_detalle(
    desde: date,
    hasta: date,
    terapeutaid: Optional[int] = Query(None),
    consultorioid: Optional[int] = Query(None),
    dia_semana: Optional[int] = Query(None, ge=0, le=6),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    desde, hasta = _default_range(desde, hasta)
    dia_semana = _validar_dia_semana(dia_semana)

    sesiones = _sesiones_reporte_filtradas(
        db=db,
        current_user=current_user,
        desde=desde,
        hasta=hasta,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
        dia_semana=dia_semana,
    )
    tratamiento_ids = {
        int(s.tratamientopacienteid)
        for s in sesiones
        if s.tratamientopacienteid is not None
    }
    if not tratamiento_ids:
        return SaldoFavorDetalleOut(desde=desde, hasta=hasta)

    cuentas = _calcular_cuentas_tratamientos(
        db,
        current_user,
        tratamiento_ids=tratamiento_ids,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
    )
    consultorios_map = _obtener_consultorios_map(db)
    terapeuta_ids = set()

    tratamientos = (
        db.query(TratamientoPaciente)
        .options(joinedload(TratamientoPaciente.paciente))
        .filter(TratamientoPaciente.id.in_(tratamiento_ids))
        .all()
    )
    for tratamiento in tratamientos:
        paciente = tratamiento.paciente
        if paciente is not None and paciente.terapeutaasignadoid is not None:
            terapeuta_ids.add(int(paciente.terapeutaasignadoid))

    terapeutas_map = {
        int(usuario.id): _nombre_usuario(usuario)
        for usuario in db.query(Usuario).filter(Usuario.id.in_(terapeuta_ids)).all()
    } if terapeuta_ids else {}

    pacientes: List[SaldoFavorPacienteOut] = []
    for tratamiento in tratamientos:
        item = cuentas.get(tratamiento.id)
        if not item:
            continue
        saldo = round(float(item.get("saldo_favor", 0.0) or 0.0), 2)
        if saldo <= 0:
            continue
        precio = round(float(item.get("precio", _precio_aplicado(tratamiento)) or 0.0), 2)
        paciente = tratamiento.paciente
        sesiones_disponibles = round(saldo / precio, 2) if precio > 0 else 0.0
        terapeuta_id = paciente.terapeutaasignadoid if paciente is not None else None
        consultorio_id = paciente.consultorioid if paciente is not None else None
        pacientes.append(
            SaldoFavorPacienteOut(
                pacienteid=int(tratamiento.pacienteid),
                paciente=_nombre_paciente(paciente),
                terapeutaid=terapeuta_id,
                terapeuta=terapeutas_map.get(terapeuta_id, "Sin fisioterapeuta"),
                consultorioid=consultorio_id,
                consultorio=consultorios_map.get(consultorio_id, "Sin consultorio"),
                tratamientopacienteid=int(tratamiento.id),
                tratamiento=tratamiento.tipotratamiento or "Tratamiento",
                valor_sesion=precio,
                saldo_favor=saldo,
                sesiones_disponibles=sesiones_disponibles,
            )
        )

    pacientes.sort(key=lambda item: (item.paciente.lower(), item.tratamiento.lower()))
    return SaldoFavorDetalleOut(
        desde=desde,
        hasta=hasta,
        total_saldo_favor=round(sum(item.saldo_favor for item in pacientes), 2),
        total_sesiones_disponibles=round(sum(item.sesiones_disponibles for item in pacientes), 2),
        pacientes=pacientes,
    )


@router.get("/deuda-acumulada", response_model=DeudaAcumuladaOut)
def reporte_deuda_acumulada(
    desde: date,
    hasta: date,
    terapeutaid: Optional[int] = Query(None),
    consultorioid: Optional[int] = Query(None),
    dia_semana: Optional[int] = Query(None, ge=0, le=6),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    desde, hasta = _default_range(desde, hasta)
    dia_semana = _validar_dia_semana(dia_semana)

    return _deuda_acumulada_reporte(
        db=db,
        current_user=current_user,
        desde=desde,
        hasta=hasta,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
        dia_semana=dia_semana,
    )


# -----------------------------------------------------------------------------
# Reporte semanal de fisioterapeutas
# -----------------------------------------------------------------------------

@router.get("/fisioterapeutas-semanal", response_model=List[FisioSemanalOut])
def reporte_fisioterapeutas_semanal(
    desde: date,
    hasta: date,
    terapeutaid: Optional[int] = Query(None),
    consultorioid: Optional[int] = Query(None),
    dia_semana: Optional[int] = Query(None, ge=0, le=6),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    desde, hasta = _default_range(desde, hasta)
    dia_semana = _validar_dia_semana(dia_semana)

    sesiones_query = (
        db.query(SesionTerapia)
        .options(
            joinedload(SesionTerapia.terapeuta),
            joinedload(SesionTerapia.paciente),
            joinedload(SesionTerapia.tratamiento_paciente),
        )
        .filter(
            SesionTerapia.fecha.between(desde, hasta),
            SesionTerapia.horasalida != None,
            SesionTerapia.tratamientopacienteid != None,
        )
    )

    sesiones_query = _aplicar_filtro_dia_sesion(sesiones_query, dia_semana)

    sesiones_query = _aplicar_filtros_sesiones(
        sesiones_query,
        current_user,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
    )

    sesiones = sesiones_query.order_by(
        SesionTerapia.fecha,
        SesionTerapia.horaingreso,
        SesionTerapia.id,
    ).all()

    consultorios_map = _obtener_consultorios_map(db)

    pagos_gimnasio_rows = _pagos_gimnasio_por_terapeuta(
        db=db,
        current_user=current_user,
        desde=desde,
        hasta=hasta,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
        dia_semana=dia_semana,
    )

    tratamiento_ids = {
        s.tratamientopacienteid
        for s in sesiones
        if s.tratamientopacienteid
    }

    cobertura_sesiones = _cobertura_sesiones_filtradas(
        db,
        sesiones,
        hasta,
    )

    data: Dict[int, Dict[str, float | int | str | None]] = {}
    sesiones_no_ecuasanitas: List[Tuple[int, int, float, date, float]] = []

    # Cantidad de atenciones por fisioterapeuta y por día.
    # Sirve para aplicar la regla: lunes-viernes > 15 atenciones => $4 todas.
    atenciones_por_terapeuta_dia: Dict[Tuple[int, date], int] = defaultdict(int)
    for s in sesiones:
        if s.terapeutaid is not None and s.fecha is not None:
            atenciones_por_terapeuta_dia[(s.terapeutaid, s.fecha)] += 1

    for sesion in sesiones:
        terapeuta_id = sesion.terapeutaid
        tratamiento_id = sesion.tratamientopacienteid
        precio = _precio_aplicado(sesion.tratamiento_paciente)
        consultorio_id = sesion.paciente.consultorioid if sesion.paciente else None

        item = data.setdefault(
            terapeuta_id,
            {
                "terapeuta": _nombre_usuario(sesion.terapeuta),
                "consultorioid": consultorio_id,
                "consultorio": consultorios_map.get(
                    consultorio_id,
                    "Sin consultorio",
                ),
                "sesiones": 0,
                "total_generado": 0.0,
                "total_ecuasanitas": 0.0,
                "total_gimnasio_pagado": 0.0,
                "ganancia_terapia_total": 0.0,
                "ganancia_terapia_ecuasanitas": 0.0,
            },
        )

        atenciones_dia = atenciones_por_terapeuta_dia.get((terapeuta_id, sesion.fecha), 0)
        ganancia_fisio = sueldo_fisio_terapia(
            sesion.fecha,
            atenciones_dia,
            _tratamiento_tarifa_especial_5(sesion.tratamiento_paciente),
        )
        item["sesiones"] = int(item["sesiones"]) + 1
        item["total_generado"] = float(item["total_generado"]) + precio
        item["ganancia_terapia_total"] = float(item["ganancia_terapia_total"]) + ganancia_fisio

        if _es_paciente_ecuasanitas(sesion.paciente):
            item["total_ecuasanitas"] = (
                float(item.get("total_ecuasanitas", 0.0)) + precio
            )
            item["ganancia_terapia_ecuasanitas"] = (
                float(item.get("ganancia_terapia_ecuasanitas", 0.0)) + ganancia_fisio
            )
        else:
            sesiones_no_ecuasanitas.append(
                (terapeuta_id, int(sesion.id), precio, sesion.fecha, ganancia_fisio)
            )

    for row in pagos_gimnasio_rows:
        terapeuta_id = int(row.terapeutaid)
        consultorio_id = row.consultorioid
        total_gimnasio_pagado = float(row.total_pagado or 0)

        item = data.setdefault(
            terapeuta_id,
            {
                "terapeuta": f"{row.nombres or ''} {row.apellidos or ''}".strip() or "Sin terapeuta",
                "consultorioid": consultorio_id,
                "consultorio": consultorios_map.get(
                    consultorio_id,
                    "Sin consultorio",
                ),
                "sesiones": 0,
                "total_generado": 0.0,
                "total_ecuasanitas": 0.0,
                "total_gimnasio_pagado": 0.0,
                "ganancia_terapia_total": 0.0,
                "ganancia_terapia_ecuasanitas": 0.0,
            },
        )

        item["total_gimnasio_pagado"] = (
            float(item.get("total_gimnasio_pagado", 0.0)) + total_gimnasio_pagado
        )

    pagado_por_terapeuta: Dict[int, float] = {tid: 0.0 for tid in data.keys()}
    pendiente_real_por_terapeuta: Dict[int, float] = {tid: 0.0 for tid in data.keys()}
    ganancia_cobrada_no_ecuasanitas_por_terapeuta: Dict[int, float] = {tid: 0.0 for tid in data.keys()}
    ganancia_pendiente_por_terapeuta: Dict[int, float] = {tid: 0.0 for tid in data.keys()}

    # El sueldo del fisioterapeuta depende de que la sesión esté cubierta/pagada.
    # Si la sesión todavía tiene saldo pendiente, el sueldo de esa sesión queda pendiente.
    for terapeuta_id, sesion_id, precio, fecha_sesion, sueldo_sesion in sesiones_no_ecuasanitas:
        aplicado, pendiente = cobertura_sesiones.get(
            int(sesion_id),
            (0.0, float(precio or 0.0)),
        )

        pagado_por_terapeuta[terapeuta_id] = (
            pagado_por_terapeuta.get(terapeuta_id, 0.0) + aplicado
        )
        pendiente_real_por_terapeuta[terapeuta_id] = (
            pendiente_real_por_terapeuta.get(terapeuta_id, 0.0) + pendiente
        )

        if pendiente <= 0.009:
            ganancia_cobrada_no_ecuasanitas_por_terapeuta[terapeuta_id] = (
                ganancia_cobrada_no_ecuasanitas_por_terapeuta.get(terapeuta_id, 0.0)
                + float(sueldo_sesion or 0.0)
            )
        else:
            ganancia_pendiente_por_terapeuta[terapeuta_id] = (
                ganancia_pendiente_por_terapeuta.get(terapeuta_id, 0.0)
                + float(sueldo_sesion or 0.0)
            )

    resultado: List[FisioSemanalOut] = []

    for tid, item in data.items():
        total_terapia_generado = float(item["total_generado"])
        total_ecuasanitas = float(item.get("total_ecuasanitas", 0.0))
        total_terapia_pagado = float(pagado_por_terapeuta.get(tid, 0.0))
        pendiente_terapia = float(pendiente_real_por_terapeuta.get(tid, 0.0))
        total_gimnasio_pagado = float(item.get("total_gimnasio_pagado", 0.0))

        ganancia_terapia_total = float(item.get("ganancia_terapia_total", 0.0))
        ganancia_terapia_ecuasanitas = float(item.get("ganancia_terapia_ecuasanitas", 0.0))
        ganancia_terapia_cobrada = (
            ganancia_terapia_ecuasanitas
            + float(ganancia_cobrada_no_ecuasanitas_por_terapeuta.get(tid, 0.0))
        )
        ganancia_terapia_pendiente = float(
            ganancia_pendiente_por_terapeuta.get(tid, 0.0)
        )
        ganancia_gimnasio_cobrada = total_gimnasio_pagado * PORCENTAJE_FISIO_GIMNASIO

        resultado.append(
            FisioSemanalOut(
                terapeutaid=tid,
                terapeuta=str(item["terapeuta"]),
                consultorioid=item.get("consultorioid"),
                consultorio=str(item.get("consultorio") or "Sin consultorio"),
                sesiones_realizadas=int(item["sesiones"]),
                total_generado=round(total_terapia_generado, 2),
                total_pagado_pacientes=round(total_terapia_pagado, 2),
                total_pendiente_pacientes=round(pendiente_terapia, 2),
                total_ecuasanitas=round(total_ecuasanitas, 2),
                total_gimnasio_pagado=round(total_gimnasio_pagado, 2),
                ganancia_terapia_total=round(ganancia_terapia_total, 2),
                ganancia_terapia_cobrada=round(ganancia_terapia_cobrada, 2),
                ganancia_terapia_pendiente=round(ganancia_terapia_pendiente, 2),
                ganancia_terapia_ecuasanitas=round(ganancia_terapia_ecuasanitas, 2),
                ganancia_gimnasio_cobrada=round(ganancia_gimnasio_cobrada, 2),
                ganancia_fisio_total=round(ganancia_terapia_total + ganancia_gimnasio_cobrada, 2),
                ganancia_fisio_cobrada=round(ganancia_terapia_cobrada + ganancia_gimnasio_cobrada, 2),
                ganancia_fisio_pendiente=round(ganancia_terapia_pendiente, 2),
            )
        )

    return sorted(
        resultado,
        key=lambda item: item.total_generado,
        reverse=True,
    )


@router.get("/fisioterapeutas-detalle", response_model=FisioDetalleOut)
def reporte_fisioterapeuta_detalle(
    terapeutaid: int,
    desde: date,
    hasta: date,
    consultorioid: Optional[int] = Query(None),
    dia_semana: Optional[int] = Query(None, ge=0, le=6),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    desde, hasta = _default_range(desde, hasta)
    dia_semana = _validar_dia_semana(dia_semana)

    _validar_filtros_para_rol(current_user, terapeutaid)

    terapeuta = db.query(Usuario).filter(Usuario.id == terapeutaid).first()

    if not terapeuta:
        raise HTTPException(
            status_code=404,
            detail="Terapeuta no encontrado",
        )

    _validar_terapeuta_para_secretario(current_user, terapeuta)

    sesiones_query = (
        db.query(SesionTerapia)
        .options(
            joinedload(SesionTerapia.paciente),
            joinedload(SesionTerapia.tratamiento_paciente),
        )
        .filter(
            SesionTerapia.terapeutaid == terapeutaid,
            SesionTerapia.fecha.between(desde, hasta),
            SesionTerapia.horasalida != None,
            SesionTerapia.tratamientopacienteid != None,
        )
    )

    sesiones_query = _aplicar_filtro_dia_sesion(sesiones_query, dia_semana)

    sesiones_query = _aplicar_filtros_sesiones(
        sesiones_query,
        current_user,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
    )

    sesiones = sesiones_query.order_by(
        SesionTerapia.fecha,
        SesionTerapia.horaingreso,
        SesionTerapia.id,
    ).all()

    consultorios_map = _obtener_consultorios_map(db)

    tratamiento_ids = {
        s.tratamientopacienteid
        for s in sesiones
        if s.tratamientopacienteid
    }

    cobertura_sesiones = _cobertura_sesiones_filtradas(
        db,
        sesiones,
        hasta,
    )

    agrupado: Dict[Tuple[int, int], Dict] = {}

    atenciones_por_dia: Dict[date, int] = defaultdict(int)
    for s in sesiones:
        if s.fecha is not None:
            atenciones_por_dia[s.fecha] += 1

    sesiones_por_dia: Dict[date, List[SesionTerapia]] = defaultdict(list)
    for s in sesiones:
        if s.fecha is not None:
            sesiones_por_dia[s.fecha].append(s)

    dias_sueldo: List[FisioDetalleDiaSueldoOut] = []
    for fecha_dia, sesiones_dia in sorted(sesiones_por_dia.items()):
        atenciones = len(sesiones_dia)
        es_fin_semana = _es_fin_semana(fecha_dia)
        es_bono_productividad = (
            not es_fin_semana
            and atenciones > UMBRAL_ATENCIONES_BONO_DIARIO
        )
        motivos_tarifa_5: Dict[str, int] = defaultdict(int)
        for sd in sesiones_dia:
            motivo_especial = _motivo_tarifa_especial_5(sd.tratamiento_paciente)
            if motivo_especial:
                motivos_tarifa_5[motivo_especial] += 1

        atenciones_tarifa_especial_5 = sum(motivos_tarifa_5.values())
        tarifa_base = sueldo_fisio_terapia(fecha_dia, atenciones, False)
        total_sueldo = 0.0
        sueldo_cobrado_dia = 0.0
        sueldo_pendiente_dia = 0.0
        sesiones_pagadas_dia = 0
        sesiones_pendientes_pago_dia = 0
        monto_pendiente_pacientes_dia = 0.0

        for sd in sesiones_dia:
            sueldo_sesion = sueldo_fisio_terapia(
                fecha_dia,
                atenciones,
                _tratamiento_tarifa_especial_5(sd.tratamiento_paciente),
            )
            total_sueldo += sueldo_sesion

            # Ecuasanitas se considera cubierta para sueldo del fisio.
            if _es_paciente_ecuasanitas(sd.paciente):
                sueldo_cobrado_dia += sueldo_sesion
                sesiones_pagadas_dia += 1
                continue

            aplicado, saldo = cobertura_sesiones.get(
                int(sd.id),
                (0.0, _precio_aplicado(sd.tratamiento_paciente)),
            )
            monto_pendiente_pacientes_dia += float(saldo or 0.0)

            # Regla CORPOFIT: si la sesión no está totalmente cubierta,
            # el sueldo de esa atención queda pendiente hasta que el paciente pague.
            if float(saldo or 0.0) <= 0.009:
                sueldo_cobrado_dia += sueldo_sesion
                sesiones_pagadas_dia += 1
            else:
                sueldo_pendiente_dia += sueldo_sesion
                sesiones_pendientes_pago_dia += 1

        detalle_tarifa_5 = ", ".join(
            f"{cantidad} {motivo}"
            for motivo, cantidad in sorted(motivos_tarifa_5.items())
        )

        reglas: List[str] = []
        if es_bono_productividad:
            reglas.append(
                f"Bono +{UMBRAL_ATENCIONES_BONO_DIARIO}: la tarifa base del día sube a $4.00"
            )
        elif es_fin_semana:
            reglas.append("Fin de semana: tarifa base $4.00")
        else:
            reglas.append("Tarifa base normal $3.50")

        if atenciones_tarifa_especial_5 > 0:
            reglas.append(f"Especial $5: {detalle_tarifa_5}")

        motivo = " • ".join(reglas)

        dias_sueldo.append(
            FisioDetalleDiaSueldoOut(
                fecha=fecha_dia,
                dia_semana=fecha_dia.weekday(),
                atenciones=atenciones,
                tarifa=round(float(tarifa_base), 2),
                total_sueldo=round(float(total_sueldo), 2),
                sueldo_cobrado=round(float(sueldo_cobrado_dia), 2),
                sueldo_pendiente=round(float(sueldo_pendiente_dia), 2),
                sesiones_pagadas=sesiones_pagadas_dia,
                sesiones_pendientes_pago=sesiones_pendientes_pago_dia,
                monto_pendiente_pacientes=round(float(monto_pendiente_pacientes_dia), 2),
                es_bono_productividad=es_bono_productividad,
                es_fin_semana=es_fin_semana,
                atenciones_multiple_extremidad=atenciones_tarifa_especial_5,
                motivo=motivo,
            )
        )

    for sesion in sesiones:
        tratamiento = sesion.tratamiento_paciente
        consultorio_id = sesion.paciente.consultorioid if sesion.paciente else None
        key = (sesion.pacienteid, sesion.tratamientopacienteid)

        item = agrupado.setdefault(
            key,
            {
                "pacienteid": sesion.pacienteid,
                "paciente": _nombre_paciente(sesion.paciente),
                "tratamientopacienteid": sesion.tratamientopacienteid,
                "tratamiento": tratamiento.tipotratamiento if tratamiento else "Sin tratamiento",
                "consultorioid": consultorio_id,
                "consultorio": consultorios_map.get(
                    consultorio_id,
                    "Sin consultorio",
                ),
                "sesiones": 0,
                "precio_sesion": _precio_aplicado(tratamiento),
                "total_generado": 0.0,
                "ganancia_total": 0.0,
                "es_ecuasanitas": _es_paciente_ecuasanitas(sesion.paciente),
                "multiple_extremidad": _tratamiento_tarifa_especial_5(tratamiento),
                "sesiones_multiple_extremidad": 0,
                "sesiones_detalle": [],
            },
        )

        precio = _precio_aplicado(tratamiento)
        atenciones_dia = atenciones_por_dia.get(sesion.fecha, 0)
        sueldo_sesion = sueldo_fisio_terapia(
            sesion.fecha,
            atenciones_dia,
            _tratamiento_tarifa_especial_5(tratamiento),
        )

        item["sesiones"] += 1
        item["total_generado"] += precio
        item["ganancia_total"] += sueldo_sesion
        if _tratamiento_tarifa_especial_5(tratamiento):
            item["multiple_extremidad"] = True
            item["sesiones_multiple_extremidad"] += 1
        item["sesiones_detalle"].append((int(sesion.id), precio, sesion.fecha, sueldo_sesion))

        if _es_paciente_ecuasanitas(sesion.paciente):
            item["es_ecuasanitas"] = True

    pacientes: List[FisioDetallePacienteOut] = []

    for (_, tratamiento_id), item in agrupado.items():
        generado = float(item["total_generado"])
        ganancia_total = float(item.get("ganancia_total", 0.0))
        es_ecuasanitas = bool(item.get("es_ecuasanitas", False))

        if es_ecuasanitas:
            pagado = 0.0
            pendiente = 0.0
            cubierto_ecuasanitas = generado
            ganancia_cobrada = ganancia_total
            ganancia_pendiente = 0.0
        else:
            pagado = 0.0
            pendiente = 0.0
            ganancia_cobrada = 0.0
            ganancia_pendiente = 0.0

            for sesion_id, precio, fecha_sesion, sueldo_sesion in item.get("sesiones_detalle", []):
                aplicado, saldo = cobertura_sesiones.get(
                    int(sesion_id),
                    (0.0, float(precio or 0.0)),
                )

                pagado += aplicado
                pendiente += saldo

                if saldo <= 0.009:
                    ganancia_cobrada += float(sueldo_sesion or 0.0)
                else:
                    ganancia_pendiente += float(sueldo_sesion or 0.0)

            cubierto_ecuasanitas = 0.0

        pacientes.append(
            FisioDetallePacienteOut(
                pacienteid=item["pacienteid"],
                paciente=item["paciente"],
                tratamientopacienteid=item["tratamientopacienteid"],
                tratamiento=item["tratamiento"],
                consultorioid=item["consultorioid"],
                consultorio=item["consultorio"],
                sesiones=item["sesiones"],
                precio_sesion=round(float(item["precio_sesion"]), 2),
                total_generado=round(generado, 2),
                pagado_paciente=round(pagado, 2),
                pendiente_paciente=round(pendiente, 2),
                es_ecuasanitas=es_ecuasanitas,
                multiple_extremidad=bool(item.get("multiple_extremidad", False)),
                sesiones_multiple_extremidad=int(item.get("sesiones_multiple_extremidad", 0)),
                cubierto_ecuasanitas=round(cubierto_ecuasanitas, 2),
                ganancia_fisio=round(ganancia_total, 2),
                ganancia_cobrada=round(ganancia_cobrada, 2),
                ganancia_pendiente=round(ganancia_pendiente, 2),
            )
        )

    return FisioDetalleOut(
        terapeutaid=terapeutaid,
        terapeuta=_nombre_usuario(terapeuta),
        desde=desde,
        hasta=hasta,
        pacientes=sorted(
            pacientes,
            key=lambda item: item.paciente,
        ),
        dias_sueldo=dias_sueldo,
    )


# -----------------------------------------------------------------------------
# Reporte semanal por clínicas / consultorios
# -----------------------------------------------------------------------------

@router.get("/clinicas-semanal", response_model=List[ClinicaSemanalOut])
def reporte_clinicas_semanal(
    desde: date,
    hasta: date,
    terapeutaid: Optional[int] = Query(None),
    consultorioid: Optional[int] = Query(None),
    dia_semana: Optional[int] = Query(None, ge=0, le=6),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    desde, hasta = _default_range(desde, hasta)
    dia_semana = _validar_dia_semana(dia_semana)

    sesiones_query = (
        db.query(SesionTerapia)
        .options(
            joinedload(SesionTerapia.paciente),
            joinedload(SesionTerapia.terapeuta),
            joinedload(SesionTerapia.tratamiento_paciente),
        )
        .filter(
            SesionTerapia.fecha.between(desde, hasta),
            SesionTerapia.horasalida != None,
            SesionTerapia.tratamientopacienteid != None,
        )
    )

    sesiones_query = _aplicar_filtro_dia_sesion(sesiones_query, dia_semana)

    sesiones_query = _aplicar_filtros_sesiones(
        sesiones_query,
        current_user,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
    )

    sesiones = sesiones_query.order_by(
        SesionTerapia.fecha,
        SesionTerapia.horaingreso,
        SesionTerapia.id,
    ).all()

    consultorios_map = _obtener_consultorios_map(db)

    pagos_gimnasio_rows = _pagos_gimnasio_por_consultorio(
        db=db,
        current_user=current_user,
        desde=desde,
        hasta=hasta,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
        dia_semana=dia_semana,
    )

    tratamiento_ids = {
        s.tratamientopacienteid
        for s in sesiones
        if s.tratamientopacienteid
    }

    disponible_pagado = _pagos_aplicados_a_rango_por_tratamiento(
        db,
        tratamiento_ids,
        desde,
        hasta,
    )

    data: Dict[Optional[int], Dict[str, float | int | str | None]] = {}
    sesiones_no_ecuasanitas: List[Tuple[Optional[int], int, float, date]] = []

    for sesion in sesiones:
        # En el reporte por clínica usamos el consultorio OPERATIVO:
        # el consultorio del fisioterapeuta que atendió la sesión.
        # Así, si un paciente de Centro Principal fue atendido por Atahualpa,
        # esa sesión cuenta para Atahualpa y no aparece como otra clínica
        # dentro del Excel del secretario de Atahualpa.
        consultorio_id = (
            sesion.terapeuta.consultorioid
            if getattr(sesion, "terapeuta", None) is not None
            else None
        )
        tratamiento_id = sesion.tratamientopacienteid
        precio = _precio_aplicado(sesion.tratamiento_paciente)

        item = data.setdefault(
            consultorio_id,
            {
                "consultorioid": consultorio_id,
                "consultorio": consultorios_map.get(
                    consultorio_id,
                    "Sin consultorio",
                ),
                "sesiones": 0,
                "total_generado": 0.0,
                "total_ecuasanitas": 0.0,
                "total_gimnasio_pagado": 0.0,
                "ganancia_fisios_terapia_total": 0.0,
                "ganancia_fisios_terapia_ecuasanitas": 0.0,
                "ganancia_clinica_terapia_total": 0.0,
                "ganancia_clinica_terapia_ecuasanitas": 0.0,
            },
        )

        ganancia_fisio = ganancia_fisio_terapia(precio, sesion.fecha)
        ganancia_clinica = ganancia_clinica_terapia(precio, sesion.fecha)

        item["sesiones"] = int(item["sesiones"]) + 1
        item["total_generado"] = float(item["total_generado"]) + precio
        item["ganancia_fisios_terapia_total"] = (
            float(item.get("ganancia_fisios_terapia_total", 0.0)) + ganancia_fisio
        )
        item["ganancia_clinica_terapia_total"] = (
            float(item.get("ganancia_clinica_terapia_total", 0.0)) + ganancia_clinica
        )

        if _es_paciente_ecuasanitas(sesion.paciente):
            item["total_ecuasanitas"] = (
                float(item.get("total_ecuasanitas", 0.0)) + precio
            )
            item["ganancia_fisios_terapia_ecuasanitas"] = (
                float(item.get("ganancia_fisios_terapia_ecuasanitas", 0.0)) + ganancia_fisio
            )
            item["ganancia_clinica_terapia_ecuasanitas"] = (
                float(item.get("ganancia_clinica_terapia_ecuasanitas", 0.0)) + ganancia_clinica
            )
        else:
            sesiones_no_ecuasanitas.append(
                (consultorio_id, tratamiento_id, precio, sesion.fecha)
            )

    for row in pagos_gimnasio_rows:
        consultorio_id = row.consultorioid
        total_gimnasio_pagado = float(row.total_pagado or 0)

        item = data.setdefault(
            consultorio_id,
            {
                "consultorioid": consultorio_id,
                "consultorio": consultorios_map.get(
                    consultorio_id,
                    "Sin consultorio",
                ),
                "sesiones": 0,
                "total_generado": 0.0,
                "total_ecuasanitas": 0.0,
                "total_gimnasio_pagado": 0.0,
                "ganancia_fisios_terapia_total": 0.0,
                "ganancia_fisios_terapia_ecuasanitas": 0.0,
                "ganancia_clinica_terapia_total": 0.0,
                "ganancia_clinica_terapia_ecuasanitas": 0.0,
            },
        )

        item["total_gimnasio_pagado"] = (
            float(item.get("total_gimnasio_pagado", 0.0)) + total_gimnasio_pagado
        )

    pagado_por_clinica: Dict[Optional[int], float] = {cid: 0.0 for cid in data.keys()}
    ganancia_fisios_cobrada_no_ecuasanitas: Dict[Optional[int], float] = {cid: 0.0 for cid in data.keys()}
    ganancia_fisios_pendiente: Dict[Optional[int], float] = {cid: 0.0 for cid in data.keys()}
    ganancia_clinica_cobrada_no_ecuasanitas: Dict[Optional[int], float] = {cid: 0.0 for cid in data.keys()}
    ganancia_clinica_pendiente: Dict[Optional[int], float] = {cid: 0.0 for cid in data.keys()}

    # Se distribuye el pago por sesión para respetar el porcentaje de cada fecha:
    # lunes-viernes 35%/65%, sábado-domingo 40%/60%.
    for consultorio_id, tratamiento_id, precio, fecha_sesion in sesiones_no_ecuasanitas:
        disponible = disponible_pagado.get(tratamiento_id, 0.0)
        aplicado = min(precio, disponible)
        pendiente = max(precio - aplicado, 0.0)
        porcentaje_fisio = porcentaje_fisio_terapia_por_fecha(fecha_sesion)
        porcentaje_clinica = porcentaje_clinica_terapia_por_fecha(fecha_sesion)

        pagado_por_clinica[consultorio_id] = (
            pagado_por_clinica.get(consultorio_id, 0.0) + aplicado
        )
        ganancia_fisios_cobrada_no_ecuasanitas[consultorio_id] = (
            ganancia_fisios_cobrada_no_ecuasanitas.get(consultorio_id, 0.0)
            + aplicado * porcentaje_fisio
        )
        ganancia_fisios_pendiente[consultorio_id] = (
            ganancia_fisios_pendiente.get(consultorio_id, 0.0)
            + pendiente * porcentaje_fisio
        )
        ganancia_clinica_cobrada_no_ecuasanitas[consultorio_id] = (
            ganancia_clinica_cobrada_no_ecuasanitas.get(consultorio_id, 0.0)
            + aplicado * porcentaje_clinica
        )
        ganancia_clinica_pendiente[consultorio_id] = (
            ganancia_clinica_pendiente.get(consultorio_id, 0.0)
            + pendiente * porcentaje_clinica
        )

        disponible_pagado[tratamiento_id] = max(disponible - aplicado, 0.0)

    resultado: List[ClinicaSemanalOut] = []

    for consultorio_id, item in data.items():
        total_terapia_generado = float(item["total_generado"])
        total_ecuasanitas = float(item.get("total_ecuasanitas", 0.0))
        total_terapia_pagado = float(pagado_por_clinica.get(consultorio_id, 0.0))
        total_no_ecuasanitas = max(total_terapia_generado - total_ecuasanitas, 0.0)
        pendiente_terapia = max(total_no_ecuasanitas - total_terapia_pagado, 0.0)
        total_gimnasio_pagado = float(item.get("total_gimnasio_pagado", 0.0))

        ganancia_fisios_terapia_total = float(item.get("ganancia_fisios_terapia_total", 0.0))
        ganancia_fisios_terapia_ecuasanitas = float(item.get("ganancia_fisios_terapia_ecuasanitas", 0.0))
        ganancia_fisios_terapia_cobrada = (
            ganancia_fisios_cobrada_no_ecuasanitas.get(consultorio_id, 0.0)
            + ganancia_fisios_terapia_ecuasanitas
        )
        ganancia_fisios_terapia_pendiente = ganancia_fisios_pendiente.get(consultorio_id, 0.0)

        ganancia_clinica_terapia_total = float(item.get("ganancia_clinica_terapia_total", 0.0))
        ganancia_clinica_terapia_ecuasanitas = float(item.get("ganancia_clinica_terapia_ecuasanitas", 0.0))
        ganancia_clinica_terapia_cobrada = (
            ganancia_clinica_cobrada_no_ecuasanitas.get(consultorio_id, 0.0)
            + ganancia_clinica_terapia_ecuasanitas
        )
        ganancia_clinica_terapia_pendiente = ganancia_clinica_pendiente.get(consultorio_id, 0.0)

        ganancia_fisios_gimnasio_cobrada = total_gimnasio_pagado * PORCENTAJE_FISIO_GIMNASIO
        ganancia_clinica_gimnasio_cobrada = total_gimnasio_pagado * PORCENTAJE_CLINICA_GIMNASIO

        resultado.append(
            ClinicaSemanalOut(
                consultorioid=consultorio_id,
                consultorio=str(item.get("consultorio") or "Sin consultorio"),
                sesiones_realizadas=int(item["sesiones"]),
                total_generado=round(total_terapia_generado, 2),
                total_pagado_pacientes=round(total_terapia_pagado, 2),
                total_pendiente_pacientes=round(pendiente_terapia, 2),
                total_ecuasanitas=round(total_ecuasanitas, 2),
                total_gimnasio_pagado=round(total_gimnasio_pagado, 2),
                ganancia_fisios_terapia_total=round(ganancia_fisios_terapia_total, 2),
                ganancia_fisios_terapia_cobrada=round(ganancia_fisios_terapia_cobrada, 2),
                ganancia_fisios_terapia_pendiente=round(ganancia_fisios_terapia_pendiente, 2),
                ganancia_fisios_terapia_ecuasanitas=round(ganancia_fisios_terapia_ecuasanitas, 2),
                ganancia_fisios_gimnasio_cobrada=round(ganancia_fisios_gimnasio_cobrada, 2),
                ganancia_clinica_terapia_total=round(ganancia_clinica_terapia_total, 2),
                ganancia_clinica_terapia_cobrada=round(ganancia_clinica_terapia_cobrada, 2),
                ganancia_clinica_terapia_pendiente=round(ganancia_clinica_terapia_pendiente, 2),
                ganancia_clinica_terapia_ecuasanitas=round(ganancia_clinica_terapia_ecuasanitas, 2),
                ganancia_clinica_gimnasio_cobrada=round(ganancia_clinica_gimnasio_cobrada, 2),
                ganancia_fisios_total=round(ganancia_fisios_terapia_total + ganancia_fisios_gimnasio_cobrada, 2),
                ganancia_fisios_cobrada=round(ganancia_fisios_terapia_cobrada + ganancia_fisios_gimnasio_cobrada, 2),
                ganancia_fisios_pendiente=round(ganancia_fisios_terapia_pendiente, 2),
                ganancia_clinica_total=round(ganancia_clinica_terapia_total + ganancia_clinica_gimnasio_cobrada, 2),
                ganancia_clinica_cobrada=round(ganancia_clinica_terapia_cobrada + ganancia_clinica_gimnasio_cobrada, 2),
                ganancia_clinica_pendiente=round(ganancia_clinica_terapia_pendiente, 2),
            )
        )

    return sorted(
        resultado,
        key=lambda item: item.total_generado,
        reverse=True,
    )


# -----------------------------------------------------------------------------
# Exportación profesional a Excel
# -----------------------------------------------------------------------------


def _excel_bool(value: bool | None) -> str:
    return "Sí" if bool(value) else "No"


def _excel_estado_pago(estado: int | None) -> str:
    return {
        1: "Pendiente",
        2: "Verificado",
        3: "Rechazado",
    }.get(int(estado or 0), "Sin estado")


def _excel_datetime_ecuador(value: Optional[datetime]) -> Optional[datetime]:
    if value is None:
        return None

    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)

    # openpyxl no acepta datetimes con zona horaria.
    return value.astimezone(ECUADOR_TZ).replace(tzinfo=None)


def _excel_fecha_desde_datetime(value: Optional[datetime]) -> Optional[date]:
    local_dt = _excel_datetime_ecuador(value)
    return local_dt.date() if local_dt else None


def _excel_hora(value) -> str:
    if value is None:
        return ""
    return value.strftime("%H:%M")


def _excel_dia(fecha_value: Optional[date]) -> str:
    if not fecha_value:
        return ""
    return DIAS_SEMANA[fecha_value.weekday()]


def _excel_safe_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _excel_nombre_persona(nombres: Optional[str], apellidos: Optional[str]) -> str:
    nombre = f"{nombres or ''} {apellidos or ''}".strip()
    return nombre or "Sin nombre"


def _excel_map_usuarios(db: Session, ids: Set[Optional[int]]) -> Dict[int, Usuario]:
    ids_limpios = {int(item) for item in ids if item is not None}
    if not ids_limpios:
        return {}
    return {u.id: u for u in db.query(Usuario).filter(Usuario.id.in_(ids_limpios)).all()}


def _excel_map_pacientes(db: Session, ids: Set[Optional[int]]) -> Dict[int, Paciente]:
    ids_limpios = {int(item) for item in ids if item is not None}
    if not ids_limpios:
        return {}
    return {p.id: p for p in db.query(Paciente).filter(Paciente.id.in_(ids_limpios)).all()}


def _excel_map_tratamientos(db: Session, ids: Set[Optional[int]]) -> Dict[int, TratamientoPaciente]:
    ids_limpios = {int(item) for item in ids if item is not None}
    if not ids_limpios:
        return {}
    return {
        t.id: t
        for t in db.query(TratamientoPaciente).filter(TratamientoPaciente.id.in_(ids_limpios)).all()
    }


def _excel_motivo_especial_5(tratamiento: Optional[TratamientoPaciente]) -> str:
    """Motivo visible para auditoría cuando una sesión paga $5 al fisio."""
    if not tratamiento:
        return ""

    if bool(getattr(tratamiento, "multiple_extremidad", False)):
        return "Más de una extremidad"

    nombre = _excel_safe_text(getattr(tratamiento, "tipotratamiento", ""))
    nombre_normalizado = _texto_normalizado(nombre)
    for nombre_especial in TERAPIAS_SUELDO_ESPECIAL_5:
        if nombre_especial in nombre_normalizado:
            return nombre or nombre_especial.title()

    return "Tarifa especial $5"


def _excel_motivo_tarifa_fisio(
    fecha_sesion: Optional[date],
    atenciones_dia: int,
    tratamiento: Optional[TratamientoPaciente],
) -> str:
    """Texto corto de la regla usada para calcular el sueldo de la sesión."""
    if _tratamiento_tarifa_especial_5(tratamiento):
        return f"Especial $5: {_excel_motivo_especial_5(tratamiento)}"

    if _es_fin_semana(fecha_sesion):
        return "Fin de semana $4"

    if atenciones_dia > UMBRAL_ATENCIONES_BONO_DIARIO:
        return f"Bono +{UMBRAL_ATENCIONES_BONO_DIARIO} atenciones $4"

    return "Normal lunes-viernes $3.50"


def _excel_estado_cobertura_sesion(
    aplicado: float,
    pendiente: float,
    precio: float,
    es_ecuasanitas: bool,
) -> str:
    if es_ecuasanitas:
        return "Ecuasanitas"
    if pendiente <= 0.009:
        return "Pagada"
    if aplicado <= 0.009:
        return "No pagada"
    if aplicado < precio:
        return "Parcial"
    return "Pagada"


def _excel_sesiones_sueldo_records(
    db: Session,
    current_user: Usuario,
    desde: date,
    hasta: date,
    terapeutaid: Optional[int] = None,
    consultorioid: Optional[int] = None,
    dia_semana: Optional[int] = None,
) -> List[Dict]:
    """Base única para auditar sesiones, deuda y sueldo del fisio en Excel.

    Usa el terapeuta REAL que atendió cada sesión (`SesionTerapia.terapeutaid`).
    Los pagos se aplican por FIFO real por paciente + tratamiento, igual que en
    los reportes de pantalla. Así un pago no se asigna completo al terapeuta que
    atendió un domingo si el tratamiento también tuvo sesiones con otros fisios.
    """
    sesiones = _sesiones_reporte_filtradas(
        db=db,
        current_user=current_user,
        desde=desde,
        hasta=hasta,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
        dia_semana=dia_semana,
    )

    cobertura_sesiones = _cobertura_sesiones_filtradas(db, sesiones, hasta)
    consultorios_map = _obtener_consultorios_map(db)

    atenciones_por_terapeuta_dia: Dict[Tuple[Optional[int], date], int] = defaultdict(int)
    for sesion in sesiones:
        if sesion.fecha is not None:
            atenciones_por_terapeuta_dia[(sesion.terapeutaid, sesion.fecha)] += 1

    records: List[Dict] = []
    for sesion in sesiones:
        paciente = sesion.paciente
        terapeuta = sesion.terapeuta
        tratamiento = sesion.tratamiento_paciente
        precio = round(float(_precio_aplicado(tratamiento) or 0.0), 2)
        es_ecuasanitas = _es_paciente_ecuasanitas(paciente)
        atenciones_dia = atenciones_por_terapeuta_dia.get((sesion.terapeutaid, sesion.fecha), 0)
        es_especial_5 = _tratamiento_tarifa_especial_5(tratamiento)
        tarifa_fisio = round(float(sueldo_fisio_terapia(sesion.fecha, atenciones_dia, es_especial_5)), 2)
        aplicado, pendiente = cobertura_sesiones.get(
            int(sesion.id),
            (0.0, precio),
        ) if sesion.id is not None else (0.0, precio)

        aplicado = round(float(aplicado or 0.0), 2)
        pendiente = round(float(pendiente or 0.0), 2)

        # Ecuasanitas no aparece como deuda del paciente dentro del reporte de terapias.
        if es_ecuasanitas:
            aplicado = precio
            pendiente = 0.0

        sueldo_cobrado = tarifa_fisio if pendiente <= 0.009 else 0.0
        sueldo_pendiente = 0.0 if pendiente <= 0.009 else tarifa_fisio
        motivo_especial = _excel_motivo_especial_5(tratamiento) if es_especial_5 else ""

        records.append(
            {
                "sesion_id": int(sesion.id) if sesion.id is not None else None,
                "fecha": sesion.fecha,
                "dia": _excel_dia(sesion.fecha),
                "paciente": _nombre_paciente(paciente),
                "cedula": _excel_safe_text(getattr(paciente, "cedula", "")),
                "fisioterapeuta": _nombre_usuario(terapeuta),
                "consultorio_atencion": consultorios_map.get(
                    getattr(terapeuta, "consultorioid", None),
                    "Sin consultorio",
                ),
                "consultorio_paciente": consultorios_map.get(
                    getattr(paciente, "consultorioid", None),
                    "Sin consultorio",
                ),
                "tratamiento": _excel_safe_text(getattr(tratamiento, "tipotratamiento", "Sin tratamiento")),
                "tratamiento_id": getattr(tratamiento, "id", None),
                "hora_ingreso": _excel_hora(sesion.horaingreso),
                "hora_salida": _excel_hora(sesion.horasalida),
                "duracion_min": sesion.duracionminutos or 0,
                "dolor_entrada": sesion.escaladolorentrada,
                "dolor_salida": sesion.escaladolorsalida,
                "precio": precio,
                "ecuasanitas": es_ecuasanitas,
                "atenciones_dia": int(atenciones_dia or 0),
                "fin_semana": _es_fin_semana(sesion.fecha),
                "bono_15": bool(
                    sesion.fecha
                    and not _es_fin_semana(sesion.fecha)
                    and atenciones_dia > UMBRAL_ATENCIONES_BONO_DIARIO
                ),
                "especial_5": es_especial_5,
                "motivo_especial": motivo_especial,
                "tarifa_fisio": tarifa_fisio,
                "motivo_tarifa": _excel_motivo_tarifa_fisio(sesion.fecha, atenciones_dia, tratamiento),
                "pago_aplicado": aplicado,
                "debe_paciente": pendiente,
                "estado_cobertura": _excel_estado_cobertura_sesion(aplicado, pendiente, precio, es_ecuasanitas),
                "sueldo_total": tarifa_fisio,
                "sueldo_cobrado": round(float(sueldo_cobrado), 2),
                "sueldo_retenido": round(float(sueldo_pendiente), 2),
                "ganancia_clinica": round(max(precio - tarifa_fisio, 0.0), 2),
            }
        )

    return records


def _excel_detalle_sueldos_fisios(
    db: Session,
    current_user: Usuario,
    desde: date,
    hasta: date,
    terapeutaid: Optional[int] = None,
    consultorioid: Optional[int] = None,
    dia_semana: Optional[int] = None,
) -> List[List]:
    records = _excel_sesiones_sueldo_records(
        db=db,
        current_user=current_user,
        desde=desde,
        hasta=hasta,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
        dia_semana=dia_semana,
    )

    rows: List[List] = []
    for item in records:
        rows.append(
            [
                item["fecha"],
                item["dia"],
                item["sesion_id"],
                item["fisioterapeuta"],
                item["consultorio_atencion"],
                item["paciente"],
                item["cedula"],
                item["tratamiento"],
                item["tratamiento_id"],
                item["precio"],
                item["pago_aplicado"],
                item["debe_paciente"],
                item["estado_cobertura"],
                item["atenciones_dia"],
                _excel_bool(item["bono_15"]),
                _excel_bool(item["fin_semana"]),
                item["tarifa_fisio"],
                item["motivo_tarifa"],
                _excel_bool(item["especial_5"]),
                item["motivo_especial"],
                item["sueldo_total"],
                item["sueldo_cobrado"],
                item["sueldo_retenido"],
                item["consultorio_paciente"],
            ]
        )

    return rows


def _excel_base_sesiones(
    db: Session,
    current_user: Usuario,
    desde: date,
    hasta: date,
    terapeutaid: Optional[int] = None,
    consultorioid: Optional[int] = None,
    dia_semana: Optional[int] = None,
) -> List[List]:
    records = _excel_sesiones_sueldo_records(
        db=db,
        current_user=current_user,
        desde=desde,
        hasta=hasta,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
        dia_semana=dia_semana,
    )

    rows: List[List] = []
    for item in records:
        rows.append(
            [
                item["sesion_id"],
                item["fecha"],
                item["dia"],
                item["paciente"],
                item["cedula"],
                item["fisioterapeuta"],
                item["consultorio_atencion"],
                item["consultorio_paciente"],
                item["tratamiento"],
                item["tratamiento_id"],
                item["hora_ingreso"],
                item["hora_salida"],
                item["duracion_min"],
                item["dolor_entrada"],
                item["dolor_salida"],
                item["precio"],
                _excel_bool(item["ecuasanitas"]),
                item["pago_aplicado"],
                item["debe_paciente"],
                item["estado_cobertura"],
                item["tarifa_fisio"],
                item["motivo_tarifa"],
                _excel_bool(item["especial_5"]),
                item["motivo_especial"],
                item["sueldo_total"],
                item["sueldo_cobrado"],
                item["sueldo_retenido"],
                item["ganancia_clinica"],
            ]
        )

    return rows


def _excel_base_pagos(
    db: Session,
    current_user: Usuario,
    desde: date,
    hasta: date,
    terapeutaid: Optional[int] = None,
    consultorioid: Optional[int] = None,
    dia_semana: Optional[int] = None,
) -> List[List]:
    consultorios_map = _obtener_consultorios_map(db)
    rows: List[List] = []

    # Pagos de terapias: incluye verificados, pendientes, rechazados, previos y anulados
    # para auditoría. La columna "Caja válida" separa lo que sí entra al cuadre.
    pagos_terapia_query = db.query(Pago).filter(
        Pago.tratamientopacienteid != None,
        filtro_fechapago_ecuador(desde, hasta),
        filtro_dia_pago_ecuador(dia_semana),
    )
    pagos_terapia_query = _aplicar_filtros_pagos(
        pagos_terapia_query,
        current_user,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
    )
    pagos_terapia = pagos_terapia_query.order_by(Pago.fechapago, Pago.id).all()

    tratamiento_ids_pagos = {
        p.tratamientopacienteid
        for p in pagos_terapia
        if p.tratamientopacienteid is not None
    }

    # Importante:
    # Los pagos de un tratamiento NO se asignan a un fisioterapeuta operativo
    # tomando la primera sesión del rango. Ese era el origen del error donde,
    # si un paciente se atendía un domingo con otro terapeuta, todos los pagos
    # del tratamiento aparecían a nombre de ese terapeuta.
    #
    # Base_Pagos muestra el terapeuta encargado/asignado del paciente. La
    # distribución real por sesión atendida se audita en la hoja
    # "Sueldos_Fisios", donde cada pago se aplica por FIFO a la sesión real.

    tratamiento_map = _excel_map_tratamientos(
        db,
        tratamiento_ids_pagos,
    )
    pacientes_map = _excel_map_pacientes(
        db,
        {p.pacienteid for p in pagos_terapia}
        | {t.pacienteid for t in tratamiento_map.values() if t is not None},
    )
    usuarios_map = _excel_map_usuarios(
        db,
        {getattr(pacientes_map.get(p.pacienteid), "terapeutaasignadoid", None) for p in pagos_terapia}
        | {p.creado_por_id for p in pagos_terapia},
    )

    def add_pago_row(
        pago: Pago,
        paciente: Optional[Paciente],
        responsable: str,
        consultorio_nombre: str,
        tipo: str,
        referencia: str,
        observacion: str = "",
    ) -> None:
        fecha_pago = _excel_fecha_desde_datetime(pago.fechapago)
        caja_valida = round(float(pago.monto or 0), 2) if (
            pago.estadopago == 2
            and not bool(pago.espagoprevio)
            and not bool(pago.anulado)
            and not _es_metodo_sin_caja(pago.metodopago)
        ) else 0.0
        rows.append(
            [
                pago.id,
                fecha_pago,
                _excel_dia(fecha_pago),
                _nombre_paciente(paciente),
                _excel_safe_text(getattr(paciente, "cedula", "")),
                responsable,
                consultorio_nombre,
                tipo,
                _excel_safe_text(pago.metodopago) or "Sin método",
                _excel_estado_pago(pago.estadopago),
                round(float(pago.monto or 0), 2),
                _excel_bool(pago.espagoprevio),
                _excel_bool(pago.esrecuperacioncartera),
                _excel_bool(pago.anulado),
                referencia,
                observacion,
                pago.creado_por_id,
                caja_valida,
            ]
        )

    for pago in pagos_terapia:
        tratamiento = tratamiento_map.get(pago.tratamientopacienteid)
        paciente = pacientes_map.get(pago.pacienteid) or pacientes_map.get(getattr(tratamiento, "pacienteid", None))
        terapeuta = usuarios_map.get(getattr(paciente, "terapeutaasignadoid", None))
        responsable_pago = _nombre_usuario(terapeuta)
        consultorio_pago = consultorios_map.get(getattr(paciente, "consultorioid", None), "Sin consultorio")

        referencia = _excel_safe_text(getattr(tratamiento, "tipotratamiento", "Terapia"))
        observacion = _excel_safe_text(
            pago.observacionpagoprevio
            or (pago.numerocomprobante if _es_metodo_sin_caja(pago.metodopago) else None)
            or pago.motivo_rechazo
            or pago.motivo_anulacion
        )
        add_pago_row(
            pago=pago,
            paciente=paciente,
            responsable=responsable_pago,
            consultorio_nombre=consultorio_pago,
            tipo="Terapia",
            referencia=referencia,
            observacion=observacion,
        )

    # Pagos de gimnasio mensual / diario. Gimnasio no se cubre por Ecuasanitas.
    consultorio_resuelto = _resolver_consultorioid_gimnasio_para_rol(current_user, consultorioid)
    responsable_expr = _gimnasio_responsable_expr()
    consultorio_expr = _gimnasio_consultorio_expr()
    pagos_gimnasio_query = (
        db.query(Pago, MembresiaGimnasio)
        .join(MembresiaGimnasio, MembresiaGimnasio.id == Pago.membresiagimnasioid)
        .join(Paciente, Paciente.id == MembresiaGimnasio.pacienteid)
        .filter(
            Pago.membresiagimnasioid != None,
            filtro_fechapago_ecuador(desde, hasta),
            filtro_dia_pago_ecuador(dia_semana),
        )
    )

    if current_user.rol == 2:
        pagos_gimnasio_query = pagos_gimnasio_query.filter(responsable_expr == current_user.id)
    elif terapeutaid is not None:
        pagos_gimnasio_query = pagos_gimnasio_query.filter(responsable_expr == terapeutaid)

    if consultorio_resuelto is not None:
        pagos_gimnasio_query = pagos_gimnasio_query.filter(consultorio_expr == consultorio_resuelto)

    pagos_gimnasio_rows = pagos_gimnasio_query.order_by(Pago.fechapago, Pago.id).all()
    pacientes_gym = _excel_map_pacientes(db, {pago.pacienteid for pago, _ in pagos_gimnasio_rows})
    usuarios_gym = _excel_map_usuarios(
        db,
        {
            membresia.responsablegimnasioid
            or getattr(pacientes_gym.get(pago.pacienteid), "terapeutaasignadoid", None)
            for pago, membresia in pagos_gimnasio_rows
        }
        | {pago.creado_por_id for pago, _ in pagos_gimnasio_rows},
    )

    for pago, membresia in pagos_gimnasio_rows:
        paciente = pacientes_gym.get(pago.pacienteid)
        responsable_id = membresia.responsablegimnasioid or getattr(paciente, "terapeutaasignadoid", None)
        terapeuta = usuarios_gym.get(responsable_id)
        consultorio_pago_id = membresia.consultorioid or getattr(paciente, "consultorioid", None)
        observacion = _excel_safe_text(
            pago.observacionpagoprevio
            or (pago.numerocomprobante if _es_metodo_sin_caja(pago.metodopago) else None)
            or pago.motivo_rechazo
            or pago.motivo_anulacion
        )
        add_pago_row(
            pago=pago,
            paciente=paciente,
            responsable=_nombre_usuario(terapeuta),
            consultorio_nombre=consultorios_map.get(consultorio_pago_id, "Sin consultorio"),
            tipo="Gimnasio",
            referencia=f"Membresía/Pase #{pago.membresiagimnasioid}",
            observacion=observacion,
        )

    # Recuperación de cartera: entra a caja, pero no reduce saldos ni genera comisión automática.
    recuperacion_query = _query_recuperacion_cartera(
        db=db,
        current_user=current_user,
        desde=desde,
        hasta=hasta,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
        dia_semana=dia_semana,
    )
    pagos_recuperacion = recuperacion_query.order_by(Pago.fechapago, Pago.id).all()
    pacientes_rec = _excel_map_pacientes(db, {p.pacienteid for p in pagos_recuperacion})
    cobradores_rec = _excel_map_usuarios(db, {p.creado_por_id for p in pagos_recuperacion})

    for pago in pagos_recuperacion:
        paciente = pacientes_rec.get(pago.pacienteid)
        cobrador = cobradores_rec.get(pago.creado_por_id)
        add_pago_row(
            pago=pago,
            paciente=paciente,
            responsable=f"Cobrador: {_nombre_usuario(cobrador)}",
            consultorio_nombre=consultorios_map.get(getattr(cobrador, "consultorioid", None), "Sin consultorio"),
            tipo="Recuperación cartera",
            referencia="Cobro anterior al sistema",
            observacion=_excel_safe_text(pago.observacion_cartera),
        )

    rows.sort(key=lambda row: (row[1] or date.min, row[0] or 0))
    return rows


def _excel_aplicar_estilo_workbook(wb) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
    from openpyxl.utils import get_column_letter

    azul = "1F4E78"
    azul_claro = "D9EAF7"
    gris = "F4F6F8"
    verde = "E2F0D9"
    blanco = "FFFFFF"
    borde = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center")
                if cell.value is not None:
                    cell.border = borde

        for cell in ws[1]:
            if cell.value:
                cell.font = Font(bold=True, color=blanco, size=12)
                cell.fill = PatternFill("solid", fgColor=azul)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Encabezados de tablas/secciones.
        for row in ws.iter_rows():
            first = row[0].value if row else None
            if isinstance(first, str) and first.startswith("▶"):
                for cell in row:
                    if cell.value:
                        cell.font = Font(bold=True, color="12355B", size=12)
                        cell.fill = PatternFill("solid", fgColor=azul_claro)

        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            letter = get_column_letter(col_idx)
            for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row):
                for item in cell:
                    value = item.value
                    if value is None:
                        continue
                    text = str(value)
                    max_len = max(max_len, len(text))
            width = min(max(max_len + 2, 10), 38)
            if letter in {"P"} and ws.title == "Base_Pagos":
                width = 42
            ws.column_dimensions[letter].width = width

        for row in ws.iter_rows():
            for cell in row:
                header = str(ws.cell(row=1, column=cell.column).value or "").lower()
                if any(word in header for word in ["monto", "total", "generado", "pagado", "pendiente", "ganancia", "caja", "ecuasanitas", "precio", "saldo", "tarifa", "sueldo", "debe", "retenido"]):
                    cell.number_format = '$#,##0.00;[Red]-$#,##0.00'
                if "porcentaje" in header:
                    cell.number_format = "0%"
                if "fecha" in header:
                    cell.number_format = "dd/mm/yyyy"

        if ws.max_row > 1:
            for cell in ws[1]:
                if cell.value:
                    cell.fill = PatternFill("solid", fgColor=azul)
                    cell.font = Font(bold=True, color=blanco)

    # Estilo específico del dashboard.
    ws = wb["Dashboard"]
    ws.sheet_view.showGridLines = False
    for row in range(1, 40):
        ws.row_dimensions[row].height = 24
    for col in range(1, 11):
        ws.column_dimensions[chr(64 + col)].width = 18

    for rng in ["A1:J2", "A4:C10", "A13:C13", "D4:F6", "G4:I6", "D8:F10", "G8:I10", "D12:F14", "G12:I14"]:
        for row in ws[rng]:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=gris)
                cell.border = borde
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws["A1"].fill = PatternFill("solid", fgColor=azul)
    ws["A1"].font = Font(bold=True, color=blanco, size=16)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    for cell_ref in ["A4", "D4", "G4", "D8", "G8", "D12", "G12", "A13"]:
        ws[cell_ref].font = Font(bold=True, color="12355B", size=11)
    for cell_ref in ["B5", "B6", "B7", "D5", "G5", "D9", "G9", "D13", "G13", "B13"]:
        ws[cell_ref].font = Font(bold=True, color="12355B", size=18 if cell_ref not in {"B5", "B6", "B7"} else 12)
        ws[cell_ref].fill = PatternFill("solid", fgColor=verde)
    ws["A4"].fill = PatternFill("solid", fgColor=azul)
    ws["A4"].font = Font(bold=True, color=blanco, size=12)


def _excel_agregar_tabla(ws, start_row: int, start_col: int, headers: List[str], rows: List[List], table_name: str, crear_tabla: bool = True) -> int:
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    header_row = start_row
    for idx, header in enumerate(headers, start=start_col):
        ws.cell(row=header_row, column=idx, value=header)

    data_rows = rows if rows else [["Sin datos"] + [None] * (len(headers) - 1)]
    for r_idx, row_values in enumerate(data_rows, start=header_row + 1):
        for c_idx, value in enumerate(row_values, start=start_col):
            ws.cell(row=r_idx, column=c_idx, value=value)

    end_row = header_row + len(data_rows)
    end_col = start_col + len(headers) - 1
    ref = f"{get_column_letter(start_col)}{header_row}:{get_column_letter(end_col)}{end_row}"
    if crear_tabla:
        tab = Table(displayName=table_name, ref=ref)
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tab)
        # IMPORTANTE:
        # No usar ws.auto_filter.ref cuando la hoja tiene tablas de Excel.
        # Cada Table ya incluye su propio autofiltro. En Microsoft Excel, combinar
        # filtros de hoja con tablas puede generar archivos .xlsx que se abren con
        # el mensaje: "Hemos encontrado un problema con contenido...".

    currency_words = [
        "monto",
        "total",
        "generado",
        "pagado",
        "pendiente",
        "ganancia",
        "caja",
        "ecuasanitas",
        "precio",
        "saldo",
        "tarifa",
        "sueldo",
        "debe",
        "retenido",
    ]
    integer_words = ["sesiones", "duración", "dolor"]

    for offset, header in enumerate(headers):
        col_idx = start_col + offset
        header_lower = header.lower()
        for row_idx in range(header_row + 1, end_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if "fecha" in header_lower:
                cell.number_format = "dd/mm/yyyy"
            elif any(word in header_lower for word in currency_words):
                cell.number_format = '$#,##0.00;[Red]-$#,##0.00'
            elif any(word in header_lower for word in integer_words):
                cell.number_format = "0"

    return end_row + 3


def _excel_escribir_titulo(ws, titulo: str, subtitulo: str) -> None:
    ws.merge_cells("A1:J1")
    ws["A1"] = titulo
    ws.merge_cells("A2:J2")
    ws["A2"] = subtitulo


def _excel_colocar_kpi(ws, cell: str, titulo: str, valor, detalle: str) -> None:
    row = ws[cell].row
    col = ws[cell].column
    ws.cell(row=row, column=col, value=titulo)
    ws.cell(row=row + 1, column=col, value=valor)
    ws.cell(row=row + 2, column=col, value=detalle)
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 2)
    ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 2)
    ws.merge_cells(start_row=row + 2, start_column=col, end_row=row + 2, end_column=col + 2)


def _excel_unicos_con_todos(valores: List[str]) -> List[str]:
    limpios = sorted({str(v).strip() for v in valores if str(v or "").strip()})
    return ["Todos"] + limpios


def _excel_crear_listas_filtros(wb, base_sesiones: List[List], base_pagos: List[List]) -> Dict[str, str]:
    """Crea listas ocultas para los desplegables del panel central de filtros."""
    ws = wb.create_sheet("Listas")

    clinicas = _excel_unicos_con_todos(
        [row[6] for row in base_sesiones if len(row) > 6]
        + [row[6] for row in base_pagos if len(row) > 6]
    )
    fisios = _excel_unicos_con_todos(
        [row[5] for row in base_sesiones if len(row) > 5]
        + [row[5] for row in base_pagos if len(row) > 5]
    )
    # Semana CORPOFIT: domingo a sábado.
    dias = ["Todos", "Domingo", "Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"]

    listas = {
        "clinicas": clinicas,
        "fisios": fisios,
        "dias": dias,
    }
    columnas = {
        "clinicas": 1,
        "fisios": 2,
        "dias": 3,
    }
    titulos = {
        "clinicas": "Clínicas",
        "fisios": "Fisioterapeutas",
        "dias": "Días",
    }

    rangos: Dict[str, str] = {}
    for nombre, valores in listas.items():
        col = columnas[nombre]
        ws.cell(row=1, column=col, value=titulos[nombre])
        for row_idx, value in enumerate(valores, start=2):
            ws.cell(row=row_idx, column=col, value=value)
        col_letter = chr(64 + col)
        rangos[nombre] = f"Listas!${col_letter}$2:${col_letter}${len(valores) + 1}"

    ws.sheet_state = "hidden"
    return rangos



def _excel_dashboard_helper_criterios(ws) -> None:
    """
    Crea criterios auxiliares simples para los filtros del Dashboard.

    IMPORTANTE:
    Antes se usaba SUMPRODUCT con condiciones booleanas. En algunas versiones de
    Excel/WPS esos arreglos daban #VALOR! al combinar tablas, textos y filtros.
    Con SUMIFS/COUNTIFS + criterios auxiliares el archivo es mucho más estable.
    """
    ws["K4"] = "Criterios internos"
    ws["K5"] = '=IF($B$5="Todos","*",$B$5)'
    ws["K6"] = '=IF($B$6="Todos","*",$B$6)'
    ws["K7"] = '=IF($B$7="Todos","*",$B$7)'
    ws.column_dimensions["K"].hidden = True


def _excel_formula_count_sesiones(clinica_criteria: str = "$K$5", fisio_criteria: str = "$K$6", dia_criteria: str = "$K$7") -> str:
    return (
        f'=COUNTIFS('
        f'tblBaseSesiones[ID Sesión],">=0",'
        f'tblBaseSesiones[Clínica / Consultorio],{clinica_criteria},'
        f'tblBaseSesiones[Fisioterapeuta],{fisio_criteria},'
        f'tblBaseSesiones[Día],{dia_criteria}'
        f')'
    )


def _excel_formula_sum_sesiones(
    columna_sumar: str,
    clinica_criteria: str = "$K$5",
    fisio_criteria: str = "$K$6",
    dia_criteria: str = "$K$7",
    extra_criteria: str = "",
) -> str:
    extra = f",{extra_criteria}" if extra_criteria else ""
    return (
        f'=SUMIFS('
        f'tblBaseSesiones[{columna_sumar}],'
        f'tblBaseSesiones[Clínica / Consultorio],{clinica_criteria},'
        f'tblBaseSesiones[Fisioterapeuta],{fisio_criteria},'
        f'tblBaseSesiones[Día],{dia_criteria}'
        f'{extra}'
        f')'
    )


def _excel_formula_sum_pagos(
    columna_sumar: str,
    clinica_criteria: str = "$K$5",
    fisio_criteria: str = "$K$6",
    dia_criteria: str = "$K$7",
    extra_criteria: str = "",
) -> str:
    extra = f",{extra_criteria}" if extra_criteria else ""
    return (
        f'=SUMIFS('
        f'tblBasePagos[{columna_sumar}],'
        f'tblBasePagos[Clínica / Consultorio],{clinica_criteria},'
        f'tblBasePagos[Fisioterapeuta / Responsable],{fisio_criteria},'
        f'tblBasePagos[Día],{dia_criteria}'
        f'{extra}'
        f')'
    )


def _excel_configurar_panel_filtros_dashboard(ws, rangos: Dict[str, str], desde: date, hasta: date) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
    from openpyxl.worksheet.datavalidation import DataValidation

    azul = "1F4E78"
    azul_claro = "D9EAF7"
    verde = "E2F0D9"
    gris = "F4F6F8"
    blanco = "FFFFFF"
    borde = Border(
        left=Side(style="thin", color="B7C9D6"),
        right=Side(style="thin", color="B7C9D6"),
        top=Side(style="thin", color="B7C9D6"),
        bottom=Side(style="thin", color="B7C9D6"),
    )

    ws.merge_cells("A4:C4")
    ws["A4"] = "CAJITA DE FILTROS"
    ws["A4"].font = Font(bold=True, color=blanco, size=12)
    ws["A4"].fill = PatternFill("solid", fgColor=azul)
    ws["A4"].alignment = Alignment(horizontal="center", vertical="center")

    labels = [
        ("A5", "Clínica / Consultorio", "B5", "Todos", "clinicas"),
        ("A6", "Fisioterapeuta", "B6", "Todos", "fisios"),
        ("A7", "Día de semana", "B7", "Todos", "dias"),
    ]
    for label_cell, label, value_cell, default, rango_key in labels:
        ws[label_cell] = label
        ws[value_cell] = default
        ws[label_cell].font = Font(bold=True, color="12355B")
        ws[value_cell].fill = PatternFill("solid", fgColor=verde)
        ws[value_cell].font = Font(bold=True, color="12355B")
        ws[value_cell].alignment = Alignment(horizontal="center", vertical="center")
        dv = DataValidation(type="list", formula1=rangos[rango_key], allow_blank=False)
        dv.error = "Selecciona una opción válida de la lista."
        dv.errorTitle = "Filtro no válido"
        ws.add_data_validation(dv)
        dv.add(ws[value_cell])

    ws["A9"] = "Periodo"
    ws["B9"] = f"{desde.strftime('%d/%m/%Y')} - {hasta.strftime('%d/%m/%Y')}"
    ws["A10"] = "Uso"
    ws["B10"] = "Cambia los 3 desplegables: KPIs y gráficas se recalculan."
    ws["B10"].alignment = Alignment(wrap_text=True, vertical="center")

    for rng in ["A4:C10"]:
        for row in ws[rng]:
            for cell in row:
                cell.border = borde
                if cell.row >= 5 and cell.column != 2:
                    cell.fill = PatternFill("solid", fgColor=gris)

    _excel_dashboard_helper_criterios(ws)

    formulas = {
        "sesiones": _excel_formula_count_sesiones(),
        "generado": _excel_formula_sum_sesiones("Precio sesión"),
        "pagado": _excel_formula_sum_pagos("Caja válida"),
        "ecuasanitas": _excel_formula_sum_sesiones(
            "Precio sesión",
            extra_criteria='tblBaseSesiones[Ecuasanitas],"Sí"',
        ),
        "ganancia_fisio": _excel_formula_sum_sesiones("Ganancia fisio"),
        "ganancia_clinica": _excel_formula_sum_sesiones("Ganancia clínica"),
        "pendiente": _excel_formula_sum_sesiones("Debe paciente"),
    }

    _excel_colocar_kpi(ws, "D4", "Sesiones filtradas", formulas["sesiones"], "Según clínica, fisio y día")
    _excel_colocar_kpi(ws, "G4", "Generado terapias", formulas["generado"], "Valor producido")
    _excel_colocar_kpi(ws, "D8", "Pagado caja", formulas["pagado"], "Cobros verificados")
    _excel_colocar_kpi(ws, "G8", "Ecuasanitas", formulas["ecuasanitas"], "Convenio terapias")
    _excel_colocar_kpi(ws, "D12", "Sueldo fisio", formulas["ganancia_fisio"], "$3.50 / $4.00 / $5.00 según regla")
    _excel_colocar_kpi(ws, "G12", "Ganancia clínica", formulas["ganancia_clinica"], "Precio sesión - sueldo fisio")

    ws["A13"] = "Pacientes deben"
    ws["B13"] = formulas["pendiente"]
    ws["C13"] = "Suma real por sesión no cubierta"
    ws["A13"].font = Font(bold=True, color="12355B")
    ws["B13"].font = Font(bold=True, color="12355B", size=14)
    ws["B13"].fill = PatternFill("solid", fgColor=verde)
    ws["C13"].alignment = Alignment(wrap_text=True)

    for cell_ref in ["G5", "D9", "G9", "D13", "G13", "B13"]:
        ws[cell_ref].number_format = '$#,##0.00;[Red]-$#,##0.00'
    ws["D5"].number_format = "0"

    ws["A16"] = "Importante"
    ws["A16"].font = Font(bold=True, color="12355B", size=12)
    ws["A17"] = "• Esta hoja es el panel principal: KPIs y gráficas obedecen a la cajita de filtros."
    ws["A18"] = "• Base_Sesiones y Base_Pagos conservan todos los datos de la semana para auditoría."
    ws["A19"] = "• Si necesitas revisar registros exactos, usa los filtros nativos de esas tablas base."
    ws["A20"] = "• Fisioterapeutas: lunes a viernes 35%; sábado y domingo 40%."
    ws["A21"] = "• Si cambias un filtro y no se actualiza, presiona F9 o guarda y vuelve a abrir el archivo."


def _excel_formula_sesiones_dia(dia_cell: str, columna_sumar: str) -> str:
    return (
        f'=IF(AND($B$7<>"Todos",{dia_cell}<>$B$7),0,'
        f'SUMIFS('
        f'tblBaseSesiones[{columna_sumar}],'
        f'tblBaseSesiones[Clínica / Consultorio],$K$5,'
        f'tblBaseSesiones[Fisioterapeuta],$K$6,'
        f'tblBaseSesiones[Día],{dia_cell}'
        f'))'
    )


def _excel_formula_count_sesiones_dia(dia_cell: str) -> str:
    return (
        f'=IF(AND($B$7<>"Todos",{dia_cell}<>$B$7),0,'
        f'COUNTIFS('
        f'tblBaseSesiones[ID Sesión],">=0",'
        f'tblBaseSesiones[Clínica / Consultorio],$K$5,'
        f'tblBaseSesiones[Fisioterapeuta],$K$6,'
        f'tblBaseSesiones[Día],{dia_cell}'
        f'))'
    )


def _excel_formula_pagos_dia(dia_cell: str, columna_sumar: str) -> str:
    return (
        f'=IF(AND($B$7<>"Todos",{dia_cell}<>$B$7),0,'
        f'SUMIFS('
        f'tblBasePagos[{columna_sumar}],'
        f'tblBasePagos[Clínica / Consultorio],$K$5,'
        f'tblBasePagos[Fisioterapeuta / Responsable],$K$6,'
        f'tblBasePagos[Día],{dia_cell}'
        f'))'
    )


def _excel_formula_pagos_metodo(metodo_cell: str) -> str:
    return (
        f'=SUMIFS('
        f'tblBasePagos[Caja válida],'
        f'tblBasePagos[Clínica / Consultorio],$K$5,'
        f'tblBasePagos[Fisioterapeuta / Responsable],$K$6,'
        f'tblBasePagos[Día],$K$7,'
        f'tblBasePagos[Método],{metodo_cell}'
        f')'
    )


def _excel_crear_graficas_dashboard(ws, base_pagos: List[List]) -> None:
    """
    Crea gráficas vinculadas a la cajita de filtros del Dashboard.

    No usa FILTER(), SORT(), UNIQUE() ni SUMPRODUCT. Las gráficas leen tablas
    auxiliares con SUMIFS/COUNTIFS para evitar #VALOR! en Excel/WPS.
    """
    from openpyxl.chart import BarChart, Reference, PieChart
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    azul = "1F4E78"
    azul_claro = "D9EAF7"
    blanco = "FFFFFF"
    gris = "F4F6F8"
    borde = Border(
        left=Side(style="thin", color="D9E2EA"),
        right=Side(style="thin", color="D9E2EA"),
        top=Side(style="thin", color="D9E2EA"),
        bottom=Side(style="thin", color="D9E2EA"),
    )

    # Tablas auxiliares visibles debajo de las gráficas. Son simples y compatibles.
    start_day = 50
    ws.cell(row=start_day - 1, column=1, value="Datos para gráficas vinculadas a filtros")
    ws.cell(row=start_day - 1, column=1).font = Font(bold=True, color="12355B", size=12)

    day_headers = ["Día", "Generado terapias", "Pagado caja", "Sesiones"]
    for col_idx, header in enumerate(day_headers, start=1):
        cell = ws.cell(row=start_day, column=col_idx, value=header)
        cell.fill = PatternFill("solid", fgColor=azul)
        cell.font = Font(bold=True, color=blanco)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borde

    dias = ["Domingo", "Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"]
    for offset, dia in enumerate(dias, start=1):
        row = start_day + offset
        dia_ref = f"$A${row}"

        ws.cell(row=row, column=1, value=dia)
        ws.cell(row=row, column=2, value=_excel_formula_sesiones_dia(dia_ref, "Precio sesión"))
        ws.cell(row=row, column=3, value=_excel_formula_pagos_dia(dia_ref, "Caja válida"))
        ws.cell(row=row, column=4, value=_excel_formula_count_sesiones_dia(dia_ref))
        for col_idx in range(1, 5):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = borde
            if col_idx in {2, 3}:
                cell.number_format = '$#,##0.00;[Red]-$#,##0.00'
            elif col_idx == 4:
                cell.number_format = "0"

    # Métodos de pago reales encontrados en la semana. Si no hay, se deja una fila neutra.
    metodos = sorted({str(row[8]).strip() for row in base_pagos if len(row) > 8 and str(row[8] or "").strip()})
    if not metodos:
        metodos = ["Sin pagos"]

    start_method_col = 6  # F:G
    method_headers = ["Método", "Caja válida"]
    for col_offset, header in enumerate(method_headers):
        cell = ws.cell(row=start_day, column=start_method_col + col_offset, value=header)
        cell.fill = PatternFill("solid", fgColor=azul)
        cell.font = Font(bold=True, color=blanco)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borde

    for offset, metodo in enumerate(metodos, start=1):
        row = start_day + offset
        metodo_ref = f"$F${row}"
        ws.cell(row=row, column=start_method_col, value=metodo)
        if metodo == "Sin pagos":
            ws.cell(row=row, column=start_method_col + 1, value=0)
        else:
            ws.cell(row=row, column=start_method_col + 1, value=_excel_formula_pagos_metodo(metodo_ref))
        for col_idx in range(start_method_col, start_method_col + 2):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = borde
            if col_idx == start_method_col + 1:
                cell.number_format = '$#,##0.00;[Red]-$#,##0.00'

    # Estilo ligero para las tablas auxiliares.
    for row in range(start_day + 1, start_day + max(len(dias), len(metodos)) + 1):
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            if row % 2 == 0 and cell.value not in (None, ""):
                cell.fill = PatternFill("solid", fgColor=gris)

    for col_letter, width in {"A": 14, "B": 18, "C": 16, "D": 12, "F": 18, "G": 16}.items():
        ws.column_dimensions[col_letter].width = max(ws.column_dimensions[col_letter].width or 0, width)

    # Gráfico 1: usa la tabla auxiliar por día, por eso cambia con la cajita de filtros.
    bar = BarChart()
    bar.title = "Generado vs pagado por día"
    bar.y_axis.title = "USD"
    bar.x_axis.title = "Día"
    data = Reference(ws, min_col=2, max_col=3, min_row=start_day, max_row=start_day + len(dias))
    cats = Reference(ws, min_col=1, min_row=start_day + 1, max_row=start_day + len(dias))
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.height = 8
    bar.width = 18
    ws.add_chart(bar, "A23")

    # Gráfico 2: método de pago, también vinculado a los filtros del Dashboard.
    pie = PieChart()
    pie.title = "Caja válida por método"
    data = Reference(
        ws,
        min_col=start_method_col + 1,
        min_row=start_day,
        max_row=start_day + len(metodos),
    )
    labels = Reference(
        ws,
        min_col=start_method_col,
        min_row=start_day + 1,
        max_row=start_day + len(metodos),
    )
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.height = 8
    pie.width = 11
    ws.add_chart(pie, "G23")


def _excel_crear_vistas_filtradas(
    ws_sesiones,
    ws_pagos,
    sesiones_headers: List[str],
    pagos_headers: List[str],
    sesiones_rows_count: int,
    pagos_rows_count: int,
) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    azul = "1F4E78"
    blanco = "FFFFFF"
    azul_claro = "D9EAF7"

    def escribir_encabezados(ws, headers: List[str]) -> None:
        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=6, column=idx, value=header)
            cell.fill = PatternFill("solid", fgColor=azul)
            cell.font = Font(bold=True, color=blanco)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def aplicar_formatos(ws, headers: List[str], start_row: int, end_row: int) -> None:
        currency_words = ["monto", "total", "generado", "pagado", "pendiente", "ganancia", "caja", "ecuasanitas", "precio", "saldo", "tarifa", "sueldo", "debe", "retenido"]
        integer_words = ["sesiones", "duración", "dolor"]
        for idx, header in enumerate(headers, start=1):
            lower = header.lower()
            for row_idx in range(start_row, end_row + 1):
                cell = ws.cell(row=row_idx, column=idx)
                if "fecha" in lower:
                    cell.number_format = "dd/mm/yyyy"
                elif "porcentaje" in lower:
                    cell.number_format = "0%"
                elif any(word in lower for word in currency_words):
                    cell.number_format = '$#,##0.00;[Red]-$#,##0.00'
                elif any(word in lower for word in integer_words):
                    cell.number_format = "0"

    # Vista de sesiones filtrada por la cajita del Dashboard.
    _excel_escribir_titulo(
        ws_sesiones,
        "CORPOFIT PRO — SESIONES FILTRADAS",
        "Esta hoja se actualiza con la cajita de filtros del Dashboard. Compatible sin FILTER()."
    )
    ws_sesiones["A4"] = "Cambia Clínica, Fisioterapeuta o Día en Dashboard. El detalle aparece aquí."
    ws_sesiones["A4"].fill = PatternFill("solid", fgColor=azul_claro)
    ws_sesiones["A4"].font = Font(bold=True, color="12355B")
    escribir_encabezados(ws_sesiones, sesiones_headers)

    total_ses_rows = max(sesiones_rows_count, 1)
    end_ses_row = total_ses_rows + 6
    for out_row in range(7, end_ses_row + 1):
        n_formula = f"ROWS($A$7:A{out_row})"
        for col_idx in range(1, len(sesiones_headers) + 1):
            col_letter = get_column_letter(col_idx)
            fallback = '"Sin resultados"' if out_row == 7 and col_idx == 1 else '""'
            ws_sesiones.cell(
                row=out_row,
                column=col_idx,
                value=f'=IFERROR(INDEX(Base_Sesiones!{col_letter}:{col_letter},MATCH({n_formula},Base_Sesiones!$W:$W,0)),{fallback})',
            )
    aplicar_formatos(ws_sesiones, sesiones_headers, 7, end_ses_row)
    ws_sesiones.freeze_panes = "A7"

    # Vista de pagos filtrada por la cajita del Dashboard.
    _excel_escribir_titulo(
        ws_pagos,
        "CORPOFIT PRO — PAGOS FILTRADOS",
        "Esta hoja se actualiza con la cajita de filtros del Dashboard. Compatible sin FILTER()."
    )
    ws_pagos["A4"] = "Cambia Clínica, Fisioterapeuta o Día en Dashboard. El detalle aparece aquí."
    ws_pagos["A4"].fill = PatternFill("solid", fgColor=azul_claro)
    ws_pagos["A4"].font = Font(bold=True, color="12355B")
    escribir_encabezados(ws_pagos, pagos_headers)

    total_pag_rows = max(pagos_rows_count, 1)
    end_pag_row = total_pag_rows + 6
    for out_row in range(7, end_pag_row + 1):
        n_formula = f"ROWS($A$7:A{out_row})"
        for col_idx in range(1, len(pagos_headers) + 1):
            col_letter = get_column_letter(col_idx)
            fallback = '"Sin resultados"' if out_row == 7 and col_idx == 1 else '""'
            ws_pagos.cell(
                row=out_row,
                column=col_idx,
                value=f'=IFERROR(INDEX(Base_Pagos!{col_letter}:{col_letter},MATCH({n_formula},Base_Pagos!$T:$T,0)),{fallback})',
            )
    aplicar_formatos(ws_pagos, pagos_headers, 7, end_pag_row)
    ws_pagos.freeze_panes = "A7"

def _crear_excel_reporte_corpofit(
    db: Session,
    current_user: Usuario,
    desde: date,
    hasta: date,
    terapeutaid: Optional[int] = None,
    consultorioid: Optional[int] = None,
    dia_semana: Optional[int] = None,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference, PieChart
    from openpyxl.styles import Alignment, Font, PatternFill

    general = reporte_terapias(
        desde=desde,
        hasta=hasta,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
        dia_semana=dia_semana,
        db=db,
        current_user=current_user,
    )
    fisios = reporte_fisioterapeutas_semanal(
        desde=desde,
        hasta=hasta,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
        dia_semana=dia_semana,
        db=db,
        current_user=current_user,
    )
    clinicas = reporte_clinicas_semanal(
        desde=desde,
        hasta=hasta,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
        dia_semana=dia_semana,
        db=db,
        current_user=current_user,
    )

    base_sesiones = _excel_base_sesiones(
        db=db,
        current_user=current_user,
        desde=desde,
        hasta=hasta,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
        dia_semana=dia_semana,
    )
    base_pagos = _excel_base_pagos(
        db=db,
        current_user=current_user,
        desde=desde,
        hasta=hasta,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
        dia_semana=dia_semana,
    )
    detalle_sueldos = _excel_detalle_sueldos_fisios(
        db=db,
        current_user=current_user,
        desde=desde,
        hasta=hasta,
        terapeutaid=terapeutaid,
        consultorioid=consultorioid,
        dia_semana=dia_semana,
    )

    wb = Workbook()
    # Forzar recálculo al abrir, porque el panel usa fórmulas vinculadas a tablas.
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass

    ws_dashboard = wb.active
    ws_dashboard.title = "Dashboard"
    ws_analisis = wb.create_sheet("Análisis")
    ws_sesiones = wb.create_sheet("Base_Sesiones")
    ws_pagos = wb.create_sheet("Base_Pagos")
    ws_sueldos = wb.create_sheet("Sueldos_Fisios")

    subtitulo = (
        f"Periodo semanal completo: {desde.strftime('%d/%m/%Y')} - {hasta.strftime('%d/%m/%Y')} | "
        f"Generado: {now_ecuador().strftime('%d/%m/%Y %H:%M')}"
    )

    # Hoja Análisis
    _excel_escribir_titulo(ws_analisis, "CORPOFIT PRO — ANÁLISIS", subtitulo)
    current_row = 4
    ws_analisis.cell(row=current_row, column=1, value="▶ Resumen por día")
    current_row += 1
    daily_headers = [
        "Fecha",
        "Día",
        "Sesiones",
        "Generado terapias",
        "Pagado caja",
        "Ecuasanitas",
        "Pendiente pacientes",
        "Sueldo fisio total",
        "Sueldo fisio cobrado",
        "Sueldo fisio retenido",
    ]
    sueldo_dia_map: Dict[date, Dict[str, float]] = defaultdict(lambda: {
        "pendiente_pacientes": 0.0,
        "sueldo_total": 0.0,
        "sueldo_cobrado": 0.0,
        "sueldo_retenido": 0.0,
    })
    for row in detalle_sueldos:
        fecha_row = row[0]
        if isinstance(fecha_row, date):
            sueldo_dia_map[fecha_row]["pendiente_pacientes"] += float(row[11] or 0.0)
            sueldo_dia_map[fecha_row]["sueldo_total"] += float(row[20] or 0.0)
            sueldo_dia_map[fecha_row]["sueldo_cobrado"] += float(row[21] or 0.0)
            sueldo_dia_map[fecha_row]["sueldo_retenido"] += float(row[22] or 0.0)

    daily_rows = []
    for item in general.sesiones_por_dia:
        sueldo_item = sueldo_dia_map.get(item.fecha, {})
        daily_rows.append(
            [
                item.fecha,
                item.dia,
                item.sesiones,
                item.total_generado,
                item.pagos_verificados,
                item.cubierto_ecuasanitas,
                round(float(sueldo_item.get("pendiente_pacientes", 0.0)), 2),
                round(float(sueldo_item.get("sueldo_total", 0.0)), 2),
                round(float(sueldo_item.get("sueldo_cobrado", 0.0)), 2),
                round(float(sueldo_item.get("sueldo_retenido", 0.0)), 2),
            ]
        )
    daily_table_start = current_row
    current_row = _excel_agregar_tabla(ws_analisis, current_row, 1, daily_headers, daily_rows, "tblAnalisisDias", crear_tabla=False)

    ws_analisis.cell(row=current_row, column=1, value="▶ Resumen por método de pago")
    current_row += 1
    metodo_headers = ["Método", "Total"]
    metodo_rows = [[m.metodo, m.total] for m in general.por_metodo_pago]
    metodo_table_start = current_row
    current_row = _excel_agregar_tabla(ws_analisis, current_row, 1, metodo_headers, metodo_rows, "tblAnalisisMetodos", crear_tabla=False)

    ws_analisis.cell(row=current_row, column=1, value="▶ Resumen por fisioterapeuta")
    current_row += 1
    fisio_headers = [
        "Terapeuta",
        "Clínica",
        "Sesiones",
        "Generado terapias",
        "Pagado pacientes",
        "Pendiente pacientes",
        "Ecuasanitas",
        "Gimnasio pagado",
        "Ganancia fisio total",
        "Ganancia fisio cobrada",
        "Ganancia fisio pendiente",
    ]
    fisio_rows = [
        [
            item.terapeuta,
            item.consultorio,
            item.sesiones_realizadas,
            item.total_generado,
            item.total_pagado_pacientes,
            item.total_pendiente_pacientes,
            item.total_ecuasanitas,
            item.total_gimnasio_pagado,
            item.ganancia_fisio_total,
            item.ganancia_fisio_cobrada,
            item.ganancia_fisio_pendiente,
        ]
        for item in fisios
    ]
    current_row = _excel_agregar_tabla(ws_analisis, current_row, 1, fisio_headers, fisio_rows, "tblAnalisisFisios", crear_tabla=False)

    ws_analisis.cell(row=current_row, column=1, value="▶ Resumen por clínica")
    current_row += 1
    clinica_headers = [
        "Clínica",
        "Sesiones",
        "Generado terapias",
        "Pagado pacientes",
        "Pendiente pacientes",
        "Ecuasanitas",
        "Gimnasio pagado",
        "Ganancia fisios total",
        "Ganancia clínica total",
        "Ganancia clínica cobrada",
        "Ganancia clínica pendiente",
    ]
    clinica_rows = [
        [
            item.consultorio,
            item.sesiones_realizadas,
            item.total_generado,
            item.total_pagado_pacientes,
            item.total_pendiente_pacientes,
            item.total_ecuasanitas,
            item.total_gimnasio_pagado,
            item.ganancia_fisios_total,
            item.ganancia_clinica_total,
            item.ganancia_clinica_cobrada,
            item.ganancia_clinica_pendiente,
        ]
        for item in clinicas
    ]
    current_row = _excel_agregar_tabla(ws_analisis, current_row, 1, clinica_headers, clinica_rows, "tblAnalisisClinicas", crear_tabla=False)

    # Base de sesiones
    sesiones_headers = [
        "ID Sesión",
        "Fecha",
        "Día",
        "Paciente",
        "Cédula",
        "Fisioterapeuta",
        "Clínica / Consultorio",
        "Clínica paciente",
        "Tratamiento",
        "Tratamiento ID",
        "Hora ingreso",
        "Hora salida",
        "Duración min",
        "Dolor entrada",
        "Dolor salida",
        "Precio sesión",
        "Ecuasanitas",
        "Pago aplicado",
        "Debe paciente",
        "Estado pago sesión",
        "Tarifa fisio",
        "Motivo tarifa",
        "Especial $5",
        "Motivo especial",
        "Ganancia fisio",
        "Sueldo fisio cobrado",
        "Sueldo fisio retenido",
        "Ganancia clínica",
    ]
    _excel_agregar_tabla(ws_sesiones, 1, 1, sesiones_headers, base_sesiones, "tblBaseSesiones")
    ws_sesiones.freeze_panes = "A2"

    # Base de pagos
    pagos_headers = [
        "ID Pago",
        "Fecha Ecuador",
        "Día",
        "Paciente",
        "Cédula",
        "Fisioterapeuta / Responsable",
        "Clínica / Consultorio",
        "Tipo",
        "Método",
        "Estado",
        "Monto",
        "Pago previo",
        "Recuperación cartera",
        "Anulado",
        "Referencia",
        "Observación",
        "Cobrador ID",
        "Caja válida",
    ]
    _excel_agregar_tabla(ws_pagos, 1, 1, pagos_headers, base_pagos, "tblBasePagos")
    ws_pagos.freeze_panes = "A2"

    # Detalle auditable de sueldos por sesión y terapeuta real que atendió.
    sueldos_headers = [
        "Fecha",
        "Día",
        "ID Sesión",
        "Fisioterapeuta que atendió",
        "Clínica atención",
        "Paciente",
        "Cédula",
        "Tratamiento",
        "Tratamiento ID",
        "Precio sesión",
        "Pago aplicado a sesión",
        "Debe paciente",
        "Estado pago sesión",
        "Atenciones fisio ese día",
        "Bono +15",
        "Fin de semana",
        "Tarifa fisio",
        "Motivo tarifa",
        "Especial $5",
        "Motivo especial",
        "Sueldo total sesión",
        "Sueldo cobrado",
        "Sueldo retenido",
        "Clínica paciente",
    ]
    _excel_agregar_tabla(ws_sueldos, 1, 1, sueldos_headers, detalle_sueldos, "tblSueldosFisios")
    ws_sueldos.freeze_panes = "A2"

    # Panel central de filtros: una sola cajita controla KPIs y gráficas.
    rangos_filtros = _excel_crear_listas_filtros(wb, base_sesiones, base_pagos)
    _excel_escribir_titulo(ws_dashboard, "CORPOFIT PRO — REPORTE EXCEL", subtitulo)
    _excel_configurar_panel_filtros_dashboard(ws_dashboard, rangos_filtros, desde, hasta)

    # Gráficas vinculadas a los filtros del Dashboard.
    # Se usan tablas auxiliares con SUMPRODUCT; no se usa FILTER() para evitar
    # archivos que Excel/WPS intenten reparar al abrir.
    _excel_crear_graficas_dashboard(ws_dashboard, base_pagos)

    # Formatos y estilo visual.
    for ws in [ws_dashboard, ws_analisis, ws_sesiones, ws_pagos, ws_sueldos]:
        ws.freeze_panes = ws.freeze_panes or "A4"
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=False)

    _excel_aplicar_estilo_workbook(wb)

    for ws in [ws_analisis, ws_sesiones, ws_pagos, ws_sueldos]:
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                header = str(ws.cell(row=1, column=cell.column).value or "").lower()
                # En Análisis los encabezados no siempre están en fila 1; reforzamos por tipo de valor.
                if isinstance(cell.value, (int, float)) and any(
                    word in str(cell.offset(row=-(cell.row - 1)).value or "").lower()
                    for word in ["monto", "total", "generado", "pagado", "pendiente", "ganancia", "caja", "precio"]
                ):
                    cell.number_format = '$#,##0.00;[Red]-$#,##0.00'
                if isinstance(cell.value, date):
                    cell.number_format = "dd/mm/yyyy"

    # Ajuste final de formatos por columnas conocidas.
    for ws in [ws_dashboard, ws_analisis, ws_sesiones, ws_pagos, ws_sueldos]:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, date):
                    cell.number_format = "dd/mm/yyyy"
                if isinstance(cell.value, (int, float)) and cell.column >= 4:
                    # La mayoría de valores monetarios están a partir de la columna D.
                    header_values = [str(ws.cell(row=r, column=cell.column).value or "").lower() for r in range(1, min(cell.row, 6) + 1)]
                    if any(
                        any(word in header for word in ["monto", "total", "generado", "pagado", "pendiente", "ganancia", "caja", "precio", "ecuasanitas", "tarifa", "sueldo", "debe", "retenido"])
                        for header in header_values
                    ):
                        cell.number_format = '$#,##0.00;[Red]-$#,##0.00'

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


@router.get("/exportar-excel")
def exportar_excel_reportes(
    desde: Optional[date] = Query(None),
    hasta: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    desde, hasta = _default_range(desde, hasta)

    # El Excel siempre contiene toda la semana permitida para el rol.
    # Clínica/Consultorio, Fisioterapeuta y Día se controlan desde una sola
    # cajita de filtros en la hoja Dashboard.
    contenido = _crear_excel_reporte_corpofit(
        db=db,
        current_user=current_user,
        desde=desde,
        hasta=hasta,
        terapeutaid=None,
        consultorioid=None,
        dia_semana=None,
    )

    filename = f"corpofit_reporte_{desde.isoformat()}_{hasta.isoformat()}.xlsx"
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "Cache-Control": "no-store",
    }

    return StreamingResponse(
        BytesIO(contenido),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
